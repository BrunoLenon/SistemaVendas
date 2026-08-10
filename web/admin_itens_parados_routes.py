# -*- coding: utf-8 -*-
"""Administração enxuta de Itens Parados.

Mantém:
* cadastro/ativação dos produtos por EMP com validade;
* regra global ou específica por EMP para base e valor de cada ponto;
* importação do modelo CODIGO, DESCRICAO e LOJA_ATIVA com validade definida na página;
* importação das vendas com cálculo de pontos inteiros e bônus por funcionário.

O cálculo acontece somente na importação; as páginas operacionais leem o
snapshot pronto e não consultam a base geral de vendas.
"""

from __future__ import annotations

import calendar
import json
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Callable, Type

from flask import current_app, render_template, request, send_file, session

from sqlalchemy import or_

from db import (
    ItensParadosPontosConfig,
    ItensParadosVendaUsuario,
    ensure_itens_parados_snapshot_schema,
)
from itens_parados_snapshot import (
    balance_row_is_seller,
    batch_warnings,
    latest_batch,
    load_point_rules,
    norm_emp,
    seller_identity_sets,
)
from sv_utils import business_today, emp_sort_key


MAX_ITEMS_IMPORT = 20000


def _clean_text(value) -> str:
    if value is None:
        return ""
    raw = str(value).replace("\xa0", " ").strip()
    if raw.lower() in {"nan", "none", "null"}:
        return ""
    return re.sub(r"\s+", " ", raw)


def _clean_code(value) -> str:
    raw = _clean_text(value)
    if re.fullmatch(r"\d+\.0+", raw):
        return raw.split(".", 1)[0]
    return raw


def _positive_decimal(value, label: str, *, allow_zero: bool = False) -> Decimal:
    raw = _clean_text(value).replace("R$", "").replace(" ", "")
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        result = Decimal(raw)
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"{label} inválido.") from exc
    if not result.is_finite():
        raise ValueError(f"{label} inválido.")
    if result < 0 or (result == 0 and not allow_zero):
        operator = "maior ou igual a zero" if allow_zero else "maior que zero"
        raise ValueError(f"{label} deve ser {operator}.")
    return result.quantize(Decimal("0.0001"))


def _required_date(value, label: str) -> date:
    """Converte datas do formulário/Excel e exige uma data válida."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = _clean_text(value)
    if not raw:
        raise ValueError(f"{label} é obrigatória.")

    # Pandas pode serializar timestamps como 2026-07-01 00:00:00.
    candidate = raw[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(candidate, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{label} inválida. Use DD/MM/AAAA.")


def _validity_period(start_value, end_value) -> tuple[date, date]:
    start = _required_date(start_value, "Data inicial")
    end = _required_date(end_value, "Data final")
    if end < start:
        raise ValueError("A data final não pode ser anterior à data inicial.")
    return start, end


def _norm_header(value) -> str:
    raw = unicodedata.normalize("NFKD", str(value or ""))
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    return re.sub(r"[^A-Za-z0-9]", "", raw).upper()


def _read_active_items_file(
    file_storage,
    *,
    default_start: date,
    default_end: date,
) -> tuple[list[dict], list[str]]:
    filename = (getattr(file_storage, "filename", "") or "").lower().strip()
    if not filename:
        raise ValueError("Selecione a planilha de itens parados.")
    if not filename.endswith((".xlsx", ".xlsm", ".csv")):
        raise ValueError("Formato inválido. Envie .xlsx, .xlsm ou .csv.")

    try:
        import pandas as pd
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("Pandas não está disponível no servidor.") from exc

    file_storage.stream.seek(0)
    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(
                file_storage.stream,
                dtype=str,
                keep_default_na=False,
                sep=None,
                engine="python",
            )
        else:
            df = pd.read_excel(file_storage.stream, dtype=str, keep_default_na=False)
    except Exception as exc:
        raise ValueError(f"Não foi possível ler a planilha: {exc}") from exc

    if df is None or df.empty:
        raise ValueError("A planilha está vazia.")
    if len(df.index) > MAX_ITEMS_IMPORT:
        raise ValueError(
            f"A planilha possui {len(df.index)} linhas. Limite: {MAX_ITEMS_IMPORT}."
        )

    aliases = {
        "codigo": {"CODIGO", "MESTRE", "CODIGOPRODUTO", "CODPRODUTO"},
        "descricao": {"DESCRICAO", "DESCRICAOPRODUTO", "PRODUTO", "NOME"},
        "emp": {"LOJAATIVA", "EMP", "EMPRESA", "LOJA", "LOJAACA0", "LOJAACAO"},
        "data_inicio": {"DATAINICIO", "INICIO", "VALIDADEINICIO", "INICIOVALIDADE"},
        "data_fim": {"DATAFIM", "FIM", "VALIDADEFIM", "FIMVALIDADE"},
    }
    mapping: dict[str, str] = {}
    for column in df.columns:
        normalized = _norm_header(column)
        for field, names in aliases.items():
            if field not in mapping and normalized in names:
                mapping[field] = column
                break

    required_fields = ("codigo", "descricao", "emp")
    missing = [field for field in required_fields if field not in mapping]
    if missing:
        labels = {
            "codigo": "CODIGO",
            "descricao": "DESCRICAO",
            "emp": "LOJA_ATIVA",
        }
        raise ValueError(
            "O modelo precisa conter CODIGO, DESCRICAO e LOJA_ATIVA. "
            "Colunas ausentes: "
            + ", ".join(labels[field] for field in missing)
            + "."
        )

    records: list[dict] = []
    warnings: list[str] = []
    seen: set[tuple[str, str]] = set()
    for index, row in df.iterrows():
        line = int(index) + 2
        codigo = _clean_code(row.get(mapping["codigo"]))
        descricao = _clean_text(row.get(mapping["descricao"]))
        emp = norm_emp(row.get(mapping["emp"]))
        raw_start = row.get(mapping["data_inicio"]) if "data_inicio" in mapping else None
        raw_end = row.get(mapping["data_fim"]) if "data_fim" in mapping else None
        if not codigo and not descricao and not emp and not _clean_text(raw_start) and not _clean_text(raw_end):
            continue
        if not codigo or not emp:
            warnings.append(
                f"Linha {line} ignorada: CODIGO e LOJA_ATIVA são obrigatórios."
            )
            continue
        if _clean_text(raw_start) or _clean_text(raw_end):
            try:
                data_inicio, data_fim = _validity_period(raw_start, raw_end)
            except ValueError as exc:
                warnings.append(f"Linha {line} ignorada: {exc}")
                continue
        else:
            data_inicio, data_fim = default_start, default_end
        key = (emp, codigo.upper())
        if key in seen:
            warnings.append(
                f"Linha {line} ignorada: código {codigo} duplicado para a EMP {emp}."
            )
            continue
        seen.add(key)
        records.append(
            {
                "linha": line,
                "codigo": codigo,
                "descricao": descricao or None,
                "emp": emp,
                "data_inicio": data_inicio,
                "data_fim": data_fim,
            }
        )
    if not records:
        raise ValueError("Nenhuma linha válida foi encontrada no modelo.")
    return records, warnings


def _create_model_workbook() -> BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl não está disponível para gerar o modelo.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Itens Parados"
    sheet.append(["CODIGO", "DESCRICAO", "LOJA_ATIVA"])
    sheet.append(["345269", "CARENAGEM LAT FAROL TITAN 160", "101"])
    sheet.append(["373126", "FILTRO AR ESPORTIVO TITAN/FAN", "501"])

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 58
    sheet.column_dimensions["C"].width = 18
    sheet.freeze_panes = "A2"
    table = Table(displayName="ModeloItensParados", ref="A1:C3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    info = workbook.create_sheet("Instrucoes")
    info.append(["Campo", "Como preencher"])
    info.append(["CODIGO", "Obrigatório. Código/MESTRE do produto da ação."])
    info.append(["DESCRICAO", "Descrição do produto para conferência."])
    info.append(
        [
            "LOJA_ATIVA",
            "Obrigatório. Informe uma EMP por linha. Para o mesmo produto em várias lojas, repita o código em uma linha para cada EMP.",
        ]
    )
    info.append(["VALIDADE", "Defina a data inicial e final na página no momento da importação. A mesma validade será aplicada a todas as linhas do arquivo."])
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 90
    for row in info.iter_rows(min_row=1, max_row=5, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font

    bio = BytesIO()
    workbook.save(bio)
    bio.seek(0)
    return bio


def register_admin_itens_parados_routes(
    app,
    *,
    SessionLocal,
    ItemParado: Type,
    login_required_fn: Callable[[], object | None],
    admin_required_fn: Callable[[], object | None],
    usuario_logado_fn: Callable[[], object],
):
    def admin_itens_parados():
        red = login_required_fn()
        if red:
            return red
        red = admin_required_fn()
        if red:
            return red

        ensure_itens_parados_snapshot_schema()
        erro = None
        ok = None
        avisos_acao: list[str] = []

        with SessionLocal() as db:
            try:
                action = (request.form.get("acao") or "").strip().lower()
                if request.method == "POST" and action:
                    if action == "salvar_regra":
                        regra_emp = norm_emp(request.form.get("regra_emp")) or None
                        base_reais = _positive_decimal(
                            request.form.get("base_reais"),
                            "Valor necessário para 1 ponto",
                        )
                        valor_por_ponto = _positive_decimal(
                            request.form.get("valor_por_ponto"),
                            "Valor de cada ponto",
                            allow_zero=True,
                        )
                        if regra_emp:
                            configs = (
                                db.query(ItensParadosPontosConfig)
                                .filter(ItensParadosPontosConfig.emp == regra_emp)
                                .order_by(ItensParadosPontosConfig.id.desc())
                                .all()
                            )
                        else:
                            configs = (
                                db.query(ItensParadosPontosConfig)
                                .filter(ItensParadosPontosConfig.emp.is_(None))
                                .order_by(ItensParadosPontosConfig.id.desc())
                                .all()
                            )
                        config = configs[0] if configs else None
                        for duplicate in configs[1:]:
                            duplicate.ativo = False
                            duplicate.atualizado_em = datetime.utcnow()
                        if config:
                            config.base_reais = float(base_reais)
                            config.valor_por_ponto = float(valor_por_ponto)
                            config.ativo = True
                            config.atualizado_em = datetime.utcnow()
                        else:
                            db.add(
                                ItensParadosPontosConfig(
                                    emp=regra_emp,
                                    base_reais=float(base_reais),
                                    valor_por_ponto=float(valor_por_ponto),
                                    ativo=True,
                                    criado_em=datetime.utcnow(),
                                    atualizado_em=datetime.utcnow(),
                                )
                            )
                        db.commit()
                        destino = f"EMP {regra_emp}" if regra_emp else "regra global"
                        ok = f"Regra de pontos salva para {destino}."

                    elif action == "criar_item":
                        emp = norm_emp(request.form.get("emp"))
                        codigo = _clean_code(request.form.get("codigo"))
                        descricao = _clean_text(request.form.get("descricao")) or None
                        data_inicio, data_fim = _validity_period(
                            request.form.get("data_inicio"),
                            request.form.get("data_fim"),
                        )
                        if not emp or not codigo:
                            raise ValueError("Informe EMP e código.")
                        item = (
                            db.query(ItemParado)
                            .filter(ItemParado.emp == emp, ItemParado.codigo == codigo)
                            .order_by(ItemParado.id.desc())
                            .first()
                        )
                        if item:
                            item.descricao = descricao or item.descricao
                            item.data_inicio = data_inicio
                            item.data_fim = data_fim
                            item.ativo = True
                            item.atualizado_em = datetime.utcnow()
                            ok = "Item atualizado e ativado."
                        else:
                            db.add(
                                ItemParado(
                                    emp=emp,
                                    codigo=codigo,
                                    descricao=descricao,
                                    quantidade=None,
                                    recompensa_pct=0,
                                    modo="PONTOS",
                                    multiplicador_pontos=1,
                                    data_inicio=data_inicio,
                                    data_fim=data_fim,
                                    ativo=True,
                                    criado_em=datetime.utcnow(),
                                    atualizado_em=datetime.utcnow(),
                                )
                            )
                            ok = "Item cadastrado."
                        db.commit()

                    elif action == "atualizar_validade":
                        item_id = int(request.form.get("item_id") or 0)
                        item = db.get(ItemParado, item_id)
                        if not item:
                            raise ValueError("Item não encontrado.")
                        data_inicio, data_fim = _validity_period(
                            request.form.get("data_inicio"),
                            request.form.get("data_fim"),
                        )
                        item.data_inicio = data_inicio
                        item.data_fim = data_fim
                        item.ativo = True
                        item.atualizado_em = datetime.utcnow()
                        db.commit()
                        ok = "Validade do item atualizada e item ativado."

                    elif action == "alternar_item":
                        item_id = int(request.form.get("item_id") or 0)
                        item = db.get(ItemParado, item_id)
                        if not item:
                            raise ValueError("Item não encontrado.")
                        item.ativo = not bool(item.ativo)
                        item.atualizado_em = datetime.utcnow()
                        db.commit()
                        ok = "Status do item atualizado."

                    elif action == "excluir_item":
                        item_id = int(request.form.get("item_id") or 0)
                        item = db.get(ItemParado, item_id)
                        if not item:
                            raise ValueError("Item não encontrado.")
                        db.delete(item)
                        db.commit()
                        ok = "Item removido."

                    elif action == "importar_itens":
                        file_storage = request.files.get("arquivo_itens")
                        validade_inicio, validade_fim = _validity_period(
                            request.form.get("validade_inicio"),
                            request.form.get("validade_fim"),
                        )
                        records, avisos_acao = _read_active_items_file(
                            file_storage,
                            default_start=validade_inicio,
                            default_end=validade_fim,
                        )
                        created = 0
                        updated = 0
                        now = datetime.utcnow()
                        for record in records:
                            item = (
                                db.query(ItemParado)
                                .filter(
                                    ItemParado.emp == record["emp"],
                                    ItemParado.codigo == record["codigo"],
                                )
                                .order_by(ItemParado.id.desc())
                                .first()
                            )
                            if item:
                                item.descricao = record["descricao"] or item.descricao
                                item.data_inicio = record["data_inicio"]
                                item.data_fim = record["data_fim"]
                                item.ativo = True
                                item.atualizado_em = now
                                updated += 1
                            else:
                                db.add(
                                    ItemParado(
                                        emp=record["emp"],
                                        codigo=record["codigo"],
                                        descricao=record["descricao"],
                                        quantidade=None,
                                        recompensa_pct=0,
                                        modo="PONTOS",
                                        multiplicador_pontos=1,
                                        data_inicio=record["data_inicio"],
                                        data_fim=record["data_fim"],
                                        ativo=True,
                                        criado_em=now,
                                        atualizado_em=now,
                                    )
                                )
                                created += 1
                        db.commit()
                        ok = (
                            f"Importação concluída: {created} item(ns) criado(s) e "
                            f"{updated} atualizado(s), com validade de "
                            f"{validade_inicio.strftime('%d/%m/%Y')} a "
                            f"{validade_fim.strftime('%d/%m/%Y')}."
                        )
                    else:
                        raise ValueError("Ação inválida.")
            except Exception as exc:
                db.rollback()
                erro = str(exc)
                current_app.logger.exception("Erro no admin de itens parados")

            filter_emp = norm_emp(request.args.get("f_emp"))
            filter_code = _clean_code(request.args.get("f_codigo"))
            filter_status = (request.args.get("f_status") or "ativos").strip().lower()
            filter_validity = (request.args.get("f_validade") or "todos").strip().lower()
            page = max(int(request.args.get("page") or 1), 1)
            limit = 150
            query = db.query(ItemParado)
            if filter_emp:
                query = query.filter(ItemParado.emp == filter_emp)
            if filter_code:
                query = query.filter(ItemParado.codigo.ilike(f"%{filter_code}%"))
            if filter_status == "ativos":
                query = query.filter(ItemParado.ativo.is_(True))
            elif filter_status == "inativos":
                query = query.filter(ItemParado.ativo.is_(False))

            today_filter = business_today()
            if filter_validity == "vigentes":
                query = query.filter(
                    ItemParado.data_inicio.isnot(None),
                    ItemParado.data_fim.isnot(None),
                    ItemParado.data_inicio <= today_filter,
                    ItemParado.data_fim >= today_filter,
                )
            elif filter_validity == "agendados":
                query = query.filter(ItemParado.data_inicio > today_filter)
            elif filter_validity == "encerrados":
                query = query.filter(ItemParado.data_fim < today_filter)
            elif filter_validity == "sem_validade":
                query = query.filter(
                    or_(ItemParado.data_inicio.is_(None), ItemParado.data_fim.is_(None))
                )
            total_items = query.count()
            items = (
                query.order_by(ItemParado.emp.asc(), ItemParado.codigo.asc())
                .offset((page - 1) * limit)
                .limit(limit)
                .all()
            )
            total_pages = max((total_items + limit - 1) // limit, 1)
            emp_options = sorted(
                {
                    str(row[0]).strip()
                    for row in db.query(ItemParado.emp).distinct().all()
                    if row and row[0]
                },
                key=emp_sort_key,
            )

            point_rules = load_point_rules(db)
            point_configs = []
            global_base, global_value = point_rules["global"]
            point_configs.append(
                {
                    "emp": "",
                    "label": "Global",
                    "base_reais": global_base,
                    "valor_por_ponto": global_value,
                }
            )
            for emp_code, values in sorted(
                point_rules["emps"].items(), key=lambda item: emp_sort_key(item[0])
            ):
                point_configs.append(
                    {
                        "emp": emp_code,
                        "label": f"EMP {emp_code}",
                        "base_reais": values[0],
                        "valor_por_ponto": values[1],
                    }
                )

            today = business_today()
            ano_vendas = int(request.args.get("ano_vendas") or today.year)
            mes_vendas = int(request.args.get("mes_vendas") or today.month)
            validade_padrao_inicio = date(ano_vendas, mes_vendas, 1)
            validade_padrao_fim = date(
                ano_vendas,
                mes_vendas,
                calendar.monthrange(ano_vendas, mes_vendas)[1],
            )
            sales_batch = latest_batch(db, ano_vendas, mes_vendas)
            sales_warnings = batch_warnings(sales_batch)
            seller_ids, seller_names = seller_identity_sets(db)
            seller_sales_rows = [
                row
                for row in (
                    db.query(ItensParadosVendaUsuario)
                    .filter(
                        ItensParadosVendaUsuario.ano == ano_vendas,
                        ItensParadosVendaUsuario.mes == mes_vendas,
                    )
                    .order_by(
                        ItensParadosVendaUsuario.emp.asc(),
                        ItensParadosVendaUsuario.bonus_total.desc(),
                    )
                    .all()
                )
                if balance_row_is_seller(row, seller_ids, seller_names)
            ]
            sales_rows = seller_sales_rows[:200]
            sales_summary = {
                "valor_total": sum(
                    (Decimal(str(row.valor_total or 0)) for row in seller_sales_rows),
                    Decimal("0"),
                ),
                "pontos_total": sum(int(row.pontos or 0) for row in seller_sales_rows),
                "bonus_total": sum(
                    (Decimal(str(row.bonus_total or 0)) for row in seller_sales_rows),
                    Decimal("0"),
                ),
                "registros_usuarios": len(seller_sales_rows),
                "linhas_importadas": sum(
                    int(row.qtd_linhas or 0) for row in seller_sales_rows
                ),
            }

        return render_template(
            "admin_itens_parados.html",
            usuario=usuario_logado_fn(),
            role=(session.get("role") or ""),
            itens=items,
            erro=erro,
            ok=ok,
            avisos_acao=avisos_acao,
            filtro_emp=filter_emp,
            filtro_codigo=filter_code,
            filtro_status=filter_status,
            filtro_validade=filter_validity,
            hoje=business_today(),
            validade_padrao_inicio=validade_padrao_inicio,
            validade_padrao_fim=validade_padrao_fim,
            pagina=page,
            total_paginas=total_pages,
            total_itens=total_items,
            emps_options=emp_options,
            point_configs=point_configs,
            ano_vendas=ano_vendas,
            mes_vendas=mes_vendas,
            sales_batch=sales_batch,
            sales_warnings=sales_warnings,
            sales_rows=sales_rows,
            sales_summary=sales_summary,
        )

    def admin_itens_parados_modelo():
        red = login_required_fn()
        if red:
            return red
        red = admin_required_fn()
        if red:
            return red
        return send_file(
            _create_model_workbook(),
            as_attachment=True,
            download_name="modelo_importacao_itens_parados.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    app.add_url_rule(
        "/admin/itens_parados/modelo",
        endpoint="admin_itens_parados_modelo",
        view_func=admin_itens_parados_modelo,
        methods=["GET"],
    )
    app.add_url_rule(
        "/admin/itens_parados",
        endpoint="admin_itens_parados",
        view_func=admin_itens_parados,
        methods=["GET", "POST"],
    )
