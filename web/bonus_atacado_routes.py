# -*- coding: utf-8 -*-
"""Importação e consulta do snapshot mensal de Bônus Atacado.

A planilha permanece responsável pelos cálculos. O sistema importa somente os
valores prontos da aba ``PremiacaoFinal`` e aplica a hierarquia cadastrada:

* vendedor, mecânico e demais perfis operacionais: somente os próprios dados;
* supervisor e gerente: todos os usuários das EMPs vinculadas, além dos valores
  de Loja Anterior e Loja Atual dessas empresas;
* administrador: todos os dados, filtros, importação e validações.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from flask import current_app, flash, redirect, render_template, request, session, url_for
from sqlalchemy import or_
from werkzeug.utils import secure_filename

from auth_helpers import _allowed_emps, _login_required, _role, _usuario_logado
from db import (
    BonusAtacadoImportacaoLote,
    BonusAtacadoUsuario,
    SessionLocal,
    Usuario,
    UsuarioEmp,
    ensure_bonus_atacado_schema,
    ensure_itens_parados_snapshot_schema,
    ensure_bonus_outros_valores_schema,
)
from security_utils import audit, normalize_role
from sv_utils import emp_sort_key
from itens_parados_snapshot import attach_saldos_to_bonus_rows
from bonus_outros_valores import attach_outros_valores_to_bonus_rows, bonus_row_options


MAX_UPLOAD_BYTES = 15 * 1024 * 1024
VALID_EXTENSIONS = {".xlsx", ".xlsm"}
MANAGER_ROLES = {"admin", "supervisor", "gerente"}

HEADER_ALIASES: dict[str, set[str]] = {
    "emp": {"EMP", "EMPRESA"},
    "funcao_planilha": {"FUNCAO", "PERFIL", "CARGO"},
    "usuario_nome": {"FUNCIONARIO", "USUARIO", "VENDEDOR"},
    "total_produtos": {"TOTAL", "TOTALPREMIOPRODUTOS", "PREMIOPRODUTOS"},
    "venda_anterior": {"VENDAANTERIOR", "VENDASANTERIOR"},
    "venda_atual": {"VENDAATUAL", "VENDASATUAL"},
    "importado": {"IMPORTADO", "VENDAIMPORTADO", "VENDIDOIMPORTADO"},
    "percentual_importado": {"PERCENTUALIMPORTADO", "IMPORTADOPERCENTUAL"},
    "loja_anterior": {"LOJAANTERIOR", "VENDAANTERIORLOJA"},
    "loja_atual": {"LOJAATUAL", "VENDAATUALLOJA"},
    "falta_valor_vendedor": {
        "FALTAVALORVENDEDOR",
        "VALORFALTAVENDEDOR",
        "FALTAVENDEDOR",
    },
    "mix": {"MIX", "MIXPRODUTOS", "MIXDEPRODUTOS"},
}
REQUIRED_FIELDS = {
    "emp",
    "usuario_nome",
    "total_produtos",
    "venda_anterior",
    "venda_atual",
    "importado",
    "percentual_importado",
    "loja_anterior",
    "loja_atual",
    "falta_valor_vendedor",
    "mix",
}
NUMERIC_FIELDS = (
    "total_produtos",
    "venda_anterior",
    "venda_atual",
    "importado",
    "percentual_importado",
    "loja_anterior",
    "loja_atual",
    "falta_valor_vendedor",
)
MONEY_FIELDS = set(NUMERIC_FIELDS) - {"percentual_importado"}
INTEGER_FIELDS = {"mix"}


def register_bonus_atacado_routes(app) -> None:
    app.add_url_rule(
        "/bonus-atacado",
        endpoint="bonus_atacado",
        view_func=bonus_atacado,
        methods=["GET"],
    )
    app.add_url_rule(
        "/admin/bonus-atacado/importar",
        endpoint="admin_bonus_atacado_importar",
        view_func=admin_bonus_atacado_importar,
        methods=["POST"],
    )


def _strip_accents(value: object) -> str:
    raw = str(value or "")
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch)
    )


def _norm_header(value: object) -> str:
    raw = _strip_accents(value).strip().upper().replace("%", " PERCENTUAL ")
    return re.sub(r"[^A-Z0-9]", "", raw)


def _norm_username(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def _canonical_username(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", _strip_accents(value).upper())


def _norm_emp(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).upper()
    if isinstance(value, (int, float, Decimal)):
        try:
            number = Decimal(str(value))
            if number == number.to_integral_value():
                return str(int(number))
        except Exception:
            pass
    raw = str(value).strip()
    if re.fullmatch(r"[+-]?\d+[.,]0+", raw):
        return raw.split(".", 1)[0].split(",", 1)[0]
    return raw.upper()


def _safe_period(value: object, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(str(value).strip())
    except Exception:
        return default
    return max(minimum, min(maximum, parsed))


def _decimal_from_cell(cell, *, is_percent: bool) -> Decimal | None:
    value = cell.value
    if getattr(cell, "data_type", None) == "e" or (
        isinstance(value, str) and value.strip().startswith("#")
    ):
        raise ValueError(f"erro do Excel ({value})")

    if value is None or (isinstance(value, str) and not value.strip()):
        return None

    had_percent_symbol = False
    if isinstance(value, str):
        raw = value.strip()
        had_percent_symbol = "%" in raw
        raw = raw.replace("R$", "").replace("%", "").replace(" ", "")
        if not raw:
            return None
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        elif raw.count(".") > 1:
            raw = raw.replace(".", "")
        value = raw

    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"valor numérico inválido ({value})") from exc
    if not result.is_finite():
        raise ValueError(f"valor numérico inválido ({value})")

    number_format = str(getattr(cell, "number_format", "") or "")
    if is_percent and not had_percent_symbol and "%" in number_format:
        # O Excel guarda 26,30% como 0,2630. Persistimos 26,30 para exibição.
        result *= Decimal("100")

    quantum = Decimal("0.000001") if is_percent else Decimal("0.0001")
    return result.quantize(quantum, rounding=ROUND_HALF_UP)


def _integer_from_cell(cell) -> int:
    value = _decimal_from_cell(cell, is_percent=False)
    if value is None:
        return 0
    integral = value.to_integral_value(rounding=ROUND_HALF_UP)
    if value != integral:
        raise ValueError(f"valor deve ser inteiro ({value})")
    if integral < 0:
        raise ValueError(f"valor não pode ser negativo ({value})")
    return int(integral)


def _find_sheet(workbook):
    for sheet in workbook.worksheets:
        if _norm_header(sheet.title) == "PREMIACAOFINAL":
            return sheet
    available = ", ".join(sheet.title for sheet in workbook.worksheets)
    raise ValueError(
        "A aba 'PremiacaoFinal' não foi encontrada. Abas disponíveis: " + available
    )


def _find_header(sheet) -> tuple[int, dict[str, int]]:
    best_row = 1
    best_mapping: dict[str, int] = {}

    for row_number, cells in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 30)), start=1
    ):
        mapping: dict[str, int] = {}
        for index, cell in enumerate(cells):
            normalized = _norm_header(cell.value)
            if not normalized:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if field not in mapping and normalized in aliases:
                    mapping[field] = index
                    break
        if len(mapping) > len(best_mapping):
            best_row, best_mapping = row_number, mapping
        if REQUIRED_FIELDS.issubset(mapping):
            return row_number, mapping

    missing = [
        field.replace("_", " ").title()
        for field in sorted(REQUIRED_FIELDS - set(best_mapping))
    ]
    raise ValueError(
        "O cabeçalho da aba PremiacaoFinal não corresponde ao modelo. "
        "Colunas não reconhecidas: " + ", ".join(missing)
    )


def _read_workbook(content: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "A biblioteca openpyxl não está disponível no servidor."
        ) from exc

    try:
        workbook = load_workbook(
            filename=BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ValueError("Não foi possível abrir o arquivo Excel enviado.") from exc

    try:
        sheet = _find_sheet(workbook)
        header_row, mapping = _find_header(sheet)
        max_col = max(mapping.values()) + 1

        records: list[dict[str, Any]] = []
        warnings: list[str] = []
        duplicate_keys: dict[tuple[str, str], int] = {}
        rows_read = 0
        rows_skipped = 0

        store_values: dict[str, tuple[Decimal | None, Decimal | None, int]] = {}

        for source_row, cells_tuple in enumerate(
            sheet.iter_rows(min_row=header_row + 1, max_col=max_col),
            start=header_row + 1,
        ):
            cells = list(cells_tuple)
            emp_raw = cells[mapping["emp"]].value
            user_raw = cells[mapping["usuario_nome"]].value
            if emp_raw in (None, "") and user_raw in (None, ""):
                continue

            rows_read += 1
            emp = _norm_emp(emp_raw)
            username = _norm_username(user_raw)
            if not emp or not username:
                rows_skipped += 1
                missing = "EMP" if not emp else "FUNCIONARIO"
                warnings.append(f"Linha {source_row} ignorada: {missing} não informado.")
                continue

            key = (emp, username)
            if key in duplicate_keys:
                first_row = duplicate_keys[key]
                raise ValueError(
                    f"Duplicidade na aba PremiacaoFinal: usuário {username}, EMP {emp}, "
                    f"linhas {first_row} e {source_row}."
                )
            duplicate_keys[key] = source_row

            record: dict[str, Any] = {
                "linha_origem": source_row,
                "usuario_nome": username,
                "funcao_planilha": None,
                "emp": emp,
            }
            function_index = mapping.get("funcao_planilha")
            if function_index is not None:
                function_raw = cells[function_index].value
                record["funcao_planilha"] = (
                    _strip_accents(function_raw).strip().upper() or None
                )

            for field in NUMERIC_FIELDS:
                try:
                    record[field] = _decimal_from_cell(
                        cells[mapping[field]],
                        is_percent=(field == "percentual_importado"),
                    )
                except ValueError as exc:
                    record[field] = None
                    warnings.append(
                        f"Linha {source_row}: {field.replace('_', ' ')} indisponível para "
                        f"{username} ({exc})."
                    )

            try:
                record["mix"] = _integer_from_cell(cells[mapping["mix"]])
            except ValueError as exc:
                record["mix"] = 0
                warnings.append(
                    f"Linha {source_row}: MIX indisponível para {username} ({exc}). "
                    "O valor foi importado como zero."
                )

            previous_store = store_values.get(emp)
            current_store = (
                record["loja_anterior"],
                record["loja_atual"],
                source_row,
            )
            if previous_store is None:
                store_values[emp] = current_store
            else:
                for label, old, new in (
                    ("Loja Anterior", previous_store[0], current_store[0]),
                    ("Loja Atual", previous_store[1], current_store[1]),
                ):
                    if old is not None and new is not None and old != new:
                        warnings.append(
                            f"EMP {emp}: {label} diverge entre as linhas "
                            f"{previous_store[2]} e {source_row}. O sistema manterá cada "
                            "valor exatamente como veio na linha importada."
                        )
                        break

            records.append(record)

        if not records:
            raise ValueError("Nenhuma linha válida foi encontrada na aba PremiacaoFinal.")

        return {
            "sheet_name": sheet.title,
            "records": records,
            "warnings": warnings,
            "rows_read": rows_read,
            "rows_skipped": rows_skipped,
        }
    finally:
        workbook.close()


def _bind_users_and_validate(db, records: list[dict[str, Any]], warnings: list[str]) -> None:
    users = db.query(Usuario).all()
    exact_by_name = {_norm_username(user.username): user for user in users}
    canonical_users: dict[str, list[Usuario]] = defaultdict(list)
    for user in users:
        canonical_users[_canonical_username(user.username)].append(user)

    links = db.query(UsuarioEmp).filter(UsuarioEmp.ativo.is_(True)).all()
    emps_by_user: dict[int, set[str]] = defaultdict(set)
    for link in links:
        emps_by_user[int(link.usuario_id)].add(_norm_emp(link.emp))

    for record in records:
        username = record["usuario_nome"]
        user = exact_by_name.get(username)
        if user is None:
            candidates = canonical_users.get(_canonical_username(username), [])
            if len(candidates) == 1:
                user = candidates[0]
            elif len(candidates) > 1:
                warnings.append(
                    f"Linha {record['linha_origem']}: nome {username} corresponde a mais de "
                    "um usuário cadastrado; vínculo automático não realizado."
                )

        if user is None:
            record["usuario_id"] = None
            warnings.append(
                f"Linha {record['linha_origem']}: usuário {username} não encontrado no cadastro."
            )
            continue

        record["usuario_id"] = int(user.id)
        allowed = set(emps_by_user.get(int(user.id), set()))
        legacy_emp = _norm_emp(getattr(user, "emp", None))
        if legacy_emp:
            allowed.add(legacy_emp)
        if allowed and record["emp"] not in allowed:
            warnings.append(
                f"Linha {record['linha_origem']}: EMP {record['emp']} não está vinculada "
                f"ao usuário {username} no cadastro atual."
            )
        elif not allowed:
            warnings.append(
                f"Linha {record['linha_origem']}: usuário {username} não possui EMP ativa "
                "no cadastro atual."
            )

        role_sheet = normalize_role(record.get("funcao_planilha"))
        role_db = normalize_role(getattr(user, "role", None))
        if record.get("funcao_planilha") and role_sheet != role_db:
            warnings.append(
                f"Linha {record['linha_origem']}: função da planilha "
                f"({record['funcao_planilha']}) difere do cadastro ({role_db.upper()}). "
                "A permissão seguirá o cadastro do sistema."
            )


def _load_warnings(batch: BonusAtacadoImportacaoLote | None) -> list[str]:
    if not batch or not batch.avisos_json:
        return []
    try:
        value = json.loads(batch.avisos_json)
        return [str(item) for item in value] if isinstance(value, list) else []
    except Exception:
        return []


def _period_options(db) -> list[tuple[int, int]]:
    rows = (
        db.query(BonusAtacadoUsuario.ano, BonusAtacadoUsuario.mes)
        .distinct()
        .order_by(BonusAtacadoUsuario.ano.desc(), BonusAtacadoUsuario.mes.desc())
        .all()
    )
    return [(int(row[0]), int(row[1])) for row in rows]


def _sum_field(rows: list[BonusAtacadoUsuario], field: str) -> Decimal:
    total = Decimal("0")
    for row in rows:
        value = getattr(row, field, None)
        if value is not None:
            total += Decimal(str(value))
    return total


def _attach_current_roles(db, rows: list[BonusAtacadoUsuario]) -> None:
    ids = sorted({int(row.usuario_id) for row in rows if row.usuario_id is not None})
    roles_by_id: dict[int, str] = {}
    if ids:
        for user_id, role in db.query(Usuario.id, Usuario.role).filter(Usuario.id.in_(ids)).all():
            roles_by_id[int(user_id)] = normalize_role(role)

    for row in rows:
        role = roles_by_id.get(int(row.usuario_id)) if row.usuario_id is not None else None
        row.funcao_exibicao = (role or normalize_role(row.funcao_planilha) or "usuario").upper()


def _store_cards(rows: list[BonusAtacadoUsuario]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        card = grouped.setdefault(
            row.emp,
            {
                "emp": row.emp,
                "loja_anterior": None,
                "loja_atual": None,
                "registros": 0,
            },
        )
        card["registros"] += 1
        if card["loja_anterior"] is None and row.loja_anterior is not None:
            card["loja_anterior"] = row.loja_anterior
        if card["loja_atual"] is None and row.loja_atual is not None:
            card["loja_atual"] = row.loja_atual
    return sorted(grouped.values(), key=lambda item: emp_sort_key(item["emp"]))


def admin_bonus_atacado_importar():
    red = _login_required()
    if red:
        return red
    if _role() != "admin":
        flash("Acesso restrito ao administrador.", "warning")
        return redirect(url_for("bonus_atacado"))

    today = date.today()
    ano = _safe_period(request.form.get("ano"), today.year, 2000, 2100)
    mes = _safe_period(request.form.get("mes"), today.month, 1, 12)
    uploaded = request.files.get("arquivo")

    if uploaded is None or not uploaded.filename:
        flash("Selecione a planilha que contém a aba PremiacaoFinal.", "warning")
        return redirect(url_for("bonus_atacado", ano=ano, mes=mes))

    filename = secure_filename(uploaded.filename) or "premiacao_atacado.xlsx"
    extension = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in VALID_EXTENSIONS:
        flash("Envie um arquivo .xlsx ou .xlsm.", "warning")
        return redirect(url_for("bonus_atacado", ano=ano, mes=mes))

    content = uploaded.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        flash("A planilha excede o limite de 15 MB.", "danger")
        return redirect(url_for("bonus_atacado", ano=ano, mes=mes))
    if not content:
        flash("O arquivo enviado está vazio.", "warning")
        return redirect(url_for("bonus_atacado", ano=ano, mes=mes))

    try:
        parsed = _read_workbook(content)
        ensure_bonus_atacado_schema()

        imported_at = datetime.utcnow()
        username = _norm_username(_usuario_logado())
        user_id = session.get("user_id")
        warnings = list(parsed["warnings"])

        with SessionLocal() as db:
            _bind_users_and_validate(db, parsed["records"], warnings)

            # Substituição atômica da competência: em caso de falha, o snapshot
            # anterior permanece intacto.
            db.query(BonusAtacadoUsuario).filter(
                BonusAtacadoUsuario.ano == ano,
                BonusAtacadoUsuario.mes == mes,
            ).delete(synchronize_session=False)

            batch = BonusAtacadoImportacaoLote(
                ano=ano,
                mes=mes,
                arquivo_origem=filename,
                aba_origem=parsed["sheet_name"],
                importado_por_user_id=int(user_id) if user_id else None,
                importado_por=username,
                importado_em=imported_at,
                linhas_lidas=int(parsed["rows_read"]),
                linhas_importadas=len(parsed["records"]),
                linhas_ignoradas=int(parsed["rows_skipped"]),
                avisos_json=json.dumps(warnings, ensure_ascii=False),
            )
            db.add(batch)
            db.flush()

            rows = []
            for record in parsed["records"]:
                rows.append(
                    BonusAtacadoUsuario(
                        **record,
                        lote_id=batch.id,
                        ano=ano,
                        mes=mes,
                        importado_em=imported_at,
                    )
                )
            db.add_all(rows)
            db.commit()

        audit(
            "bonus_atacado_snapshot_imported",
            ano=ano,
            mes=mes,
            arquivo=filename,
            linhas=len(parsed["records"]),
            ignoradas=parsed["rows_skipped"],
            avisos=len(warnings),
        )
        flash(
            f"Bônus Atacado de {mes:02d}/{ano} importado: "
            f"{len(parsed['records'])} registro(s).",
            "success",
        )
        if warnings:
            flash(
                f"A importação terminou com {len(warnings)} aviso(s). "
                "Consulte a validação na página.",
                "warning",
            )
    except Exception as exc:
        current_app.logger.exception("Falha ao importar Bônus Atacado")
        audit(
            "bonus_atacado_snapshot_import_failed",
            ano=ano,
            mes=mes,
            arquivo=filename,
            erro=str(exc),
        )
        flash(f"Não foi possível importar a planilha: {exc}", "danger")

    return redirect(url_for("bonus_atacado", ano=ano, mes=mes))


def bonus_atacado():
    red = _login_required()
    if red:
        return red

    role = normalize_role(_role())
    today = date.today()

    try:
        ensure_bonus_atacado_schema()
    except Exception:
        current_app.logger.exception("Falha ao garantir schema de Bônus Atacado")
        flash(
            "As tabelas do Bônus Atacado ainda não estão disponíveis. "
            "Execute o SQL de implantação no Supabase.",
            "danger",
        )
        return render_template(
            "bonus_atacado.html",
            role=role,
            ano=today.year,
            mes=today.month,
            rows=[],
            periods=[],
            emp_options=[],
            user_options=[],
            emp_filter="",
            user_filter="",
            summary={},
            store_cards=[],
            batch=None,
            batch_warnings=[],
            can_view_store_metrics=(role in MANAGER_ROLES),
            admin_user_options=[],
            db_unavailable=True,
        )

    current_username = _norm_username(_usuario_logado())
    current_user_id = session.get("user_id")

    with SessionLocal() as db:
        periods = _period_options(db)
        default_year, default_month = periods[0] if periods else (today.year, today.month)
        ano = _safe_period(request.args.get("ano"), default_year, 2000, 2100)
        mes = _safe_period(request.args.get("mes"), default_month, 1, 12)

        period_query = db.query(BonusAtacadoUsuario).filter(
            BonusAtacadoUsuario.ano == ano,
            BonusAtacadoUsuario.mes == mes,
        )

        emp_filter = ""
        user_filter = ""
        emp_options: list[str] = []
        user_options: list[str] = []
        all_period_rows: list[BonusAtacadoUsuario] = []

        if role == "admin":
            all_period_rows = period_query.all()
            emp_filter = _norm_emp(request.args.get("emp"))
            user_filter = _norm_username(request.args.get("usuario"))
            query = period_query
            if emp_filter:
                query = query.filter(BonusAtacadoUsuario.emp == emp_filter)
            if user_filter:
                query = query.filter(BonusAtacadoUsuario.usuario_nome == user_filter)
            rows = query.all()
            emp_options = sorted(
                {row.emp for row in all_period_rows if row.emp}, key=emp_sort_key
            )
            user_options = sorted(
                {row.usuario_nome for row in all_period_rows if row.usuario_nome}
            )
        elif role in {"supervisor", "gerente"}:
            allowed_emps = sorted(
                {_norm_emp(emp) for emp in (_allowed_emps() or []) if _norm_emp(emp)},
                key=emp_sort_key,
            )
            if allowed_emps:
                rows = period_query.filter(BonusAtacadoUsuario.emp.in_(allowed_emps)).all()
            else:
                own_filter = [BonusAtacadoUsuario.usuario_nome == current_username]
                if current_user_id:
                    own_filter.append(BonusAtacadoUsuario.usuario_id == int(current_user_id))
                rows = period_query.filter(or_(*own_filter)).all()
        else:
            # Vendedor, mecânico e demais perfis operacionais jamais recebem
            # registros de outro usuário.
            own_filter = [BonusAtacadoUsuario.usuario_nome == current_username]
            if current_user_id:
                own_filter.append(BonusAtacadoUsuario.usuario_id == int(current_user_id))
            rows = period_query.filter(or_(*own_filter)).all()

        rows.sort(key=lambda row: (emp_sort_key(row.emp), row.usuario_nome))
        _attach_current_roles(db, rows)
        try:
            ensure_itens_parados_snapshot_schema()
            itens_parados_summary = attach_saldos_to_bonus_rows(
                db, rows, ano=ano, mes=mes
            )
        except Exception:
            current_app.logger.exception(
                "Falha ao carregar saldo de itens parados no Bônus Atacado"
            )
            itens_parados_summary = {
                "total_visivel": Decimal("0"),
                "total_lojas": Decimal("0"),
            }
            for row in rows:
                row.saldo_itens_parados = Decimal("0")
                row.saldo_itens_parados_usuario = Decimal("0")
                row.saldo_itens_parados_loja = Decimal("0")

        try:
            ensure_bonus_outros_valores_schema()
            outros_valores_summary = attach_outros_valores_to_bonus_rows(
                db,
                rows,
                ano=ano,
                mes=mes,
                origem="ATACADO",
                bonus_field="total_produtos",
            )
        except Exception:
            current_app.logger.exception(
                "Falha ao carregar Outros Valores no Bônus Atacado"
            )
            outros_valores_summary = {
                "total_visivel": Decimal("0"),
                "quantidade": 0,
            }
            for row in rows:
                row.outros_valores_total = Decimal("0")
                row.outros_valores_detalhes = []
                row.valor_bonus_base = Decimal(str(row.total_produtos or 0))
                row.total_geral = row.valor_bonus_base + Decimal(
                    str(getattr(row, "saldo_itens_parados", 0) or 0)
                )

        batch = (
            db.query(BonusAtacadoImportacaoLote)
            .filter(
                BonusAtacadoImportacaoLote.ano == ano,
                BonusAtacadoImportacaoLote.mes == mes,
            )
            .order_by(BonusAtacadoImportacaoLote.importado_em.desc())
            .first()
        )

        bonus_total = _sum_field(rows, "total_produtos")
        summary = {
            "registros": len(rows),
            "emps": len({row.emp for row in rows if row.emp}),
            "total_produtos": bonus_total,
            "importado": _sum_field(rows, "importado"),
            "itens_parados": itens_parados_summary["total_visivel"],
            "outros_valores": outros_valores_summary["total_visivel"],
            "total_geral": (
                bonus_total
                + itens_parados_summary["total_visivel"]
                + outros_valores_summary["total_visivel"]
            ),
        }
        can_view_store_metrics = role in MANAGER_ROLES
        stores = _store_cards(rows) if can_view_store_metrics else []
        for store in stores:
            store_rows = [row for row in rows if row.emp == store["emp"]]
            store["itens_parados"] = next(
                (
                    row.saldo_itens_parados_loja
                    for row in store_rows
                ),
                Decimal("0"),
            )
            store["bonus"] = sum(
                (Decimal(str(row.total_produtos or 0)) for row in store_rows),
                Decimal("0"),
            )
            store["outros_valores"] = sum(
                (Decimal(str(row.outros_valores_total or 0)) for row in store_rows),
                Decimal("0"),
            )
            store["total_geral"] = (
                store["bonus"]
                + store["itens_parados"]
                + store["outros_valores"]
            )
        admin_user_options = bonus_row_options(all_period_rows) if role == "admin" else []
        batch_warnings = _load_warnings(batch) if role == "admin" else []

    return render_template(
        "bonus_atacado.html",
        role=role,
        ano=ano,
        mes=mes,
        rows=rows,
        periods=periods,
        emp_options=emp_options,
        user_options=user_options,
        emp_filter=emp_filter,
        user_filter=user_filter,
        summary=summary,
        store_cards=stores,
        batch=batch,
        batch_warnings=batch_warnings,
        can_view_store_metrics=can_view_store_metrics,
        admin_user_options=admin_user_options,
        db_unavailable=False,
    )
