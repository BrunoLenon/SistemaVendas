# -*- coding: utf-8 -*-
"""Importação e snapshot mensal de Itens Parados por pontos.

Fluxo:
1. O administrador cadastra os produtos ativos por EMP.
2. Define a regra de pontos (global ou específica por EMP).
3. Importa a planilha de vendas.
4. Somente as linhas cujo código esteja ativo na mesma EMP entram no cálculo.
5. O valor líquido elegível é consolidado por funcionário/EMP.
6. Pontos inteiros = piso(valor elegível / base em reais).
7. Bônus = pontos inteiros x valor por ponto.

As páginas de Bônus e Financeiro leem o bônus pronto do snapshot e não
recalculam a planilha a cada acesso.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_FLOOR, ROUND_HALF_UP
from io import BytesIO
from typing import Any, Iterable

from flask import current_app, flash, redirect, request, session, url_for
from werkzeug.utils import secure_filename

from auth_helpers import _login_required, _role, _usuario_logado
from db import (
    ItemParado,
    ItensParadosPontosConfig,
    ItensParadosVendaImportacaoLote,
    ItensParadosVendaUsuario,
    SessionLocal,
    Usuario,
    UsuarioEmp,
    ensure_itens_parados_snapshot_schema,
)
from security_utils import audit, normalize_role


MAX_UPLOAD_BYTES = 20 * 1024 * 1024
VALID_EXTENSIONS = {".xlsx", ".xlsm"}

SALE_MOVEMENTS = {"OA", "OV", "SV", "VA", "VV"}
NEGATIVE_MOVEMENTS = {"CA", "DS"}
ALLOWED_MOVEMENTS = SALE_MOVEMENTS | NEGATIVE_MOVEMENTS

HEADER_ALIASES: dict[str, set[str]] = {
    "codigo": {"MESTRE", "CODIGO", "CODIGOPRODUTO", "CODPRODUTO"},
    "descricao": {"DESCRICAO", "DESCRICAOPRODUTO", "PRODUTO"},
    "movimento": {"MOVIMENTO", "DATA", "DATAVENDA", "EMISSAO"},
    "tipo_movimento": {
        "MOVTIPOMOVTO",
        "TIPOMOVIMENTO",
        "MOVIMENTOTIPO",
        "TIPOMOVTO",
    },
    "usuario_nome": {"VENDEDOR", "FUNCIONARIO", "USUARIO"},
    "emp": {"EMP", "EMPRESA", "LOJA"},
    "quantidade": {"QTDADEVENDIDA", "QUANTIDADE", "QTD"},
    "valor_total": {"VALORTOTAL", "TOTAL", "VALORVENDA"},
}
REQUIRED_FIELDS = {"codigo", "movimento", "usuario_nome", "emp", "valor_total"}

DEFAULT_BASE_REAIS = Decimal("100")
DEFAULT_VALOR_PONTO = Decimal("10")


def register_itens_parados_snapshot_routes(app) -> None:
    app.add_url_rule(
        "/admin/itens-parados/vendas/importar",
        endpoint="admin_itens_parados_vendas_importar",
        view_func=admin_itens_parados_vendas_importar,
        methods=["POST"],
    )


def _strip_accents(value: object) -> str:
    raw = str(value or "")
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch)
    )


def norm_header(value: object) -> str:
    raw = _strip_accents(value).strip().upper().replace("%", " PERCENTUAL ")
    return re.sub(r"[^A-Z0-9]", "", raw)


def norm_username(value: object) -> str:
    raw = re.sub(r"\s+", " ", str(value or "").strip())
    return raw.upper()


def canonical_username(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", _strip_accents(value).upper())


def norm_emp(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).upper()
    if isinstance(value, (int, float, Decimal)):
        try:
            number = Decimal(str(value))
            if number == number.to_integral_value():
                return str(int(number))
        except Exception:
            pass
    raw = str(value).strip()
    if re.fullmatch(r"[+-]?\d+[.,]0+", raw):
        return raw.split(".", 1)[0].split(",", 1)[0]
    return raw.upper()


def norm_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).upper()
    if isinstance(value, (int, float, Decimal)):
        try:
            number = Decimal(str(value))
            if number == number.to_integral_value():
                return str(int(number))
        except Exception:
            pass
    raw = str(value).strip()
    if re.fullmatch(r"[+-]?\d+[.,]0+", raw):
        raw = raw.split(".", 1)[0].split(",", 1)[0]
    return re.sub(r"\s+", "", raw).upper()


def _safe_period(value: object, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(str(value).strip())
    except Exception:
        return default
    return max(minimum, min(maximum, parsed))


def _decimal(value: object) -> Decimal:
    if value is None or (isinstance(value, str) and not value.strip()):
        return Decimal("0")
    if isinstance(value, Decimal):
        result = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        result = Decimal(str(value))
    else:
        raw = str(value).strip().replace("R$", "").replace(" ", "")
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        elif raw.count(".") > 1:
            raw = raw.replace(".", "")
        try:
            result = Decimal(raw)
        except (InvalidOperation, ValueError, TypeError) as exc:
            raise ValueError(f"valor numérico inválido ({value})") from exc
    if not result.is_finite():
        raise ValueError(f"valor numérico inválido ({value})")
    return result.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


def _date_from_cell(value: object) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except Exception:
            return None
    raw = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw[:10], fmt).date()
        except Exception:
            continue
    return None


def _find_sheet_and_header(workbook) -> tuple[Any, int, dict[str, int]]:
    best: tuple[Any, int, dict[str, int]] | None = None
    for sheet in workbook.worksheets:
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 30)), start=1
        ):
            mapping: dict[str, int] = {}
            for index, cell in enumerate(cells):
                normalized = norm_header(cell.value)
                if not normalized:
                    continue
                for field, aliases in HEADER_ALIASES.items():
                    if field not in mapping and normalized in aliases:
                        mapping[field] = index
                        break
            if best is None or len(mapping) > len(best[2]):
                best = (sheet, row_number, mapping)
            if REQUIRED_FIELDS.issubset(mapping):
                return sheet, row_number, mapping
    if best is None:
        raise ValueError("A planilha está vazia.")
    missing = sorted(REQUIRED_FIELDS - set(best[2]))
    labels = ", ".join(field.replace("_", " ").upper() for field in missing)
    raise ValueError(f"Colunas obrigatórias não encontradas: {labels}.")


def load_active_item_windows(db) -> dict[tuple[str, str], list[tuple[date | None, date | None]]]:
    """Retorna os produtos ativos por EMP/código, incluindo validade opcional."""
    result: defaultdict[
        tuple[str, str], list[tuple[date | None, date | None]]
    ] = defaultdict(list)
    rows = (
        db.query(
            ItemParado.emp,
            ItemParado.codigo,
            ItemParado.data_inicio,
            ItemParado.data_fim,
        )
        .filter(ItemParado.ativo.is_(True))
        .all()
    )
    for emp, codigo, data_inicio, data_fim in rows:
        emp_code = norm_emp(emp)
        product_code = norm_code(codigo)
        if emp_code and product_code:
            result[(emp_code, product_code)].append((data_inicio, data_fim))
    return dict(result)


def load_point_rules(db) -> dict[str, Any]:
    """Carrega a regra global e a regra mais recente de cada EMP."""
    global_rule = (DEFAULT_BASE_REAIS, DEFAULT_VALOR_PONTO)
    by_emp: dict[str, tuple[Decimal, Decimal]] = {}

    configs = (
        db.query(ItensParadosPontosConfig)
        .filter(ItensParadosPontosConfig.ativo.is_(True))
        .order_by(ItensParadosPontosConfig.id.asc())
        .all()
    )
    for config in configs:
        base = _decimal(getattr(config, "base_reais", DEFAULT_BASE_REAIS))
        point_value = _decimal(getattr(config, "valor_por_ponto", DEFAULT_VALOR_PONTO))
        if base <= 0 or point_value < 0:
            continue
        emp = norm_emp(getattr(config, "emp", None))
        if emp:
            by_emp[emp] = (base, point_value)
        else:
            global_rule = (base, point_value)

    return {"global": global_rule, "emps": by_emp}


def point_rule_for_emp(point_rules: dict[str, Any], emp: object) -> tuple[Decimal, Decimal]:
    emp_code = norm_emp(emp)
    by_emp = point_rules.get("emps") or {}
    return by_emp.get(emp_code) or point_rules.get("global") or (
        DEFAULT_BASE_REAIS,
        DEFAULT_VALOR_PONTO,
    )


def _item_is_active(
    active_items: dict[tuple[str, str], list[tuple[date | None, date | None]]],
    *,
    emp: str,
    codigo: str,
    movement_date: date,
) -> bool:
    windows = active_items.get((emp, codigo)) or []
    for start, end in windows:
        if start is not None and movement_date < start:
            continue
        if end is not None and movement_date > end:
            continue
        return True
    return False


def _parse_sales_workbook(
    content: bytes,
    *,
    ano: int,
    mes: int,
    active_items: dict[tuple[str, str], list[tuple[date | None, date | None]]],
    point_rules: dict[str, Any],
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("A biblioteca openpyxl não está disponível no servidor.") from exc

    try:
        workbook = load_workbook(
            filename=BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ValueError("Não foi possível abrir a planilha de vendas de itens parados.") from exc

    try:
        sheet, header_row, mapping = _find_sheet_and_header(workbook)
        max_col = max(mapping.values()) + 1
        grouped: dict[tuple[str, str], dict[str, Any]] = {}
        warnings: list[str] = []
        rows_read = 0
        rows_imported = 0
        rows_skipped = 0
        outside_period = 0
        not_eligible = 0
        invalid_movements: defaultdict[str, int] = defaultdict(int)
        data_min: date | None = None
        data_max: date | None = None

        for source_row, cells_tuple in enumerate(
            sheet.iter_rows(min_row=header_row + 1, max_col=max_col),
            start=header_row + 1,
        ):
            cells = list(cells_tuple)
            values = [cells[mapping[field]].value for field in REQUIRED_FIELDS]
            if not any(value not in (None, "") for value in values):
                continue
            rows_read += 1

            username = norm_username(cells[mapping["usuario_nome"]].value)
            emp = norm_emp(cells[mapping["emp"]].value)
            codigo = norm_code(cells[mapping["codigo"]].value)
            movement_date = _date_from_cell(cells[mapping["movimento"]].value)
            if not username or not emp or not codigo or movement_date is None:
                rows_skipped += 1
                missing = []
                if not username:
                    missing.append("VENDEDOR")
                if not emp:
                    missing.append("EMP")
                if not codigo:
                    missing.append("CODIGO/MESTRE")
                if movement_date is None:
                    missing.append("MOVIMENTO")
                warnings.append(
                    f"Linha {source_row} ignorada: {', '.join(missing)} inválido(s) ou vazio(s)."
                )
                continue

            data_min = movement_date if data_min is None or movement_date < data_min else data_min
            data_max = movement_date if data_max is None or movement_date > data_max else data_max
            if movement_date.year != ano or movement_date.month != mes:
                outside_period += 1
                rows_skipped += 1
                continue

            movement_type = ""
            if "tipo_movimento" in mapping:
                movement_type = str(
                    cells[mapping["tipo_movimento"]].value or ""
                ).strip().upper()
                if movement_type and movement_type not in ALLOWED_MOVEMENTS:
                    invalid_movements[movement_type] += 1
                    rows_skipped += 1
                    continue

            if "quantidade" in mapping:
                try:
                    quantity = _decimal(cells[mapping["quantidade"]].value)
                except ValueError:
                    rows_skipped += 1
                    warnings.append(f"Linha {source_row} ignorada: quantidade inválida.")
                    continue
                if quantity == 0:
                    rows_skipped += 1
                    continue
                if movement_type not in NEGATIVE_MOVEMENTS and quantity < 0:
                    rows_skipped += 1
                    continue

            if not _item_is_active(
                active_items,
                emp=emp,
                codigo=codigo,
                movement_date=movement_date,
            ):
                not_eligible += 1
                rows_skipped += 1
                continue

            try:
                total = _decimal(cells[mapping["valor_total"]].value)
            except ValueError as exc:
                rows_skipped += 1
                warnings.append(f"Linha {source_row} ignorada: {exc}.")
                continue
            if total == 0:
                rows_skipped += 1
                continue

            signed_total = -abs(total) if movement_type in NEGATIVE_MOVEMENTS else abs(total)
            key = (emp, username)
            item = grouped.setdefault(
                key,
                {
                    "emp": emp,
                    "usuario_nome": username,
                    "valor_total": Decimal("0"),
                    "qtd_linhas": 0,
                    "codigos": set(),
                },
            )
            item["valor_total"] += signed_total
            item["qtd_linhas"] += 1
            item["codigos"].add(codigo)
            rows_imported += 1

        if outside_period:
            warnings.append(
                f"{outside_period} linha(s) fora da competência {mes:02d}/{ano} foram ignoradas."
            )
        if not_eligible:
            warnings.append(
                f"{not_eligible} linha(s) não correspondiam a produtos ativos na mesma EMP e foram ignoradas."
            )
        for movement, count in sorted(invalid_movements.items()):
            warnings.append(
                f"{count} linha(s) com movimento {movement} foram ignoradas; "
                "são aceitos OA, OV, SV, VA, VV, CA e DS."
            )
        if not grouped:
            raise ValueError(
                "Nenhuma venda válida de produto ativo foi encontrada para a competência selecionada."
            )

        records: list[dict[str, Any]] = []
        valor_total_geral = Decimal("0")
        pontos_total_geral = 0
        bonus_total_geral = Decimal("0")

        for item in grouped.values():
            base_reais, valor_por_ponto = point_rule_for_emp(point_rules, item["emp"])
            valor_liquido = item["valor_total"].quantize(
                Decimal("0.0001"), rounding=ROUND_HALF_UP
            )
            valor_base_pontos = max(valor_liquido, Decimal("0"))
            pontos = (
                int((valor_base_pontos / base_reais).to_integral_value(rounding=ROUND_FLOOR))
                if base_reais > 0
                else 0
            )
            bonus = (Decimal(pontos) * valor_por_ponto).quantize(
                Decimal("0.0001"), rounding=ROUND_HALF_UP
            )
            record = {
                "emp": item["emp"],
                "usuario_nome": item["usuario_nome"],
                "valor_total": valor_liquido,
                "pontos": pontos,
                "base_reais": base_reais,
                "valor_por_ponto": valor_por_ponto,
                "bonus_total": bonus,
                "qtd_linhas": int(item["qtd_linhas"]),
                "qtd_itens": len(item["codigos"]),
            }
            records.append(record)
            valor_total_geral += valor_liquido
            pontos_total_geral += pontos
            bonus_total_geral += bonus

        records.sort(key=lambda row: (row["emp"], row["usuario_nome"]))
        return {
            "sheet_name": sheet.title,
            "records": records,
            "warnings": warnings,
            "rows_read": rows_read,
            "rows_imported": rows_imported,
            "rows_skipped": rows_skipped,
            "rows_not_eligible": not_eligible,
            "data_min": data_min,
            "data_max": data_max,
            "valor_total": valor_total_geral.quantize(
                Decimal("0.0001"), rounding=ROUND_HALF_UP
            ),
            "pontos_total": pontos_total_geral,
            "bonus_total": bonus_total_geral.quantize(
                Decimal("0.0001"), rounding=ROUND_HALF_UP
            ),
        }
    finally:
        workbook.close()


def _resolve_users(db, records: list[dict[str, Any]], warnings: list[str]) -> None:
    users = db.query(Usuario.id, Usuario.username, Usuario.emp).all()
    by_canonical: defaultdict[str, list[tuple[int, str, str | None]]] = defaultdict(list)
    for user_id, username, primary_emp in users:
        by_canonical[canonical_username(username)].append(
            (int(user_id), str(username or ""), norm_emp(primary_emp))
        )

    user_emps: defaultdict[int, set[str]] = defaultdict(set)
    for user_id, emp in (
        db.query(UsuarioEmp.usuario_id, UsuarioEmp.emp)
        .filter(UsuarioEmp.ativo.is_(True))
        .all()
    ):
        user_emps[int(user_id)].add(norm_emp(emp))
    for user_id, _, primary_emp in users:
        if primary_emp:
            user_emps[int(user_id)].add(norm_emp(primary_emp))

    warned_missing: set[str] = set()
    warned_emp: set[tuple[int, str]] = set()
    for record in records:
        canonical = canonical_username(record["usuario_nome"])
        candidates = by_canonical.get(canonical) or []
        if len(candidates) == 1:
            user_id, db_username, _ = candidates[0]
            record["usuario_id"] = user_id
            record["usuario_nome"] = norm_username(db_username)
            if record["emp"] not in user_emps.get(user_id, set()):
                key = (user_id, record["emp"])
                if key not in warned_emp:
                    warnings.append(
                        f"{record['usuario_nome']}: EMP {record['emp']} não está vinculada "
                        "ao usuário no cadastro. O resultado foi mantido para conferência."
                    )
                    warned_emp.add(key)
        elif len(candidates) > 1:
            record["usuario_id"] = None
            warnings.append(
                f"{record['usuario_nome']}: mais de um usuário corresponde ao nome; "
                "vínculo automático não realizado."
            )
        else:
            record["usuario_id"] = None
            if canonical not in warned_missing:
                warnings.append(
                    f"{record['usuario_nome']}: usuário não localizado no cadastro do sistema."
                )
                warned_missing.add(canonical)


def admin_itens_parados_vendas_importar():
    red = _login_required()
    if red:
        return red
    if normalize_role(_role()) != "admin":
        flash("Acesso restrito ao administrador.", "warning")
        return redirect(url_for("itens_parados"))

    today = date.today()
    ano = _safe_period(request.form.get("ano"), today.year, 2000, 2100)
    mes = _safe_period(request.form.get("mes"), today.month, 1, 12)
    uploaded = request.files.get("arquivo_vendas_itens")
    redirect_args = {"ano_vendas": ano, "mes_vendas": mes}

    if uploaded is None or not uploaded.filename:
        flash("Selecione a planilha de vendas dos itens parados.", "warning")
        return redirect(url_for("admin_itens_parados", **redirect_args))

    filename = secure_filename(uploaded.filename) or "itens_parados.xlsx"
    extension = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in VALID_EXTENSIONS:
        flash("Envie um arquivo .xlsx ou .xlsm.", "warning")
        return redirect(url_for("admin_itens_parados", **redirect_args))

    content = uploaded.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        flash("A planilha excede o limite de 20 MB.", "danger")
        return redirect(url_for("admin_itens_parados", **redirect_args))
    if not content:
        flash("O arquivo enviado está vazio.", "warning")
        return redirect(url_for("admin_itens_parados", **redirect_args))

    try:
        ensure_itens_parados_snapshot_schema()
        with SessionLocal() as db:
            active_items = load_active_item_windows(db)
            point_rules = load_point_rules(db)
        if not active_items:
            raise ValueError(
                "Nenhum produto ativo foi cadastrado. Importe ou cadastre os itens antes das vendas."
            )

        parsed = _parse_sales_workbook(
            content,
            ano=ano,
            mes=mes,
            active_items=active_items,
            point_rules=point_rules,
        )
        warnings = list(parsed["warnings"])
        imported_at = datetime.utcnow()
        current_user_id = session.get("user_id")
        current_username = norm_username(_usuario_logado())

        with SessionLocal() as db:
            _resolve_users(db, parsed["records"], warnings)
            db.query(ItensParadosVendaUsuario).filter(
                ItensParadosVendaUsuario.ano == ano,
                ItensParadosVendaUsuario.mes == mes,
            ).delete(synchronize_session=False)
            db.query(ItensParadosVendaImportacaoLote).filter(
                ItensParadosVendaImportacaoLote.ano == ano,
                ItensParadosVendaImportacaoLote.mes == mes,
            ).delete(synchronize_session=False)

            batch = ItensParadosVendaImportacaoLote(
                ano=ano,
                mes=mes,
                arquivo_origem=filename,
                importado_por_user_id=int(current_user_id) if current_user_id else None,
                importado_por=current_username,
                importado_em=imported_at,
                data_inicio=parsed["data_min"],
                data_fim=parsed["data_max"],
                linhas_lidas=int(parsed["rows_read"]),
                linhas_importadas=int(parsed["rows_imported"]),
                linhas_ignoradas=int(parsed["rows_skipped"]),
                registros_usuarios=len(parsed["records"]),
                valor_total=parsed["valor_total"],
                pontos_total=int(parsed["pontos_total"]),
                bonus_total=parsed["bonus_total"],
                linhas_nao_elegiveis=int(parsed["rows_not_eligible"]),
                avisos_json=json.dumps(warnings[:500], ensure_ascii=False),
            )
            db.add(batch)
            db.flush()

            db.add_all(
                [
                    ItensParadosVendaUsuario(
                        lote_id=batch.id,
                        ano=ano,
                        mes=mes,
                        usuario_id=record.get("usuario_id"),
                        usuario_nome=record["usuario_nome"],
                        emp=record["emp"],
                        valor_total=record["valor_total"],
                        pontos=record["pontos"],
                        base_reais=record["base_reais"],
                        valor_por_ponto=record["valor_por_ponto"],
                        bonus_total=record["bonus_total"],
                        qtd_linhas=record["qtd_linhas"],
                        qtd_itens=record["qtd_itens"],
                        importado_em=imported_at,
                    )
                    for record in parsed["records"]
                ]
            )
            db.commit()

        audit(
            "itens_parados_vendas_imported",
            ano=ano,
            mes=mes,
            arquivo=filename,
            linhas=parsed["rows_imported"],
            usuarios=len(parsed["records"]),
            valor_elegivel=str(parsed["valor_total"]),
            pontos=parsed["pontos_total"],
            bonus=str(parsed["bonus_total"]),
        )
        valor_fmt = f"{parsed['valor_total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        bonus_fmt = f"{parsed['bonus_total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        flash(
            f"Itens Parados de {mes:02d}/{ano} calculados: "
            f"R$ {valor_fmt} em vendas elegíveis, {parsed['pontos_total']} ponto(s) "
            f"e R$ {bonus_fmt} de bônus.",
            "success",
        )
        if warnings:
            flash(
                f"A importação terminou com {len(warnings)} aviso(s). Consulte o resumo na página.",
                "warning",
            )
    except Exception as exc:
        current_app.logger.exception("Falha ao importar vendas de itens parados")
        audit(
            "itens_parados_vendas_import_failed",
            ano=ano,
            mes=mes,
            arquivo=filename,
            erro=str(exc),
        )
        flash(f"Não foi possível importar a planilha: {exc}", "danger")

    return redirect(url_for("admin_itens_parados", **redirect_args))


def load_period_rows(db, ano: int, mes: int):
    return (
        db.query(ItensParadosVendaUsuario)
        .filter(
            ItensParadosVendaUsuario.ano == int(ano),
            ItensParadosVendaUsuario.mes == int(mes),
        )
        .all()
    )


def period_options(db) -> list[tuple[int, int]]:
    rows = (
        db.query(ItensParadosVendaUsuario.ano, ItensParadosVendaUsuario.mes)
        .distinct()
        .order_by(
            ItensParadosVendaUsuario.ano.desc(),
            ItensParadosVendaUsuario.mes.desc(),
        )
        .all()
    )
    return [(int(row[0]), int(row[1])) for row in rows]


def latest_batch(db, ano: int, mes: int):
    return (
        db.query(ItensParadosVendaImportacaoLote)
        .filter(
            ItensParadosVendaImportacaoLote.ano == int(ano),
            ItensParadosVendaImportacaoLote.mes == int(mes),
        )
        .order_by(ItensParadosVendaImportacaoLote.importado_em.desc())
        .first()
    )


def batch_warnings(batch: ItensParadosVendaImportacaoLote | None) -> list[str]:
    if not batch or not batch.avisos_json:
        return []
    try:
        value = json.loads(batch.avisos_json)
        return [str(item) for item in value] if isinstance(value, list) else []
    except Exception:
        return []


def attach_saldos_to_bonus_rows(
    db,
    rows: Iterable[Any],
    *,
    ano: int,
    mes: int,
) -> dict[str, Decimal]:
    """Anexa o bônus calculado de Itens Parados aos objetos de Bônus."""
    rows = list(rows)
    emps = sorted(
        {
            norm_emp(getattr(row, "emp", ""))
            for row in rows
            if norm_emp(getattr(row, "emp", ""))
        }
    )
    if not emps:
        return {"total_visivel": Decimal("0"), "total_lojas": Decimal("0")}

    balances = (
        db.query(ItensParadosVendaUsuario)
        .filter(
            ItensParadosVendaUsuario.ano == int(ano),
            ItensParadosVendaUsuario.mes == int(mes),
            ItensParadosVendaUsuario.emp.in_(emps),
        )
        .all()
    )
    by_key = {
        (norm_emp(item.emp), norm_username(item.usuario_nome)): Decimal(
            str(item.bonus_total or 0)
        )
        for item in balances
    }
    sales_by_key = {
        (norm_emp(item.emp), norm_username(item.usuario_nome)): Decimal(
            str(item.valor_total or 0)
        )
        for item in balances
    }
    points_by_key = {
        (norm_emp(item.emp), norm_username(item.usuario_nome)): int(item.pontos or 0)
        for item in balances
    }

    by_emp: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    sales_by_emp: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    points_by_emp: defaultdict[str, int] = defaultdict(int)
    for item in balances:
        emp = norm_emp(item.emp)
        by_emp[emp] += Decimal(str(item.bonus_total or 0))
        sales_by_emp[emp] += Decimal(str(item.valor_total or 0))
        points_by_emp[emp] += int(item.pontos or 0)

    visible_keys: set[tuple[str, str]] = set()
    manager_emps: set[str] = set()
    for row in rows:
        emp = norm_emp(getattr(row, "emp", ""))
        username = norm_username(getattr(row, "usuario_nome", ""))
        key = (emp, username)
        individual = by_key.get(key, Decimal("0"))
        store = by_emp.get(emp, Decimal("0"))

        setattr(row, "saldo_itens_parados_usuario", individual)
        setattr(row, "saldo_itens_parados_loja", store)
        setattr(row, "itens_parados_vendas_usuario", sales_by_key.get(key, Decimal("0")))
        setattr(row, "itens_parados_vendas_loja", sales_by_emp.get(emp, Decimal("0")))
        setattr(row, "itens_parados_pontos_usuario", points_by_key.get(key, 0))
        setattr(row, "itens_parados_pontos_loja", points_by_emp.get(emp, 0))

        role_value = (
            getattr(row, "funcao_exibicao", None)
            or getattr(row, "funcao", None)
            or getattr(row, "funcao_planilha", None)
            or ""
        )
        role = normalize_role(role_value)
        shown = store if role in {"gerente", "supervisor"} else individual
        setattr(row, "saldo_itens_parados", shown)
        if role in {"gerente", "supervisor"}:
            manager_emps.add(emp)
            continue
        visible_keys.add(key)

    individual_emps = {emp for emp, _ in visible_keys}
    total_visible = sum(
        (by_key.get(key, Decimal("0")) for key in visible_keys), Decimal("0")
    )
    total_visible += sum(
        (by_emp.get(emp, Decimal("0")) for emp in manager_emps if emp not in individual_emps),
        Decimal("0"),
    )
    return {
        "total_visivel": total_visible,
        "total_lojas": sum((by_emp[emp] for emp in emps), Decimal("0")),
    }
