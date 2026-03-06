from __future__ import annotations

from typing import Callable

from flask import request, render_template, send_file


def register_operacoes_vendas_produtos_routes(
    app,
    *,
    login_required_fn,
    allowed_emps_fn: Callable[[], list[str]],
    role_fn: Callable[[], str],
    emp_fn: Callable[[], str],
) -> None:
    """
    Rotas de Operações (refatoração pura):
    - /operacoes/vendas-produtos (HTML + export Excel)
    - /operacoes/api/produtos_suggest (typeahead)

    Mantém endpoints/URLs/contratos do app.py original.
    """

    @login_required_fn
    def operacoes_vendas_produto():
        """
        Pesquisa de vendas por produto (descrição) e/ou marca.
        - Busca flexível por "começa com" + termos em sequência (ex.: 'PNEU TRAS 90 90 18' encontra 'PNEU TRASEIRO 90/90-18')
        - Retorna matriz: EMP + Vendedor + colunas por mês + média mensal
        - Exporta Excel via ?export=1
        """
        from db import SessionLocal, Venda, Usuario, UsuarioEmp
        from sqlalchemy import func, case
        from flask import jsonify
        from io import BytesIO
        from datetime import date, timedelta

        def _norm(s: str) -> str:
            import unicodedata, re
            s = (s or "").strip()
            s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
            s = re.sub(r"\s+", " ", s).strip().lower()
            return s

        def _month_range(ano_i: int, mes_i: int, ano_f: int, mes_f: int):
            out = []
            y, m = ano_i, mes_i
            while (y < ano_f) or (y == ano_f and m <= mes_f):
                out.append((y, m))
                m += 1
                if m == 13:
                    m = 1
                    y += 1
            return out

        def _last_day(y: int, m: int):
            if m == 12:
                return date(y, 12, 31)
            return date(y, m + 1, 1) - timedelta(days=1)

        # ===== filtros =====
        produto_raw = (request.args.get("produto") or "").strip()
        marca_raw = (request.args.get("marca") or "").strip()
        mestre_raw = (request.args.get("mestre") or "").strip()

        # meses (default: Jan -> mês atual do ano atual)
        today = date.today()
        meses_select = []
        # gera lista de meses (Jan do ano atual até Dez do ano atual) para selects
        for mm in range(1, 13):
            meses_select.append({"value": f"{today.year:04d}-{mm:02d}", "label": f"{mm:02d}/{today.year:04d}"})

        mes_ini = (request.args.get("mes_ini") or f"{today.year:04d}-01").strip()
        mes_fim = (request.args.get("mes_fim") or f"{today.year:04d}-{today.month:02d}").strip()

        try:
            ano_i, m_i = [int(x) for x in mes_ini.split("-")]
            ano_f, m_f = [int(x) for x in mes_fim.split("-")]
        except Exception:
            ano_i, m_i = today.year, 1
            ano_f, m_f = today.year, today.month

        # evita inversão
        if (ano_i, m_i) > (ano_f, m_f):
            ano_i, m_i, ano_f, m_f = ano_f, m_f, ano_i, m_i

        months = _month_range(ano_i, m_i, ano_f, m_f)
        start_dt = date(ano_i, m_i, 1)
        end_dt = _last_day(ano_f, m_f)

        # EMP filter: múltiplos ?emp=101&emp=1001
        emps_sel = request.args.getlist("emp") or []
        emps_sel = [str(e).strip() for e in emps_sel if str(e).strip()]

        allowed = allowed_emps_fn()  # [] => admin all
        if allowed:
            if emps_sel:
                emps_sel = [e for e in emps_sel if e in allowed]
            emps_disponiveis = allowed
        else:
            # admin all: mostra chips com as EMPs que existirem em vendas (limitado)
            emps_disponiveis = []

        # ===== monta query =====
        resultados = None
        db = SessionLocal()
        try:
            # carrega emps_disponiveis para admin all (somente para UI de chips)
            if not allowed:
                emps_disponiveis = [r[0] for r in db.query(Venda.emp).filter(Venda.emp.isnot(None)).distinct().order_by(Venda.emp.asc()).limit(60).all()]

            do_search = bool(produto_raw or marca_raw or mestre_raw)

            if do_search:
                campo_desc = func.lower(func.trim(func.coalesce(Venda.descricao_norm, Venda.descricao, "")))
                conds = [Venda.movimento >= start_dt, Venda.movimento <= end_dt]

                # Escopo EMP
                if allowed:
                    conds.append(Venda.emp.in_(allowed))
                if emps_sel:
                    conds.append(Venda.emp.in_(emps_sel))

                # produto (começa com + termos em sequência)
                if produto_raw:
                    qn = _norm(produto_raw)
                    tokens = [t for t in qn.split(" ") if t]
                    if tokens:
                        pattern = tokens[0] + "%" + "%".join(tokens[1:]) + "%"
                        conds.append(campo_desc.like(pattern))

                # mestre (opcional) - começa com (prefixo). Combina com produto/descrição quando ambos informados.
                if mestre_raw:
                    mn = _norm(mestre_raw)
                    campo_mestre = func.lower(func.trim(func.coalesce(Venda.mestre, "")))
                    conds.append(campo_mestre.like(mn + "%"))

                # marca (opcional) - começa com
                if marca_raw:
                    mn = _norm(marca_raw)
                    campo_marca = func.lower(func.trim(func.coalesce(Venda.marca, "")))
                    conds.append(campo_marca.like(mn + "%"))

                signed_qty = case(
                    (Venda.mov_tipo_movto.in_(["DS", "CA"]), -func.coalesce(Venda.qtdade_vendida, 0.0)),
                    else_=func.coalesce(Venda.qtdade_vendida, 0.0),
                )
                q_year = func.extract("year", Venda.movimento).label("ano")
                q_month = func.extract("month", Venda.movimento).label("mes")

                rows = (
                    db.query(
                        Venda.emp.label("emp"),
                        Usuario.username.label("vendedor"),
                        q_year,
                        q_month,
                        func.coalesce(func.sum(signed_qty), 0.0).label("qtd"),
                    )
                    # Somente vendedores cadastrados (usuarios.role='vendedor') e vinculados à EMP via usuario_emps (ativo=TRUE)
                    .join(
                        Usuario,
                        func.lower(func.trim(func.coalesce(Venda.vendedor, "")))
                        == func.lower(func.trim(func.coalesce(Usuario.username, ""))),
                    )
                    .join(
                        UsuarioEmp,
                        (UsuarioEmp.usuario_id == Usuario.id)
                        & (UsuarioEmp.emp == Venda.emp)
                        & (UsuarioEmp.ativo.is_(True)),
                    )
                    .filter(Usuario.role == "vendedor")
                    .filter(*conds)
                    .group_by(Venda.emp, Usuario.username, q_year, q_month)
                    .order_by(Venda.emp.asc(), Usuario.username.asc(), q_year.asc(), q_month.asc())
                    .all()
                )

                # pivot
                meses_meta = [{"key": f"{int(y):04d}-{int(m):02d}", "label": f"{int(m):02d}/{int(y):04d}"} for (y, m) in months]
                by_key = {}
                for r in rows:
                    emp = (r.emp or "").strip()
                    vend = (r.vendedor or "").strip()
                    key = (emp, vend)
                    mm_key = f"{int(r.ano):04d}-{int(r.mes):02d}"
                    by_key.setdefault(key, {})[mm_key] = float(r.qtd or 0.0)

                linhas = []
                total_qtd_all = 0.0
                for (emp, vend), mp in by_key.items():
                    total = 0.0
                    count_mes = 0
                    for mm in meses_meta:
                        v = float(mp.get(mm["key"], 0.0))
                        total += v
                        if abs(v) > 1e-9:
                            count_mes += 1
                    media = (total / count_mes) if count_mes else 0.0
                    total_qtd_all += total
                    linhas.append({
                        "emp": emp,
                        "vendedor": vend,
                        "por_mes": mp,
                        "total": total,
                        "media": media,
                    })

                # ordena por total desc
                linhas.sort(key=lambda x: (x["emp"], x["vendedor"]))

                media_mensal_global = 0.0
                # média global: total / meses com resultado (considera meses com qualquer dado em qualquer linha)
                meses_com_dado = set()
                for l in linhas:
                    for k, v in l["por_mes"].items():
                        if abs(float(v)) > 1e-9:
                            meses_com_dado.add(k)
                if meses_com_dado:
                    media_mensal_global = total_qtd_all / len(meses_com_dado)

                resultados = {
                    "meses": meses_meta,
                    "linhas": linhas,
                    "total_qtd": total_qtd_all,
                    "media_mensal": media_mensal_global,
                }

                # ===== export excel =====
                if (request.args.get("export") or "").strip() == "1":
                    import openpyxl
                    from openpyxl.utils import get_column_letter
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Vendas por Produto"

                    headers = ["EMP", "Vendedor"] + [m["label"] for m in meses_meta] + ["Total", "Média"]
                    ws.append(headers)

                    for l in linhas:
                        row = [l["emp"], l["vendedor"]]
                        for m in meses_meta:
                            row.append(float(l["por_mes"].get(m["key"], 0.0)))
                        row += [float(l["total"]), float(l["media"])]
                        ws.append(row)

                    # ajustes simples
                    for col in range(1, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col)].width = 14 if col > 2 else 18
                    ws.freeze_panes = "C2"

                    bio = BytesIO()
                    wb.save(bio)
                    bio.seek(0)
                    filename = f"vendas_produto_{today.strftime('%Y%m%d')}.xlsx"
                    return send_file(
                        bio,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        as_attachment=True,
                        download_name=filename,
                    )

            filtros = {
                "produto": produto_raw,
                "marca": marca_raw,
                "mestre": mestre_raw,
                "mes_ini": mes_ini,
                "mes_fim": mes_fim,
                "emps_sel": emps_sel,
                "emps_disponiveis": emps_disponiveis,
            }

            return render_template(
                "operacoes_vendas_produto.html",
                role=role_fn(),
                emp=emp_fn(),
                filtros=filtros,
                meses_select=meses_select,
                resultados=resultados,
            )
        finally:
            try:
                db.close()
            except Exception:
                pass

    @login_required_fn
    def api_produtos_suggest():
        """Sugestões rápidas de descrição (typeahead) para a pesquisa de produtos."""
        from db import SessionLocal, Venda, Usuario, UsuarioEmp
        from sqlalchemy import func
        from flask import jsonify
        import unicodedata, re
        from datetime import date

        def _norm(s: str) -> str:
            s = (s or "").strip()
            s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
            s = re.sub(r"\s+", " ", s).strip().lower()
            return s

        q = (request.args.get("q") or "").strip()
        if len(q) < 3:
            return jsonify({"items": []})

        qn = _norm(q)
        tokens = [t for t in qn.split(" ") if t]
        if not tokens:
            return jsonify({"items": []})

        pattern = tokens[0] + "%" + "%".join(tokens[1:]) + "%"

        marca_raw = (request.args.get("marca") or "").strip()
        emps_sel = request.args.getlist("emp") or []
        emps_sel = [str(e).strip() for e in emps_sel if str(e).strip()]

        allowed = allowed_emps_fn()
        db = SessionLocal()
        try:
            campo_desc = func.lower(func.trim(func.coalesce(Venda.descricao_norm, Venda.descricao, "")))
            conds = [campo_desc.like(pattern)]
            if allowed:
                conds.append(Venda.emp.in_(allowed))
            if emps_sel:
                conds.append(Venda.emp.in_(emps_sel))
            if marca_raw:
                mn = _norm(marca_raw)
                campo_marca = func.lower(func.trim(func.coalesce(Venda.marca, "")))
                conds.append(campo_marca.like(mn + "%"))

            # tenta usar o ano atual como prioridade
            today = date.today()
            subq = (
                db.query(func.coalesce(Venda.descricao, Venda.mestre).label("d"))
                .filter(*conds)
                .filter(Venda.movimento >= date(today.year, 1, 1))
                .distinct()
                .subquery()
            )
            # Postgres exige que expressões do ORDER BY apareçam no SELECT quando usamos DISTINCT.
            # Por isso fazemos o DISTINCT em subquery e ordenamos fora.
            rows = (
                db.query(subq.c.d)
                .order_by(func.length(subq.c.d).asc())
                .limit(18)
                .all()
            )
            items = [r.d for r in rows if r.d]
            return jsonify({"items": items})
        finally:
            try:
                db.close()
            except Exception:
                pass

    # Registrar rotas (endpoints preservados)
    app.add_url_rule(
        "/operacoes/vendas-produtos",
        endpoint="operacoes_vendas_produto",
        view_func=operacoes_vendas_produto,
        methods=["GET"],
    )
    app.add_url_rule(
        "/operacoes/api/produtos_suggest",
        endpoint="api_produtos_suggest",
        view_func=api_produtos_suggest,
        methods=["GET"],
    )
