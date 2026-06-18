"""Importação de vendas (Render-friendly) com deduplicação configurável.

IMPORTANTE:
- Para o seu caso, o MOV_TIPO_MOVTO (ex: 'DS' / 'CA' / 'OA') DEVE fazer parte
  da chave de duplicidade, senão DS/CA "somem" do cálculo.
- Recomendado: criar um índice UNIQUE no Postgres com as colunas da chave,
  para o ON CONFLICT funcionar.

Suporta:
  - .csv (recomendado p/ arquivos grandes) -> streaming por chunks
  - .xlsx (arquivos menores)              -> read_only + batches

API:
  importar_planilha(filepath, modo="ignorar_duplicados", chave=..., ...)

chaves disponíveis:
  - mestre_vendedor_nota_emp           (antiga)
  - mestre_movimento_vendedor_nota_emp      (inclui MOVIMENTO/data)
  - mestre_movimento_vendedor_nota_tipo_emp (inclui MOVIMENTO + MOV_TIPO_MOVTO)  <-- RECOMENDADA (usa índice com MARCA no banco)
  - mestre_marca_movimento_vendedor_nota_tipo_emp (explicita MARCA no nome; equivalente ao índice atual)
  - mestre_movimento_vendedor_nota_tipo     (sem EMP, se você não usa EMP na chave)
"""

from __future__ import annotations

import datetime as dt
import os
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from sqlalchemy.dialects.postgresql import insert as pg_insert

from db import SessionLocal, Venda
from dashboard_cache import refresh_dashboard_cache
from sv_utils import MOVIMENTOS_VENDA, MOVIMENTOS_NEGATIVOS, classificar_movimento


REQUIRED_COLS = [
    "MESTRE",
    "MARCA",
    "MOVIMENTO",
    "MOV_TIPO_MOVTO",
    "VENDEDOR",
    "NOTA",
    "EMP",
    "UNIT",
    "DES",
    "QTDADE_VENDIDA",
    "VALOR_TOTAL",
]


OPTIONAL_COLS = [
    "DESCRICAO",
    "RAZAO",
    "CIDADE",
    "CNPJ_CPF",
]


def _norm_cols(cols: List[Any]) -> List[str]:
    return [str(c).strip().upper() if c is not None else "" for c in cols]


def _to_date(value: Any) -> dt.date:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        raise ValueError("MOVIMENTO vazio")
    if isinstance(value, dt.date) and not isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.datetime):
        return value.date()
    return pd.to_datetime(value).date()


def _to_float(value: Any) -> Optional[float]:
    """Converte números vindos do Excel/CSV com segurança.

    - Se vier numérico (int/float), retorna float direto.
    - Se vier string no formato brasileiro (contém vírgula), remove separador de milhar '.' e troca ',' por '.'
    - Se vier string no formato internacional (somente '.' como decimal), NÃO remove '.' (evita multiplicar por 10/100)
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value)

        if isinstance(value, str):
            s = value.strip()
            if not s:
                return None
            if "," in s:
                s = s.replace(".", "").replace(",", ".")
                return float(s)
            return float(s)

        return float(value)
    except Exception:
        return None



def _sanitize_qtd_valor(qtd_raw: Any, valor_raw: Any) -> tuple[float, float, bool]:
    """Normaliza quantidade/valor de venda importada.

    Regra de segurança: linha com QTDADE_VENDIDA vazia, nula ou 0 não pode gerar
    faturamento, campanha, meta, mix ou qualquer cálculo operacional.
    Mantemos a linha importável para rastreabilidade/atualização, mas zeramos
    o valor_total dela.
    """
    qtd = _to_float(qtd_raw)
    valor = _to_float(valor_raw) or 0.0
    try:
        qtd_f = float(qtd or 0.0)
    except Exception:
        qtd_f = 0.0
    if abs(qtd_f) <= 1e-12:
        return 0.0, 0.0, True
    return qtd_f, float(valor or 0.0), False


def _norm_str(value: Any) -> Optional[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip()
    return s if s else None



def _norm_text(value: Any) -> Optional[str]:
    """Normaliza texto para comparações estáveis (lowercase, sem acento, espaços colapsados)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip()
    if not s:
        return None
    # remove acentos
    import unicodedata, re
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s if s else None


def _client_id_norm(cnpj_cpf: Any, razao: Any) -> Optional[str]:
    """Gera um identificador estável do cliente.

    Preferência: CNPJ/CPF (normalizado, só dígitos). Fallback: hash da razão social normalizada.
    """
    import hashlib, re
    c = _norm_str(cnpj_cpf)
    if c:
        digits = re.sub(r"\D+", "", c)
        if digits:
            return f"doc:{digits}"
    r = _norm_text(razao)
    if r:
        h = hashlib.sha256(r.encode("utf-8")).hexdigest()
        return f"razao:{h}"
    return None


def _month_period_from_date(emp: str, mov: dt.date) -> Tuple[str, int, int]:
    return (str(emp), int(mov.year), int(mov.month))


def _safe_date_key(emp: Any, mov: Any) -> Optional[Tuple[str, dt.date]]:
    """Retorna chave segura (EMP, data) para reprocessamento automático."""
    try:
        e = _norm_str(emp)
        d = _to_date(mov)
        if not e or not d:
            return None
        return (str(e), d)
    except Exception:
        return None


def _delete_affected_dates(datas: List[Tuple[str, dt.date]]) -> Dict[str, Any]:
    """Remove vendas já existentes apenas para os recortes exatos do arquivo.

    Segurança: em vez de apagar o mês inteiro, apaga somente EMP + DATA presentes
    na planilha. Isso evita perder vendas de outros dias quando a importação diária
    contém apenas parte da competência mensal.
    """
    unique_dates = sorted({(str(emp), mov) for emp, mov in datas if emp and mov})
    if not unique_dates:
        return {"ok": True, "tipo": "EMP_DATA", "recortes": 0, "linhas_apagadas": 0, "detalhes": []}

    db = SessionLocal()
    detalhes = []
    total = 0
    try:
        for emp, mov in unique_dates:
            apagadas = (
                db.query(Venda)
                .filter(Venda.emp == str(emp))
                .filter(Venda.movimento == mov)
                .delete(synchronize_session=False)
            )
            apagadas = int(apagadas or 0)
            total += apagadas
            detalhes.append({"emp": str(emp), "data": mov.isoformat(), "linhas_apagadas": apagadas})
        db.commit()
        return {
            "ok": True,
            "tipo": "EMP_DATA",
            "recortes": len(unique_dates),
            "linhas_apagadas": int(total),
            "detalhes": detalhes,
        }
    except Exception:
        db.rollback()
        raise
    finally:
        try:
            db.close()
        except Exception:
            pass

def _conflict_cols_from_key(chave: str) -> List[str]:
    """Mapeia o nome da chave para colunas do banco.

    IMPORTANTE: estas colunas precisam bater 100% com o índice UNIQUE existente,
    senão o Postgres retorna:
      'there is no unique or exclusion constraint matching the ON CONFLICT specification'

    Seu índice atual (vendas_unique_import) é:
      (mestre, marca, vendedor, movimento, mov_tipo_movto, nota, emp)
    """
    # nomes da tabela/ORM: mestre, marca, movimento, vendedor, nota, emp, mov_tipo_movto
    if chave == "mestre_marca_movimento_vendedor_nota_tipo_emp":
        return ["mestre", "marca", "vendedor", "movimento", "mov_tipo_movto", "nota", "emp"]

    # compatibilidade com a chave antiga (sem 'marca' no nome), mas inclui 'marca' pois é o índice do banco
    if chave == "mestre_movimento_vendedor_nota_tipo_emp":
        return ["mestre", "marca", "vendedor", "movimento", "mov_tipo_movto", "nota", "emp"]

    if chave == "mestre_movimento_vendedor_nota_tipo":
        return ["mestre", "marca", "vendedor", "movimento", "mov_tipo_movto", "nota"]

    if chave == "mestre_movimento_vendedor_nota_emp":
        return ["mestre", "marca", "vendedor", "movimento", "nota", "emp"]

    # fallback antigo: manter, mas alinhado ao índice principal quando possível
    return ["mestre", "marca", "vendedor", "movimento", "nota", "emp"]


def _build_stmt(records: List[dict], modo: str, conflict_cols: List[str]):
    stmt = pg_insert(Venda.__table__).values(records)
    if modo == "atualizar":
        update_cols = {
            "marca": stmt.excluded.marca,
            "movimento": stmt.excluded.movimento,
            "mov_tipo_movto": stmt.excluded.mov_tipo_movto,
            "unit": stmt.excluded.unit,
            "des": stmt.excluded.des,
            "qtdade_vendida": stmt.excluded.qtdade_vendida,
            "valor_total": stmt.excluded.valor_total,
            "emp": stmt.excluded.emp,
            "nota": stmt.excluded.nota,
            "vendedor": stmt.excluded.vendedor,
            "mestre": stmt.excluded.mestre,
            "descricao": stmt.excluded.descricao,
            "razao": stmt.excluded.razao,
            "cidade": stmt.excluded.cidade,
            "cnpj_cpf": stmt.excluded.cnpj_cpf,
            "descricao_norm": stmt.excluded.descricao_norm,
            "razao_norm": stmt.excluded.razao_norm,
            "cidade_norm": stmt.excluded.cidade_norm,
            "cliente_id_norm": stmt.excluded.cliente_id_norm,
        }
        return stmt.on_conflict_do_update(index_elements=conflict_cols, set_=update_cols)
    return stmt.on_conflict_do_nothing(index_elements=conflict_cols)



def scan_metadata_xlsx(filepath: str, check_required: bool = False) -> Dict[str, Any]:
    """Varre o XLSX (read_only) para descobrir EMP(s), datas e competência(s) presentes.

    Retorna:
      {"ok": True, "emps": [...], "datas": [(emp, date)], "periodos": [(emp, ano, mes)]}
    """
    from openpyxl import load_workbook
    wb = None
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return {"ok": False, "msg": "Planilha vazia."}
        cols = _norm_cols(list(header))
        col_index = {c: i for i, c in enumerate(cols)}
        required = REQUIRED_COLS if check_required else ["EMP", "MOVIMENTO"]
        missing = [c for c in required if c not in col_index]
        if missing:
            return {"ok": False, "msg": "Colunas faltando para varredura.", "faltando": missing, "lidas": cols}
        periodos = set()
        datas = set()
        emps = set()
        for r in rows:
            key = _safe_date_key(r[col_index["EMP"]], r[col_index["MOVIMENTO"]])
            if not key:
                continue
            emp, mov = key
            emps.add(str(emp))
            datas.add((str(emp), mov))
            periodos.add(_month_period_from_date(str(emp), mov))
        return {
            "ok": True,
            "emps": sorted(emps),
            "datas": sorted(datas),
            "periodos": sorted(periodos),
            "multi_competencia": (len({(a, m) for _, a, m in periodos}) > 1),
        }
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def scan_metadata_csv(filepath: str, check_required: bool = False, csv_chunksize: int = 3000) -> Dict[str, Any]:
    """Varre CSV para descobrir EMP(s), datas e competência(s) presentes sem importar."""
    periodos = set()
    datas = set()
    emps = set()
    first = True
    seen_cols = []
    required = REQUIRED_COLS if check_required else ["EMP", "MOVIMENTO"]
    for chunk in pd.read_csv(filepath, chunksize=csv_chunksize, dtype=str, encoding_errors="ignore"):
        chunk.columns = _norm_cols(list(chunk.columns))
        if first:
            seen_cols = list(chunk.columns)
            missing = [c for c in required if c not in chunk.columns]
            if missing:
                return {"ok": False, "msg": "Colunas faltando para varredura.", "faltando": missing, "lidas": seen_cols}
            first = False
        for _emp, _mov in chunk[["EMP", "MOVIMENTO"]].dropna().itertuples(index=False, name=None):
            key = _safe_date_key(_emp, _mov)
            if not key:
                continue
            emp, mov = key
            emps.add(str(emp))
            datas.add((str(emp), mov))
            periodos.add(_month_period_from_date(str(emp), mov))
    if first:
        return {"ok": False, "msg": "Arquivo CSV vazio."}
    return {
        "ok": True,
        "emps": sorted(emps),
        "datas": sorted(datas),
        "periodos": sorted(periodos),
        "multi_competencia": (len({(a, m) for _, a, m in periodos}) > 1),
    }


def importar_planilha(
    filepath: str,
    modo: str = "ignorar_duplicados",
    chave: str = "mestre_movimento_vendedor_nota_tipo_emp",
    batch_size: int = 300,
    csv_chunksize: int = 3000,
    xlsx_max_mb: int = 12,
    reprocessar_competencia: bool = False,
    atualizar_cache_dashboard: bool | None = None,
) -> Dict[str, Any]:
    """Importa vendas no banco com inserção em lotes.

    Por padrão, NÃO recalcula dashboard_cache dentro da requisição de importação.
    Esse recálculo pode levar minutos e derrubar o worker do Render. Para ativar
    pontualmente, use atualizar_cache_dashboard=True ou a variável
    IMPORT_REFRESH_DASHBOARD_CACHE_SYNC=1.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if atualizar_cache_dashboard is None:
        atualizar_cache_dashboard = (os.getenv("IMPORT_REFRESH_DASHBOARD_CACHE_SYNC", "0").strip().lower() in {"1", "true", "sim", "yes", "on"})

    conflict_cols = _conflict_cols_from_key(chave)
    reprocess_info: Dict[str, Any] = {"ok": True, "tipo": "EMP_DATA", "recortes": 0, "linhas_apagadas": 0, "detalhes": []}

    if ext == ".xlsx":
        try:
            mb = os.path.getsize(filepath) / (1024 * 1024)
            if mb > float(xlsx_max_mb):
                return {
                    "ok": False,
                    "msg": f"Arquivo XLSX grande ({mb:.1f}MB). Para evitar erro no Render, exporte para CSV e importe o CSV.",
                    "inseridas": 0,
                    "atualizadas": 0,
                    "ignoradas": 0,
                    "erros_linha": 0,
                }
        except Exception:
            pass

        if reprocessar_competencia:
            meta = scan_metadata_xlsx(filepath, check_required=True)
            if not meta.get("ok"):
                return {"ok": False, "msg": meta.get("msg", "Falha ao varrer planilha."), "faltando": meta.get("faltando"), "lidas": meta.get("lidas")}
            reprocess_info = _delete_affected_dates(meta.get("datas") or [])

        from openpyxl import load_workbook

        wb = None
        db = SessionLocal()
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)

            header = next(rows, None)
            if not header:
                return {"ok": False, "msg": "Planilha vazia.", "inseridas": 0, "atualizadas": 0, "ignoradas": 0, "erros_linha": 0}

            cols = _norm_cols(list(header))
            col_index = {c: i for i, c in enumerate(cols)}
            missing = [c for c in REQUIRED_COLS if c not in col_index]
            if missing:
                return {"ok": False, "msg": "Colunas faltando.", "faltando": missing, "lidas": cols}

            total_linhas = 0
            validas = 0
            erros_linha = 0
            inseridas = 0
            atualizadas = 0
            batch: List[dict] = []
            affected_periods = set()
            total_bruto = 0.0
            total_cancelamentos = 0.0
            total_devolucoes = 0.0
            total_ignorados_movimento = 0.0
            cancelamento_linhas = 0
            devolucao_linhas = 0
            movimentos_ignorados_linhas = 0
            linhas_qtd_zero = 0

            for r in rows:
                total_linhas += 1
                try:
                    def get(col: str):
                        return r[col_index[col]] if col in col_index else None

                    mestre = _norm_str(get("MESTRE"))
                    vendedor = _norm_str(get("VENDEDOR"))
                    vendedor = vendedor.upper() if vendedor else None
                    if not mestre or not vendedor:
                        erros_linha += 1
                        continue

                    nota = _norm_str(get("NOTA"))
                    emp = _norm_str(get("EMP"))
                    mov = _to_date(get("MOVIMENTO"))
                    # registra período para atualizar cache depois
                    if emp:
                        affected_periods.add((str(emp), int(mov.year), int(mov.month)))
                    mov_tipo = (_norm_str(get("MOV_TIPO_MOVTO")) or "").upper()
                    if not mov_tipo:
                        erros_linha += 1
                        continue

                    qtd_importada, valor_importado, qtd_zero_calc = _sanitize_qtd_valor(get("QTDADE_VENDIDA"), get("VALOR_TOTAL"))
                    if qtd_zero_calc:
                        linhas_qtd_zero += 1

                    rec = {
                        "mestre": mestre,
                        "marca": _norm_str(get("MARCA")),
                        "movimento": mov,
                        "mov_tipo_movto": mov_tipo,
                        "vendedor": vendedor,
                        "nota": nota,
                        "emp": emp,
                        "unit": _to_float(get("UNIT")),
                        "des": _to_float(get("DES")),
                        "qtdade_vendida": qtd_importada,
                        "valor_total": valor_importado,
                        "descricao": _norm_str(get("DESCRICAO")),
                        "razao": _norm_str(get("RAZAO")),
                        "cidade": _norm_str(get("CIDADE")),
                        "cnpj_cpf": _norm_str(get("CNPJ_CPF")),
                        "descricao_norm": _norm_text(get("DESCRICAO")),
                        "razao_norm": _norm_text(get("RAZAO")),
                        "cidade_norm": _norm_text(get("CIDADE")),
                        "cliente_id_norm": _client_id_norm(get("CNPJ_CPF"), get("RAZAO")),
                    }

                    batch.append(rec)
                    validas += 1

                    try:
                        v = float(rec.get('valor_total') or 0.0)
                    except Exception:
                        v = 0.0
                    classe_mov = classificar_movimento(mov_tipo)
                    if classe_mov == "VENDA":
                        total_bruto += v
                    elif classe_mov == "CANCELAMENTO":
                        total_cancelamentos += abs(v)
                        cancelamento_linhas += 1
                    elif classe_mov == "DEVOLUCAO":
                        total_devolucoes += abs(v)
                        devolucao_linhas += 1
                    else:
                        total_ignorados_movimento += abs(v)
                        movimentos_ignorados_linhas += 1

                    if len(batch) >= batch_size:
                        stmt = _build_stmt(batch, modo, conflict_cols)
                        res = db.execute(stmt)
                        db.commit()
                        if modo == "atualizar":
                            atualizadas += int(res.rowcount or 0)
                        else:
                            inseridas += int(res.rowcount or 0)
                        batch.clear()

                except Exception:
                    erros_linha += 1
                    continue

            if batch:
                stmt = _build_stmt(batch, modo, conflict_cols)
                res = db.execute(stmt)
                db.commit()
                if modo == "atualizar":
                    atualizadas += int(res.rowcount or 0)
                else:
                    inseridas += int(res.rowcount or 0)
                batch.clear()

            # rowcount = quantos foram inseridos/atualizados; o resto são ignorados (duplicados)
            efetivadas = atualizadas if modo == "atualizar" else inseridas
            ignoradas = max(validas - efetivadas, 0)

            # Atualiza cache (dashboard_cache) somente quando solicitado.
            # Não fazer isso em toda importação evita POST de 2-3 minutos e 502 no Render.
            cache_info = []
            cache_skipped = True
            if atualizar_cache_dashboard:
                cache_skipped = False
                try:
                    for e,a,m in sorted(affected_periods):
                        cache_info.append({"emp": e, "ano": a, "mes": m, **refresh_dashboard_cache(e,a,m)})
                except Exception:
                    # Não bloqueia importação se o cache falhar
                    cache_info = []

            return {
                "ok": True,
                "msg": "Importação finalizada.",
                "total_linhas": int(total_linhas),
                "validas": int(validas),
                "inseridas": int(inseridas),
                "atualizadas": int(atualizadas),
                "ignoradas": int(ignoradas),
                "erros_linha": int(erros_linha),
                "modo": modo,
                "chave": chave,
                "batch_size": int(batch_size),
                "conflict_cols": conflict_cols,
                "cache": cache_info,
                "cache_dashboard_atualizado": bool(atualizar_cache_dashboard),
                "cache_dashboard_pendente": bool(cache_skipped and affected_periods),
                "total_bruto": float(total_bruto),
                "total_cancelamentos": float(total_cancelamentos),
                "total_devolucoes": float(total_devolucoes),
                "total_ca": float(total_cancelamentos + total_devolucoes),  # compatibilidade com telas antigas
                "total_movimentos_ignorados": float(total_ignorados_movimento),
                "total_liquido": float(total_bruto - total_cancelamentos - total_devolucoes),
                "cancelamento_linhas": int(cancelamento_linhas),
                "devolucao_linhas": int(devolucao_linhas),
                "ca_linhas": int(cancelamento_linhas + devolucao_linhas),  # compatibilidade
                "movimentos_ignorados_linhas": int(movimentos_ignorados_linhas),
                "movimentos_venda_considerados": list(MOVIMENTOS_VENDA),
                "movimentos_negativos_considerados": list(MOVIMENTOS_NEGATIVOS),
                "linhas_qtd_zero_ignoradas_calculo": int(linhas_qtd_zero),
                "reprocessar": bool(reprocessar_competencia),
                "reprocessamento": reprocess_info,
                "linhas_reprocessadas_apagadas": int((reprocess_info or {}).get("linhas_apagadas") or 0),
                "affected_periods": sorted(list(affected_periods)),
            }

        finally:
            try:
                db.close()
            except Exception:
                pass
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

    # CSV (recomendado)
    if ext != ".csv":
        return {"ok": False, "msg": "Formato não suportado. Use .csv (recomendado) ou .xlsx."}

    total_linhas = 0
    validas = 0
    erros_linha = 0
    inseridas = 0
    atualizadas = 0
    affected_periods = set()
    total_bruto = 0.0
    total_cancelamentos = 0.0
    total_devolucoes = 0.0
    total_ignorados_movimento = 0.0
    cancelamento_linhas = 0
    devolucao_linhas = 0
    movimentos_ignorados_linhas = 0
    linhas_qtd_zero = 0

    if reprocessar_competencia:
        meta = scan_metadata_csv(filepath, check_required=True, csv_chunksize=csv_chunksize)
        if not meta.get("ok"):
            return {"ok": False, "msg": meta.get("msg", "Falha ao varrer CSV."), "faltando": meta.get("faltando"), "lidas": meta.get("lidas")}
        reprocess_info = _delete_affected_dates(meta.get("datas") or [])

    db = SessionLocal()
    try:
        for chunk in pd.read_csv(filepath, chunksize=csv_chunksize, dtype=str, encoding_errors="ignore"):
            chunk.columns = _norm_cols(list(chunk.columns))
            missing = [c for c in REQUIRED_COLS if c not in chunk.columns]
            if missing:
                return {"ok": False, "msg": "Colunas faltando.", "faltando": missing, "lidas": list(chunk.columns)}

            chunk["MOVIMENTO"] = pd.to_datetime(chunk["MOVIMENTO"], errors="coerce").dt.date
            try:
                # períodos (emp, ano, mes) presentes no chunk
                for _emp, _mov in chunk[["EMP","MOVIMENTO"]].dropna().itertuples(index=False, name=None):
                    _e = str(_emp).strip()
                    if not _e or not _mov:
                        continue
                    affected_periods.add((_e, int(_mov.year), int(_mov.month)))
            except Exception:
                pass

            records: List[dict] = []
            for _, row in chunk.iterrows():
                total_linhas += 1
                try:
                    mestre = _norm_str(row.get("MESTRE"))
                    vendedor = _norm_str(row.get("VENDEDOR"))
                    vendedor = vendedor.upper() if vendedor else None
                    if not mestre or not vendedor:
                        erros_linha += 1
                        continue

                    nota = _norm_str(row.get("NOTA"))
                    emp = _norm_str(row.get("EMP"))
                    mov = row.get("MOVIMENTO")
                    if mov is None or pd.isna(mov):
                        erros_linha += 1
                        continue

                    mov_tipo = (_norm_str(row.get("MOV_TIPO_MOVTO")) or "").upper()
                    if not mov_tipo:
                        erros_linha += 1
                        continue

                    qtd_importada, valor_importado, qtd_zero_calc = _sanitize_qtd_valor(row.get("QTDADE_VENDIDA"), row.get("VALOR_TOTAL"))
                    if qtd_zero_calc:
                        linhas_qtd_zero += 1

                    rec = {
                        "mestre": mestre,
                        "marca": _norm_str(row.get("MARCA")),
                        "movimento": mov,
                        "mov_tipo_movto": mov_tipo,
                        "vendedor": vendedor,
                        "nota": nota,
                        "emp": emp,
                        "unit": _to_float(row.get("UNIT")),
                        "des": _to_float(row.get("DES")),
                        "qtdade_vendida": qtd_importada,
                        "valor_total": valor_importado,
                        "descricao": _norm_str(row.get("DESCRICAO")),
                        "razao": _norm_str(row.get("RAZAO")),
                        "cidade": _norm_str(row.get("CIDADE")),
                        "cnpj_cpf": _norm_str(row.get("CNPJ_CPF")),
                        "descricao_norm": _norm_text(row.get("DESCRICAO")),
                        "razao_norm": _norm_text(row.get("RAZAO")),
                        "cidade_norm": _norm_text(row.get("CIDADE")),
                        "cliente_id_norm": _client_id_norm(row.get("CNPJ_CPF"), row.get("RAZAO")),
                        "descricao": _norm_str(row.get("DESCRICAO")),
                        "razao": _norm_str(row.get("RAZAO")),
                        "cidade": _norm_str(row.get("CIDADE")),
                        "cnpj_cpf": _norm_str(row.get("CNPJ_CPF")),
                        "descricao_norm": _norm_text(row.get("DESCRICAO")),
                        "razao_norm": _norm_text(row.get("RAZAO")),
                        "cidade_norm": _norm_text(row.get("CIDADE")),
                        "cliente_id_norm": _client_id_norm(row.get("CNPJ_CPF"), row.get("RAZAO")),
                    }
                    records.append(rec)
                    validas += 1

                    try:
                        v = float(rec.get('valor_total') or 0.0)
                    except Exception:
                        v = 0.0
                    classe_mov = classificar_movimento(mov_tipo)
                    if classe_mov == "VENDA":
                        total_bruto += v
                    elif classe_mov == "CANCELAMENTO":
                        total_cancelamentos += abs(v)
                        cancelamento_linhas += 1
                    elif classe_mov == "DEVOLUCAO":
                        total_devolucoes += abs(v)
                        devolucao_linhas += 1
                    else:
                        total_ignorados_movimento += abs(v)
                        movimentos_ignorados_linhas += 1

                    if len(records) >= batch_size:
                        stmt = _build_stmt(records, modo, conflict_cols)
                        res = db.execute(stmt)
                        db.commit()
                        if modo == "atualizar":
                            atualizadas += int(res.rowcount or 0)
                        else:
                            inseridas += int(res.rowcount or 0)
                        records.clear()

                except Exception:
                    erros_linha += 1
                    continue

            if records:
                stmt = _build_stmt(records, modo, conflict_cols)
                res = db.execute(stmt)
                db.commit()
                if modo == "atualizar":
                    atualizadas += int(res.rowcount or 0)
                else:
                    inseridas += int(res.rowcount or 0)
                records.clear()

    finally:
        db.close()

    efetivadas = atualizadas if modo == "atualizar" else inseridas
    ignoradas = max(validas - efetivadas, 0)

    # Atualiza cache (dashboard_cache) somente quando solicitado.
    cache_info = []
    cache_skipped = True
    if atualizar_cache_dashboard:
        cache_skipped = False
        try:
            for e,a,m in sorted(affected_periods):
                cache_info.append({"emp": e, "ano": a, "mes": m, **refresh_dashboard_cache(e,a,m)})
        except Exception:
            cache_info = []

    return {
        "ok": True,
        "msg": "Importação finalizada.",
        "total_linhas": int(total_linhas),
        "validas": int(validas),
        "inseridas": int(inseridas),
        "atualizadas": int(atualizadas),
        "ignoradas": int(ignoradas),
        "erros_linha": int(erros_linha),
        "modo": modo,
        "chave": chave,
        "batch_size": int(batch_size),
        "csv_chunksize": int(csv_chunksize),
        "conflict_cols": conflict_cols,
        "cache": cache_info,
        "cache_dashboard_atualizado": bool(atualizar_cache_dashboard),
        "cache_dashboard_pendente": bool(cache_skipped and affected_periods),
        "total_bruto": float(total_bruto),
        "total_cancelamentos": float(total_cancelamentos),
        "total_devolucoes": float(total_devolucoes),
        "total_ca": float(total_cancelamentos + total_devolucoes),  # compatibilidade com telas antigas
        "total_movimentos_ignorados": float(total_ignorados_movimento),
        "total_liquido": float(total_bruto - total_cancelamentos - total_devolucoes),
        "cancelamento_linhas": int(cancelamento_linhas),
        "devolucao_linhas": int(devolucao_linhas),
        "ca_linhas": int(cancelamento_linhas + devolucao_linhas),  # compatibilidade
        "movimentos_ignorados_linhas": int(movimentos_ignorados_linhas),
        "movimentos_venda_considerados": list(MOVIMENTOS_VENDA),
        "movimentos_negativos_considerados": list(MOVIMENTOS_NEGATIVOS),
        "linhas_qtd_zero_ignoradas_calculo": int(linhas_qtd_zero),
        "reprocessar": bool(reprocessar_competencia),
        "reprocessamento": reprocess_info,
        "linhas_reprocessadas_apagadas": int((reprocess_info or {}).get("linhas_apagadas") or 0),
        "affected_periods": sorted(list(affected_periods)),
    }
