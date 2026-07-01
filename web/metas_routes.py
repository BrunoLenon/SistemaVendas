# -*- coding: utf-8 -*-
"""Rotas oficiais do novo modulo de Metas.

Substitui a estrutura antiga mantendo os endpoints publicos:
- /admin/metas
- /metas
"""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
import unicodedata

import pandas as pd
from flask import flash, redirect, render_template, request, send_file, session, url_for
from sqlalchemy import func, text

from auth_helpers import _allowed_emps, _emp, _login_required, _role
from db import (
    Emp,
    MetaBaseManual,
    MetaEscala,
    MetaMarca,
    MetaMargemVendedor,
    MetaPrograma,
    MetaProgramaEmp,
    SessionLocal,
    Usuario,
    ensure_metas_lojas_schema,
)
from metas_helpers import (
    calcular_meta,
    get_active_emps,
    get_meta_emps,
    get_meta_escalas,
    get_meta_marcas,
    get_margem_vendedor,
    get_gerentes_para_metas,
    get_mecanicos_para_metas,
    get_participantes_para_meta,
    get_pessoas_para_metas,
    get_vendedores_para_metas,
    is_meta_gerente,
    is_meta_mecanico,
    metas_ativas_periodo,
    montar_resultados_periodo,
    normalize_emp,
    normalize_text,
    query_valor_mecanico_mes,
    query_valor_mes,
    upsert_base_manual,
)

TIPOS_META = ("CRESCIMENTO", "MIX", "SHARE_MARCA", "MECANICO_FATURAMENTO")


def register_metas_routes(app) -> None:
    app.add_url_rule("/metas", endpoint="metas", view_func=metas, methods=["GET"])
    app.add_url_rule("/admin/metas", endpoint="admin_metas", view_func=admin_metas, methods=["GET"])
    app.add_url_rule("/admin/metas/criar", endpoint="admin_metas_criar", view_func=admin_metas_criar, methods=["POST"])
    app.add_url_rule("/admin/metas/toggle/<int:meta_id>", endpoint="admin_metas_toggle", view_func=admin_metas_toggle, methods=["POST"])
    app.add_url_rule("/admin/metas/excluir/<int:meta_id>", endpoint="admin_metas_excluir", view_func=admin_metas_excluir, methods=["POST"])
    app.add_url_rule("/admin/metas/emps/<int:meta_id>", endpoint="admin_metas_emps_salvar", view_func=admin_metas_emps_salvar, methods=["POST"])
    app.add_url_rule("/admin/metas/regra/<int:meta_id>", endpoint="admin_metas_regra_salvar", view_func=admin_metas_regra_salvar, methods=["POST"])
    app.add_url_rule("/admin/metas/recalcular", endpoint="admin_metas_recalcular", view_func=admin_metas_recalcular, methods=["POST"])
    app.add_url_rule("/admin/metas/bases/<int:meta_id>", endpoint="admin_meta_bases", view_func=admin_meta_bases, methods=["GET"])
    app.add_url_rule("/admin/metas/bases/<int:meta_id>/salvar", endpoint="admin_meta_bases_salvar", view_func=admin_meta_bases_salvar, methods=["POST"])
    app.add_url_rule("/admin/metas/mecanicos/adicionar", endpoint="admin_metas_mecanico_adicionar", view_func=admin_metas_mecanico_adicionar, methods=["POST"])
    app.add_url_rule("/admin/metas/margens/modelo", endpoint="admin_metas_margens_modelo", view_func=admin_metas_margens_modelo, methods=["GET"])
    app.add_url_rule("/admin/metas/margens/importar", endpoint="admin_metas_margens_importar", view_func=admin_metas_margens_importar, methods=["POST"])


def _safe_int(v, default):
    try:
        return int(v)
    except Exception:
        return default


def _safe_float(v, default=0.0):
    if v is None:
        return default
    raw = str(v).strip()
    if not raw:
        return default
    try:
        s = raw.replace("R$", "").replace("%", "").replace(" ", "").strip()
        # pt-BR: 100.000,00 -> 100000.00 / 0,10 -> 0.10
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        # US decimal simples: 0.10 -> 0.10; milhar puro: 100.000 -> 100000
        elif s.count(".") == 1:
            left, right = s.split(".", 1)
            if len(right) == 3 and len(left) >= 2:
                s = left + right
        elif s.count(".") > 1:
            s = s.replace(".", "")
        return float(s)
    except Exception:
        return default


def _norm_col_name(value: object) -> str:
    raw = str(value or "").strip().upper()
    raw = "".join(ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9]", "", raw)


def _parse_margem_percentual(value, default=None):
    if value is None:
        return default
    raw = str(value).strip()
    if not raw or raw.lower() in ("nan", "none", "null"):
        return default
    raw = raw.replace("%", "").replace(" ", "")
    try:
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        return float(raw)
    except Exception:
        return default


def _format_ptbr_percent(value: float) -> str:
    try:
        return f"{float(value):.2f}".replace(".", ",")
    except Exception:
        return "0,00"

def _margem_minima_efetiva(db, meta_id: int | None, emp: str, vendedor: str, margem_padrao: float = 0.0) -> tuple[float, str]:
    """Retorna (margem mínima efetiva, origem).

    Origem:
    - individual: cadastrada em Bases dos Vendedores
    - padrao: regra geral cadastrada na Meta Crescimento
    - vazio: sem exigência de margem
    """
    margem_padrao = float(margem_padrao or 0.0)
    if meta_id:
        try:
            base = (
                db.query(MetaBaseManual)
                .filter(
                    MetaBaseManual.meta_id == int(meta_id),
                    MetaBaseManual.emp == normalize_emp(emp),
                    MetaBaseManual.vendedor == normalize_text(vendedor),
                )
                .first()
            )
            margem_individual = float(getattr(base, "margem_percentual", 0.0) or 0.0) if base else 0.0
            if margem_individual > 0:
                return margem_individual, "individual"
        except Exception:
            pass
    if margem_padrao > 0:
        return margem_padrao, "padrao"
    return 0.0, ""


def _tipo_label(tipo: str) -> str:
    tipo = normalize_text(tipo)
    if tipo == "CRESCIMENTO":
        return "Crescimento"
    if tipo == "MIX":
        return "Mix de Itens Únicos"
    if tipo == "SHARE_MARCA":
        return "Meta Marcas"
    if tipo == "MECANICO_FATURAMENTO":
        return "Meta Mecânicos"
    return tipo or "-"


def _tipo_badge(tipo: str) -> str:
    tipo = normalize_text(tipo)
    if tipo == "CRESCIMENTO":
        return "📈 Crescimento"
    if tipo == "MIX":
        return "🧩 Mix"
    if tipo == "SHARE_MARCA":
        return "🏷️ Marcas"
    if tipo == "MECANICO_FATURAMENTO":
        return "🔧 Mecânicos"
    return tipo or "-"


def _escopo_label(escopo: str) -> str:
    escopo_n = normalize_text(escopo) or "VENDEDOR"
    if escopo_n == "GERENTE":
        return "Gerente / EMP inteira"
    if escopo_n == "MECANICO":
        return "Mecânico / Oficina"
    return "Vendedor"


def _escopo_badge(escopo: str) -> str:
    escopo_n = normalize_text(escopo) or "VENDEDOR"
    if escopo_n == "GERENTE":
        return "🏢 Gerente"
    if escopo_n == "MECANICO":
        return "🔧 Mecânico"
    return "👤 Vendedor"


def _parse_escalas(raw: str, tipo: str) -> list[tuple[float, float]]:
    """Aceita linhas como 5=0,10 ou 750=100."""
    out: list[tuple[float, float]] = []
    for line in str(raw or "").splitlines():
        ln = line.strip()
        if not ln:
            continue
        # remove textos comuns que o usuario possa colar da planilha
        ln = ln.replace("%", "").replace("R$", "")
        if "=" in ln:
            left, right = ln.split("=", 1)
        elif ":" in ln:
            left, right = ln.split(":", 1)
        elif ";" in ln:
            left, right = ln.split(";", 1)
        else:
            parts = re.split(r"\s+", ln)
            if len(parts) < 2:
                continue
            left, right = parts[0], parts[1]
        lim = _safe_float(left, None)
        val = _safe_float(right, None)
        if lim is None or val is None:
            continue
        out.append((float(lim), float(val)))
    # remove duplicadas por limite, preservando maior valor informado por ultimo
    merged: dict[float, float] = {}
    for lim, val in out:
        merged[lim] = val
    return sorted(merged.items(), key=lambda x: x[0])


def _parse_marcas(raw: str) -> list[str]:
    parts = re.split(r"[,;\n/]+", str(raw or ""))
    marcas = []
    seen = set()
    for p in parts:
        marca = normalize_text(p)
        if marca and marca not in seen:
            seen.add(marca)
            marcas.append(marca)
    return marcas


def _period_from_request():
    hoje = date.today()
    ano = _safe_int(request.values.get("ano"), hoje.year)
    mes = _safe_int(request.values.get("mes"), hoje.month)
    if mes < 1 or mes > 12:
        mes = hoje.month
    return ano, mes


def _allowed_admin_emps(db) -> list[str]:
    role = (_role() or "").lower()
    allowed = _allowed_emps()
    if role == "admin":
        return get_active_emps(db, [])
    return get_active_emps(db, allowed)


def _meta_payload(db, meta: MetaPrograma) -> dict:
    emps = get_meta_emps(db, int(meta.id))
    escalas = get_meta_escalas(db, int(meta.id))
    marcas = get_meta_marcas(db, int(meta.id))
    bases_count = db.query(func.count(MetaBaseManual.id)).filter(MetaBaseManual.meta_id == meta.id).scalar() or 0
    return {
        "meta": meta,
        "emps": emps,
        "escalas": escalas,
        "marcas": marcas,
        "bases_count": int(bases_count or 0),
    }


def _admin_guard():
    red = _login_required()
    if red:
        return red
    if (_role() or "").lower() != "admin":
        flash("Acesso restrito ao administrador.", "warning")
        return redirect(url_for("dashboard"))
    return None


def metas():
    ensure_metas_lojas_schema()
    red = _login_required()
    if red:
        return red

    ano, mes = _period_from_request()
    role = (_role() or "").lower()
    usuario = normalize_text(session.get("usuario"))

    emp_filtro = normalize_emp(request.args.get("emp"))
    vendedor_filtro = normalize_text(request.args.get("vendedor"))

    with SessionLocal() as db:
        allowed_emps = _allowed_emps()
        emps_choices = _allowed_admin_emps(db)

        if role in ("vendedor", "mecanico"):
            vendedor_filtro = usuario
            # Se o participante tem EMPs vinculadas, usa elas; senao usa EMP da sessao.
            emps_user = [normalize_emp(e) for e in (allowed_emps or []) if normalize_emp(e)]
            if not emps_user and _emp():
                emps_user = [normalize_emp(_emp())]
            if emp_filtro and emp_filtro in set(emps_user):
                emps_calc = [emp_filtro]
            else:
                emps_calc = emps_user
        elif role in ("supervisor", "gerente"):
            allowed = set([normalize_emp(e) for e in (allowed_emps or []) if normalize_emp(e)])
            emps_calc = [emp_filtro] if emp_filtro and emp_filtro in allowed else list(allowed)
        else:
            emps_calc = [emp_filtro] if emp_filtro else []

        vendedores_choices = []
        if role != "vendedor":
            vendedores_choices = get_pessoas_para_metas(db, ano, mes, emps_calc or emps_choices)
            if vendedor_filtro and vendedor_filtro not in vendedores_choices:
                vendedor_filtro = ""

        metas_list, resultados = montar_resultados_periodo(
            db,
            ano,
            mes,
            emps=emps_calc,
            vendedor=vendedor_filtro,
            persist=True,
        )
        db.commit()

        totais = {
            "venda": round(sum(float(r.get("valor_mes") or 0.0) for r in resultados), 2),
            "premio": round(sum(float(r.get("total_premios") or 0.0) for r in resultados), 2),
            "vendedores": len(resultados),
            "metas": len(metas_list),
        }

        return render_template(
            "metas.html",
            role=role,
            emp=_emp(),
            ano=ano,
            mes=mes,
            emp_filtro=emp_filtro,
            vendedor_filtro=vendedor_filtro,
            emps_choices=emps_choices,
            vendedores_choices=vendedores_choices,
            metas_list=metas_list,
            resultados=resultados,
            totais=totais,
            tipo_label={"CRESCIMENTO": "Crescimento", "MIX": "Mix", "SHARE_MARCA": "Marcas", "MECANICO_FATURAMENTO": "Mecânicos"},
            tipo_label_fn=_tipo_label,
        )


def admin_metas():
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red

    ano, mes = _period_from_request()
    emp_selected = normalize_emp(request.args.get("emp_sel"))
    vendedor_selected = normalize_text(request.args.get("vendedor"))

    with SessionLocal() as db:
        emps_rows = db.query(Emp).filter(Emp.ativo.is_(True)).order_by(Emp.codigo.asc()).all()
        emps_codes = [normalize_emp(e.codigo) for e in emps_rows]
        if emp_selected and emp_selected not in set(emps_codes):
            emp_selected = ""

        metas_db = (
            db.query(MetaPrograma)
            .filter(MetaPrograma.ano == ano, MetaPrograma.mes == mes)
            .order_by(MetaPrograma.tipo.asc(), MetaPrograma.nome.asc(), MetaPrograma.id.asc())
            .all()
        )
        metas_payload = [_meta_payload(db, m) for m in metas_db]

        crescimento_meta = next((p["meta"] for p in metas_payload if normalize_text(p["meta"].tipo) == "CRESCIMENTO" and normalize_text(getattr(p["meta"], "escopo", "VENDEDOR")) not in ("GERENTE", "MECANICO") and p["meta"].ativo), None)
        crescimento_gerente_meta = next((p["meta"] for p in metas_payload if normalize_text(p["meta"].tipo) == "CRESCIMENTO" and normalize_text(getattr(p["meta"], "escopo", "VENDEDOR")) == "GERENTE" and p["meta"].ativo), None)
        mecanico_meta = next((p["meta"] for p in metas_payload if normalize_text(p["meta"].tipo) == "MECANICO_FATURAMENTO" and p["meta"].ativo), None)
        margem_minima_ativa = float(getattr(crescimento_meta, "margem_minima", 0.0) or 0.0) if crescimento_meta else 0.0
        margem_minima_gerente_ativa = float(getattr(crescimento_gerente_meta, "margem_minima", 0.0) or 0.0) if crescimento_gerente_meta else 0.0
        base_rows = []
        gerente_base_rows = []
        mecanico_rows = []
        vendedores_base = []
        if crescimento_meta and emp_selected:
            vendedores_base = get_vendedores_para_metas(db, ano, mes, [emp_selected])
            if vendedor_selected:
                vendedores_base = [v for v in vendedores_base if v == vendedor_selected]
            for vend in vendedores_base:
                base = (
                    db.query(MetaBaseManual)
                    .filter(MetaBaseManual.meta_id == crescimento_meta.id, MetaBaseManual.emp == emp_selected, MetaBaseManual.vendedor == vend)
                    .first()
                )
                venda_mes = query_valor_mes(db, ano, mes, emp_selected, vend)
                base_valor = float(getattr(base, "base_valor", 0.0) or 0.0) if base else 0.0
                crescimento_pct = ((venda_mes - base_valor) / base_valor * 100.0) if base_valor > 0 else 0.0
                margem_individual = float(getattr(base, "margem_percentual", 0.0) or 0.0) if base else 0.0
                margem_efetiva, margem_origem = _margem_minima_efetiva(db, crescimento_meta.id, emp_selected, vend, margem_minima_ativa)
                base_rows.append({
                    "emp": emp_selected,
                    "vendedor": vend,
                    "venda_mes": venda_mes,
                    "base_valor": base_valor,
                    "crescimento_pct": crescimento_pct,
                    "margem_minima_individual": margem_individual,
                    "margem_minima_efetiva": margem_efetiva,
                    "margem_minima_origem": margem_origem,
                    "observacao": getattr(base, "observacao", "") if base else "",
                })

        if crescimento_gerente_meta and emp_selected:
            gerentes_base = get_gerentes_para_metas(db, ano, mes, [emp_selected])
            if vendedor_selected:
                gerentes_base = [g for g in gerentes_base if g == vendedor_selected]
            for gerente in gerentes_base:
                base = (
                    db.query(MetaBaseManual)
                    .filter(MetaBaseManual.meta_id == crescimento_gerente_meta.id, MetaBaseManual.emp == emp_selected, MetaBaseManual.vendedor == gerente)
                    .first()
                )
                venda_mes = query_valor_mes(db, ano, mes, emp_selected, None)
                base_valor = float(getattr(base, "base_valor", 0.0) or 0.0) if base else 0.0
                crescimento_pct = ((venda_mes - base_valor) / base_valor * 100.0) if base_valor > 0 else 0.0
                margem_individual = float(getattr(base, "margem_percentual", 0.0) or 0.0) if base else 0.0
                margem_efetiva, margem_origem = _margem_minima_efetiva(db, crescimento_gerente_meta.id, emp_selected, gerente, margem_minima_gerente_ativa)
                gerente_base_rows.append({
                    "emp": emp_selected,
                    "vendedor": gerente,
                    "venda_mes": venda_mes,
                    "base_valor": base_valor,
                    "crescimento_pct": crescimento_pct,
                    "margem_minima_individual": margem_individual,
                    "margem_minima_efetiva": margem_efetiva,
                    "margem_minima_origem": margem_origem,
                    "observacao": getattr(base, "observacao", "") if base else "",
                })

        if mecanico_meta and emp_selected:
            mecanicos_base = get_mecanicos_para_metas(db, ano, mes, [emp_selected])
            if vendedor_selected:
                mecanicos_base = [m for m in mecanicos_base if m == vendedor_selected]
            for mecanico in mecanicos_base:
                base = (
                    db.query(MetaBaseManual)
                    .filter(MetaBaseManual.meta_id == mecanico_meta.id, MetaBaseManual.emp == emp_selected, MetaBaseManual.vendedor == mecanico)
                    .first()
                )
                calc = calcular_meta(db, mecanico_meta, emp_selected, mecanico, persist=False)
                mecanico_rows.append({
                    "emp": emp_selected,
                    "mecanico": mecanico,
                    "vendedor": mecanico,
                    "faturamento": float(getattr(calc, "valor_mes", 0.0) or 0.0),
                    "faixa_limite": float(getattr(calc, "faixa_limite", 0.0) or 0.0) if getattr(calc, "faixa_limite", None) is not None else None,
                    "bonus_percentual": float(getattr(calc, "bonus_percentual", 0.0) or 0.0),
                    "premio": float(getattr(calc, "premio", 0.0) or 0.0),
                    "observacao": getattr(base, "observacao", "") if base else "",
                })

        sim_emps = [emp_selected] if emp_selected else []
        sim_vendedor = vendedor_selected if vendedor_selected else None
        metas_sim, resultados_sim = montar_resultados_periodo(db, ano, mes, emps=sim_emps, vendedor=sim_vendedor, persist=False)

        # Margem atual é individual por vendedor, não por EMP.
        # Quando uma EMP estiver filtrada, mostramos apenas os vendedores daquela EMP,
        # mas a busca da margem continua usando ANO + MÊS + VENDEDOR.
        vendedores_margem_filtro: list[str] = []
        if vendedor_selected:
            vendedores_margem_filtro = [vendedor_selected]
        elif emp_selected:
            vendedores_margem_filtro = [normalize_text(v) for v in get_vendedores_para_metas(db, ano, mes, [emp_selected])]

        margens_q = db.query(MetaMargemVendedor).filter(MetaMargemVendedor.ano == ano, MetaMargemVendedor.mes == mes)
        if vendedores_margem_filtro:
            margens_q = margens_q.filter(MetaMargemVendedor.vendedor.in_(vendedores_margem_filtro))
        margens_rows_db_all = (
            margens_q
            .order_by(MetaMargemVendedor.vendedor.asc(), MetaMargemVendedor.importado_em.desc(), MetaMargemVendedor.id.desc())
            .limit(1200)
            .all()
        )

        # Pode existir histórico legado por EMP. Para a nova regra, mantém só a última margem por vendedor.
        margens_rows_db = []
        vistos_vendedores = set()
        for item in margens_rows_db_all:
            vend_key = normalize_text(getattr(item, "vendedor", ""))
            if not vend_key or vend_key in vistos_vendedores:
                continue
            vistos_vendedores.add(vend_key)
            margens_rows_db.append(item)
            if len(margens_rows_db) >= 600:
                break

        margens_rows = []
        crescimento_meta_id = int(crescimento_meta.id) if crescimento_meta else None
        for r in margens_rows_db:
            vend_norm = normalize_text(r.vendedor)
            emp_para_regra = emp_selected or normalize_emp(getattr(r, "emp", ""))
            margem_efetiva, margem_origem = _margem_minima_efetiva(
                db,
                crescimento_meta_id,
                emp_para_regra,
                vend_norm,
                margem_minima_ativa,
            )
            margem_atual = float(getattr(r, "margem_percentual", 0.0) or 0.0)
            margens_rows.append({
                "emp": emp_para_regra if emp_para_regra and emp_para_regra != "GERAL" else "Geral",
                "vendedor": vend_norm,
                "margem_percentual": margem_atual,
                "margem_minima_efetiva": margem_efetiva,
                "margem_minima_origem": margem_origem,
                "margem_atingida": True if margem_efetiva <= 0 else margem_atual >= margem_efetiva,
                "importado_em": getattr(r, "importado_em", None),
                "observacao": getattr(r, "observacao", "") or "",
            })

        margens_pendentes = []
        if emp_selected:
            margem_vendedores = {normalize_text(r.get("vendedor")) for r in margens_rows}
            for vend in get_vendedores_para_metas(db, ano, mes, [emp_selected]):
                vend_norm = normalize_text(vend)
                if vend_norm not in margem_vendedores:
                    margens_pendentes.append({"emp": emp_selected, "vendedor": vend_norm})

        totais = {
            "metas": len(metas_db),
            "ativas": sum(1 for m in metas_db if m.ativo),
            "bases": sum(int(p["bases_count"] or 0) for p in metas_payload),
            "premio_simulado": round(sum(float(r.get("total_premios") or 0.0) for r in resultados_sim), 2),
            "vendedores_simulados": len(resultados_sim),
            "margens": len(margens_rows),
            "margens_pendentes": len(margens_pendentes),
        }

        return render_template(
            "admin_metas.html",
            role=_role(),
            emp=_emp(),
            ano=ano,
            mes=mes,
            emp_selected=emp_selected,
            vendedor_selected=vendedor_selected,
            emps_rows=emps_rows,
            metas_payload=metas_payload,
            crescimento_meta=crescimento_meta,
            crescimento_gerente_meta=crescimento_gerente_meta,
            mecanico_meta=mecanico_meta,
            base_rows=base_rows,
            gerente_base_rows=gerente_base_rows,
            mecanico_rows=mecanico_rows,
            vendedores_choices=get_pessoas_para_metas(db, ano, mes, [emp_selected] if emp_selected else emps_codes),
            metas_sim=metas_sim,
            resultados_sim=resultados_sim,
            margens_rows=margens_rows,
            margem_minima_ativa=margem_minima_ativa,
            margens_pendentes=margens_pendentes,
            totais=totais,
            tipo_label=_tipo_label,
            tipo_badge=_tipo_badge,
            escopo_label=_escopo_label,
            escopo_badge=_escopo_badge,
        )


def admin_metas_criar():
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red

    ano, mes = _period_from_request()
    tipo = normalize_text(request.form.get("tipo"))
    escopo = normalize_text(request.form.get("escopo") or "VENDEDOR")
    if tipo == "MECANICO_FATURAMENTO":
        escopo = "MECANICO"
    if escopo not in ("VENDEDOR", "GERENTE", "MECANICO"):
        escopo = "VENDEDOR"
    nome = (request.form.get("nome") or _tipo_label(tipo)).strip()
    emps = [normalize_emp(e) for e in request.form.getlist("emps") if normalize_emp(e)]
    escalas = _parse_escalas(request.form.get("escalas"), tipo)
    marcas = _parse_marcas(request.form.get("marcas"))
    teto_faturamento = _safe_float(request.form.get("teto_faturamento"), 0.0)
    margem_minima = _safe_float(request.form.get("margem_minima"), 0.0)
    faturamento_default = 0.0 if tipo == "MECANICO_FATURAMENTO" or escopo == "MECANICO" else 70000.0
    faturamento_minimo = _safe_float(request.form.get("faturamento_minimo"), faturamento_default)

    if tipo not in TIPOS_META:
        flash("Tipo de meta inválido.", "danger")
        return redirect(url_for("admin_metas", ano=ano, mes=mes))
    if not emps:
        flash("Selecione pelo menos uma EMP para a meta.", "warning")
        return redirect(url_for("admin_metas", ano=ano, mes=mes))
    if not escalas:
        flash("Cadastre pelo menos uma faixa válida. Exemplo: 5=0,10 ou 750=100.", "warning")
        return redirect(url_for("admin_metas", ano=ano, mes=mes))
    if escopo == "GERENTE" and tipo not in ("CRESCIMENTO", "SHARE_MARCA"):
        flash("Meta de gerente aceita apenas Crescimento e Marcas.", "warning")
        return redirect(url_for("admin_metas", ano=ano, mes=mes))
    if escopo == "MECANICO" and tipo != "MECANICO_FATURAMENTO":
        flash("Meta de mecânico aceita apenas Faturamento de Oficina.", "warning")
        return redirect(url_for("admin_metas", ano=ano, mes=mes))
    if tipo == "SHARE_MARCA" and not marcas:
        flash("Informe pelo menos uma marca para a Meta Marcas.", "warning")
        return redirect(url_for("admin_metas", ano=ano, mes=mes))

    with SessionLocal() as db:
        meta = MetaPrograma(
            nome=nome,
            tipo=tipo,
            ano=ano,
            mes=mes,
            ativo=True,
            escopo=escopo,
            faturamento_minimo=faturamento_minimo if faturamento_minimo > 0 else 0.0,
            margem_minima=margem_minima if tipo == "CRESCIMENTO" and margem_minima > 0 else 0.0,
            # No crescimento, este campo representa a trava: venda >= teto aplica maior faixa.
            teto_faturamento=teto_faturamento if tipo == "CRESCIMENTO" and teto_faturamento > 0 else None,
            teto_bonus_percentual=None,
            created_by_user_id=session.get("user_id"),
        )
        db.add(meta)
        db.flush()

        for emp_codigo in sorted(set(emps)):
            db.add(MetaProgramaEmp(meta_id=meta.id, emp=emp_codigo, ativo=True))
        for idx, (lim, valor) in enumerate(escalas, start=1):
            db.add(MetaEscala(meta_id=meta.id, ordem=idx, limite_min=float(lim), bonus_percentual=float(valor)))
        for marca in marcas:
            db.add(MetaMarca(meta_id=meta.id, marca=marca))
        db.commit()

    flash(f"Meta '{nome}' criada com sucesso para {_escopo_label(escopo).lower()}.", "success")
    return redirect(url_for("admin_metas", ano=ano, mes=mes, emp_sel=emps[0] if emps else ""))


def admin_metas_emps_salvar(meta_id: int):
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red

    ano, mes = _period_from_request()
    emp_selected = normalize_emp(request.form.get("emp_sel"))
    vendedor_selected = normalize_text(request.form.get("vendedor"))
    novas_emps = sorted(set(normalize_emp(e) for e in request.form.getlist("emps") if normalize_emp(e)))

    if not novas_emps:
        flash("Selecione pelo menos uma EMP para a meta.", "warning")
        return redirect(url_for("admin_metas", ano=ano, mes=mes, emp_sel=emp_selected, vendedor=vendedor_selected))

    with SessionLocal() as db:
        meta = db.query(MetaPrograma).filter(MetaPrograma.id == int(meta_id)).first()
        if not meta:
            flash("Meta não encontrada.", "danger")
            return redirect(url_for("admin_metas", ano=ano, mes=mes, emp_sel=emp_selected, vendedor=vendedor_selected))
        if not bool(meta.ativo):
            flash("Ative a meta antes de alterar as EMPs participantes.", "warning")
            return redirect(url_for("admin_metas", ano=meta.ano, mes=meta.mes, emp_sel=emp_selected, vendedor=vendedor_selected))

        existentes = {normalize_emp(r.emp): r for r in db.query(MetaProgramaEmp).filter(MetaProgramaEmp.meta_id == meta.id).all()}
        agora = datetime.utcnow()
        adicionadas = 0
        removidas = 0

        for emp_codigo in novas_emps:
            vinc = existentes.get(emp_codigo)
            if vinc:
                if not bool(getattr(vinc, "ativo", True)):
                    adicionadas += 1
                vinc.ativo = True
                vinc.removido_em = None
                vinc.atualizado_em = agora
                db.add(vinc)
            else:
                db.add(MetaProgramaEmp(meta_id=meta.id, emp=emp_codigo, ativo=True, criado_em=agora, atualizado_em=agora))
                adicionadas += 1

        novas_set = set(novas_emps)
        for emp_codigo, vinc in existentes.items():
            if emp_codigo and emp_codigo not in novas_set and bool(getattr(vinc, "ativo", True)):
                vinc.ativo = False
                vinc.removido_em = agora
                vinc.atualizado_em = agora
                db.add(vinc)
                removidas += 1

        db.commit()

    flash(f"EMPs da meta atualizadas. Adicionadas: {adicionadas}. Removidas: {removidas}.", "success")
    return redirect(url_for("admin_metas", ano=ano, mes=mes, emp_sel=emp_selected or novas_emps[0], vendedor=vendedor_selected))


def admin_metas_regra_salvar(meta_id: int):
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red

    ano, mes = _period_from_request()
    emp_selected = normalize_emp(request.form.get("emp_sel"))
    vendedor_selected = normalize_text(request.form.get("vendedor"))
    faturamento_minimo = _safe_float(request.form.get("faturamento_minimo"), 0.0)
    margem_minima = _safe_float(request.form.get("margem_minima"), 0.0)
    teto_faturamento = _safe_float(request.form.get("teto_faturamento"), 0.0)
    escalas_raw = request.form.get("escalas")

    with SessionLocal() as db:
        meta = db.query(MetaPrograma).filter(MetaPrograma.id == int(meta_id)).first()
        if not meta:
            flash("Meta não encontrada.", "danger")
            return redirect(url_for("admin_metas", ano=ano, mes=mes, emp_sel=emp_selected, vendedor=vendedor_selected))

        tipo_meta = normalize_text(meta.tipo)
        if escalas_raw is not None:
            escalas = _parse_escalas(escalas_raw, tipo_meta)
            if not escalas:
                flash("Informe pelo menos uma faixa válida.", "warning")
                return redirect(url_for("admin_metas", ano=ano, mes=mes, emp_sel=emp_selected, vendedor=vendedor_selected))
            db.query(MetaEscala).filter(MetaEscala.meta_id == int(meta.id)).delete(synchronize_session=False)
            for idx, (lim, valor) in enumerate(escalas, start=1):
                db.add(MetaEscala(meta_id=meta.id, ordem=idx, limite_min=float(lim), bonus_percentual=float(valor)))

        meta.faturamento_minimo = faturamento_minimo if faturamento_minimo > 0 else 0.0
        if tipo_meta == "CRESCIMENTO":
            meta.margem_minima = margem_minima if margem_minima > 0 else 0.0
            meta.teto_faturamento = teto_faturamento if teto_faturamento > 0 else None
        else:
            meta.margem_minima = 0.0
        db.add(meta)
        db.commit()

    flash("Regras da meta atualizadas.", "success")
    return redirect(url_for("admin_metas", ano=ano, mes=mes, emp_sel=emp_selected, vendedor=vendedor_selected))



def admin_metas_mecanico_adicionar():
    """Compatibilidade para versões antigas da tela.

    O cadastro oficial de mecânicos foi centralizado em Admin > Usuários para
    manter todos os perfis na mesma página.
    """
    red = _admin_guard()
    if red:
        return red
    flash("Cadastre ou atualize mecânicos em Admin > Usuários, selecionando o perfil Mecânico e a EMP vinculada.", "info")
    return redirect(url_for("admin_usuarios"))


def admin_metas_margens_modelo():
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red

    ano, mes = _period_from_request()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Modelo Margens"
        headers = ["ANO", "MES", "VENDEDOR", "MARGEM_PERCENTUAL"]
        ws.append(headers)
        exemplos = [
            [ano, mes, "CARLOS_SALDANHA", 8.35],
            [ano, mes, "JOAO_MASO", 7.80],
            [ano, mes, "VENDEDOR_EXEMPLO", -1.20],
        ]
        for row in exemplos:
            ws.append(row)
        header_fill = PatternFill("solid", fgColor="FF8A00")
        thin = Side(style="thin", color="333333")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1E1200")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        widths = [10, 10, 34, 24]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"
        ws["F1"] = "Instruções"
        ws["F1"].font = Font(bold=True)
        ws["F2"] = "A margem é individual por vendedor. A última margem importada substitui a anterior para ANO+MES+VENDEDOR."
        ws["F3"] = "Informe a margem como percentual. Ex.: 8,35 representa 8,35%."
        ws["F4"] = "Não informe EMP. O sistema identifica pelo usuário/vendedor cadastrado, que é único."
        ws["F5"] = "Não soma, não acumula e não faz média. Só a última margem importada vale."
        ws["F6"] = "VENDEDOR deve bater com o usuário/vendedor do sistema."
        ws.column_dimensions["F"].width = 78
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name=f"modelo_importacao_margens_{mes:02d}_{ano}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        flash(f"Não foi possível gerar o modelo de margens: {exc}", "danger")
        return redirect(url_for("admin_metas", ano=ano, mes=mes))


def admin_metas_margens_importar():
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red

    ano_req, mes_req = _period_from_request()
    arquivo = request.files.get("arquivo_margens")
    if not arquivo or not arquivo.filename:
        flash("Selecione uma planilha de margens para importar.", "warning")
        return redirect(url_for("admin_metas", ano=ano_req, mes=mes_req))

    filename = arquivo.filename or "margens"
    try:
        if filename.lower().endswith(".csv"):
            df = pd.read_csv(arquivo, dtype=str, sep=None, engine="python")
        else:
            df = pd.read_excel(arquivo, dtype=str)
    except Exception as exc:
        flash(f"Não foi possível ler a planilha de margens: {exc}", "danger")
        return redirect(url_for("admin_metas", ano=ano_req, mes=mes_req))

    col_map = {_norm_col_name(c): c for c in df.columns}
    def col(*names):
        for name in names:
            key = _norm_col_name(name)
            if key in col_map:
                return col_map[key]
        return None

    c_ano = col("ANO")
    c_mes = col("MES")
    c_vendedor = col("VENDEDOR", "USUARIO", "LOGIN")
    c_margem = col("MARGEM_PERCENTUAL", "MARGEM", "MARGEM%", "PERCENTUAL_MARGEM")

    if not c_ano or not c_mes or not c_vendedor or not c_margem:
        flash("Planilha inválida. Colunas obrigatórias: ANO, MES, VENDEDOR e MARGEM_PERCENTUAL.", "danger")
        return redirect(url_for("admin_metas", ano=ano_req, mes=mes_req))

    upserts: dict[tuple[int, int, str], dict] = {}
    erros: list[str] = []
    vendedores_planilha: set[str] = set()
    for idx, row in df.iterrows():
        try:
            ano = _safe_int(row.get(c_ano), 0)
            mes = _safe_int(row.get(c_mes), 0)
            vendedor = normalize_text(row.get(c_vendedor))
            margem = _parse_margem_percentual(row.get(c_margem), None)
            if not (1 <= int(mes) <= 12) or int(ano) <= 2000:
                raise ValueError("competência inválida")
            if not vendedor:
                raise ValueError("VENDEDOR vazio")
            if margem is None:
                raise ValueError("margem inválida")
            vendedores_planilha.add(vendedor)
            # Se houver duplicidade do mesmo vendedor na própria planilha, a última linha vence.
            upserts[(int(ano), int(mes), vendedor)] = {
                "margem": float(margem),
            }
        except Exception as exc:
            if len(erros) < 10:
                erros.append(f"Linha {idx + 2}: {exc}")

    if not upserts:
        flash("Nenhuma margem válida foi encontrada na planilha. " + (" | ".join(erros) if erros else ""), "danger")
        return redirect(url_for("admin_metas", ano=ano_req, mes=mes_req))

    importados = 0
    atualizados = 0
    agora = datetime.utcnow()
    usuario = normalize_text(session.get("usuario"))
    vendedores_nao_encontrados: list[str] = []
    emp_global = "GERAL"

    with SessionLocal() as db:
        def _compact_user_key(value: object) -> str:
            return re.sub(r"[^A-Z0-9]", "", normalize_text(value))

        usuarios_lookup: dict[str, str] = {}
        for u in db.query(Usuario).filter(Usuario.username.isnot(None)).all():
            username = normalize_text(u.username)
            if not username:
                continue
            usuarios_lookup.setdefault(username, username)
            usuarios_lookup.setdefault(_compact_user_key(username), username)

        def _resolve_vendedor_importado(vendedor_raw: str) -> str | None:
            vendedor_raw = normalize_text(vendedor_raw)
            return usuarios_lookup.get(vendedor_raw) or usuarios_lookup.get(_compact_user_key(vendedor_raw))

        for vendedor in sorted(vendedores_planilha):
            if not _resolve_vendedor_importado(vendedor):
                vendedores_nao_encontrados.append(vendedor)

        for (ano, mes, vendedor_importado), payload in upserts.items():
            vendedor = _resolve_vendedor_importado(vendedor_importado)
            if not vendedor:
                continue

            # Nova regra: uma única margem por vendedor/competência, independente da EMP.
            # Mantém emp="GERAL" por compatibilidade com a tabela atual.
            item = (
                db.query(MetaMargemVendedor)
                .filter(
                    MetaMargemVendedor.ano == ano,
                    MetaMargemVendedor.mes == mes,
                    MetaMargemVendedor.emp == emp_global,
                    MetaMargemVendedor.vendedor == vendedor,
                )
                .first()
            )
            if item:
                atualizados += 1
            else:
                item = MetaMargemVendedor(ano=ano, mes=mes, emp=emp_global, vendedor=vendedor)
                importados += 1
            item.margem_percentual = float(payload["margem"])
            item.observacao = "Margem individual por vendedor"
            item.arquivo_origem = filename[:255]
            item.importado_por = usuario
            item.importado_em = agora
            db.add(item)
        db.commit()

    msg = f"Margens importadas com sucesso. Novas: {importados}. Atualizadas: {atualizados}."
    if vendedores_nao_encontrados:
        msg += " Vendedores não encontrados/ignorados: " + ", ".join(vendedores_nao_encontrados[:10])
        if len(vendedores_nao_encontrados) > 10:
            msg += f" e mais {len(vendedores_nao_encontrados) - 10}."
    if erros:
        msg += " Algumas linhas foram ignoradas: " + " | ".join(erros)
    flash(msg, "warning" if vendedores_nao_encontrados or erros else "success")
    # Volta para a competência do filtro; se a planilha tinha outra competência, o usuário pode filtrar depois.
    return redirect(url_for("admin_metas", ano=ano_req, mes=mes_req))


def admin_metas_toggle(meta_id: int):
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red
    ano, mes = _period_from_request()
    with SessionLocal() as db:
        meta = db.query(MetaPrograma).filter(MetaPrograma.id == int(meta_id)).first()
        if not meta:
            flash("Meta não encontrada.", "danger")
            return redirect(url_for("admin_metas", ano=ano, mes=mes))
        meta.ativo = not bool(meta.ativo)
        db.commit()
    flash("Status da meta atualizado.", "success")
    return redirect(url_for("admin_metas", ano=ano, mes=mes))


def admin_metas_excluir(meta_id: int):
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red
    ano, mes = _period_from_request()
    with SessionLocal() as db:
        meta = db.query(MetaPrograma).filter(MetaPrograma.id == int(meta_id)).first()
        if not meta:
            flash("Meta não encontrada.", "danger")
            return redirect(url_for("admin_metas", ano=ano, mes=mes))
        db.query(MetaBaseManual).filter(MetaBaseManual.meta_id == meta.id).delete(synchronize_session=False)
        db.query(MetaEscala).filter(MetaEscala.meta_id == meta.id).delete(synchronize_session=False)
        db.query(MetaMarca).filter(MetaMarca.meta_id == meta.id).delete(synchronize_session=False)
        db.query(MetaProgramaEmp).filter(MetaProgramaEmp.meta_id == meta.id).delete(synchronize_session=False)
        # metas_resultados pode nao estar importada aqui; delete via SQL simples para evitar dependencias.
        db.execute(text("DELETE FROM metas_resultados WHERE meta_id = :mid"), {"mid": meta.id})
        db.delete(meta)
        db.commit()
    flash("Meta excluída com sucesso.", "success")
    return redirect(url_for("admin_metas", ano=ano, mes=mes))


def admin_metas_recalcular():
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red
    ano, mes = _period_from_request()
    emp_selected = normalize_emp(request.form.get("emp_sel"))
    vendedor_selected = normalize_text(request.form.get("vendedor"))

    with SessionLocal() as db:
        emps = [emp_selected] if emp_selected else []
        _, resultados = montar_resultados_periodo(db, ano, mes, emps=emps, vendedor=vendedor_selected or None, persist=True)
        db.commit()

    flash(f"Metas recalculadas: {len(resultados)} participante(s)/EMP.", "success")
    return redirect(url_for("admin_metas", ano=ano, mes=mes, emp_sel=emp_selected, vendedor=vendedor_selected))


def admin_meta_bases(meta_id: int):
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red

    with SessionLocal() as db:
        meta = db.query(MetaPrograma).filter(MetaPrograma.id == int(meta_id)).first()
        if not meta:
            flash("Meta não encontrada.", "danger")
            return redirect(url_for("admin_metas"))
        meta_mecanico = is_meta_mecanico(meta)
        if normalize_text(meta.tipo) != "CRESCIMENTO" and not meta_mecanico:
            flash("Esta tela é usada para bases de crescimento ou cadastro de mecânicos.", "warning")
            return redirect(url_for("admin_metas", ano=meta.ano, mes=meta.mes))

        emps = get_meta_emps(db, int(meta.id))
        emp_selected = normalize_emp(request.args.get("emp")) or (emps[0] if emps else "")
        if emp_selected and emp_selected not in set(emps):
            emp_selected = emps[0] if emps else ""

        meta_gerente = is_meta_gerente(meta)
        if meta_mecanico:
            vendedores = get_mecanicos_para_metas(db, meta.ano, meta.mes, [emp_selected] if emp_selected else emps)
        else:
            vendedores = (
                get_gerentes_para_metas(db, meta.ano, meta.mes, [emp_selected] if emp_selected else emps)
                if meta_gerente
                else get_vendedores_para_metas(db, meta.ano, meta.mes, [emp_selected] if emp_selected else emps)
            )
        linhas = []
        for vend in vendedores:
            base = (
                db.query(MetaBaseManual)
                .filter(MetaBaseManual.meta_id == meta.id, MetaBaseManual.emp == emp_selected, MetaBaseManual.vendedor == vend)
                .first()
            )
            if meta_mecanico:
                calc = calcular_meta(db, meta, emp_selected, vend, persist=False)
                linhas.append({
                    "emp": emp_selected,
                    "vendedor": vend,
                    "venda_mes": float(getattr(calc, "valor_mes", 0.0) or 0.0),
                    "base_valor": 0.0,
                    "crescimento_pct": 0.0,
                    "margem_minima_individual": 0.0,
                    "margem_minima_efetiva": 0.0,
                    "margem_minima_origem": "",
                    "faixa_limite": float(getattr(calc, "faixa_limite", 0.0) or 0.0) if getattr(calc, "faixa_limite", None) is not None else None,
                    "bonus_percentual": float(getattr(calc, "bonus_percentual", 0.0) or 0.0),
                    "premio": float(getattr(calc, "premio", 0.0) or 0.0),
                    "observacao": getattr(base, "observacao", "") if base else "",
                })
                continue

            venda_mes = query_valor_mes(db, meta.ano, meta.mes, emp_selected, None if meta_gerente else vend)
            base_valor = float(getattr(base, "base_valor", 0.0) or 0.0) if base else 0.0
            crescimento_pct = ((venda_mes - base_valor) / base_valor * 100.0) if base_valor > 0 else 0.0
            margem_individual = float(getattr(base, "margem_percentual", 0.0) or 0.0) if base else 0.0
            margem_padrao = float(getattr(meta, "margem_minima", 0.0) or 0.0)
            margem_efetiva = margem_individual if margem_individual > 0 else margem_padrao
            margem_origem = "individual" if margem_individual > 0 else ("padrao" if margem_padrao > 0 else "")
            linhas.append({
                "emp": emp_selected,
                "vendedor": vend,
                "venda_mes": venda_mes,
                "base_valor": base_valor,
                "crescimento_pct": crescimento_pct,
                "margem_minima_individual": margem_individual,
                "margem_minima_efetiva": margem_efetiva,
                "margem_minima_origem": margem_origem,
                "observacao": getattr(base, "observacao", "") if base else "",
            })

        return render_template(
            "admin_meta_bases.html",
            role=_role(),
            emp=_emp(),
            meta=meta,
            emps=emps,
            emp_selected=emp_selected,
            linhas=linhas,
            margem_minima_padrao=float(getattr(meta, "margem_minima", 0.0) or 0.0),
            meta_gerente=meta_gerente,
            meta_mecanico=meta_mecanico,
            participante_label="Mecânico" if meta_mecanico else ("Gerente" if meta_gerente else "Vendedor"),
            tipo_label=_tipo_badge,
        )


def admin_meta_bases_salvar(meta_id: int):
    ensure_metas_lojas_schema()
    red = _admin_guard()
    if red:
        return red

    with SessionLocal() as db:
        meta = db.query(MetaPrograma).filter(MetaPrograma.id == int(meta_id)).first()
        if not meta:
            flash("Meta não encontrada.", "danger")
            return redirect(url_for("admin_metas"))
        emp_selected = normalize_emp(request.form.get("emp_selected"))
        if not emp_selected:
            flash("EMP não informada.", "warning")
            return redirect(url_for("admin_meta_bases", meta_id=meta.id))

        salvos = 0
        for key, value in request.form.items():
            if not key.startswith("base__"):
                continue
            vendedor = normalize_text(key.replace("base__", "", 1))
            base_valor = _safe_float(value, 0.0)
            margem_individual = _safe_float(request.form.get(f"margem__{vendedor}"), 0.0)
            obs = request.form.get(f"obs__{vendedor}") or ""
            upsert_base_manual(db, meta.id, emp_selected, vendedor, base_valor, obs, margem_individual)
            salvos += 1
        db.commit()

    flash(f"Bases salvas com sucesso: {salvos} participante(s).", "success")
    return redirect(url_for("admin_meta_bases", meta_id=meta_id, emp=emp_selected))
