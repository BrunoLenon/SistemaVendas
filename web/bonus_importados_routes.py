# -*- coding: utf-8 -*-
"""Importação e consulta do snapshot mensal da aba ``BONUS FINAL``.

A planilha continua responsável por todas as regras e cálculos. O SistemaVendas
apenas lê as colunas autorizadas para cada função, grava os valores prontos e
aplica a hierarquia de visualização:

* vendedor: somente os próprios dados;
* mecânico: somente os próprios dados;
* gerente: os próprios dados e os vendedores/mecânicos das EMPs vinculadas;
* administrador: todos os registros e a importação da competência.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from flask import current_app, flash, redirect, render_template, request, session, url_for
from sqlalchemy import or_
from werkzeug.utils import secure_filename

from auth_helpers import _allowed_emps, _login_required, _role, _usuario_logado
from db import (
    BonusImportacaoLote,
    BonusUsuarioImportado,
    SessionLocal,
    Usuario,
    UsuarioEmp,
    ensure_bonus_importados_schema,
    ensure_itens_parados_snapshot_schema,
    ensure_bonus_outros_valores_schema,
)
from security_utils import audit, normalize_role
from sv_utils import business_today, emp_sort_key
from itens_parados_snapshot import attach_saldos_to_bonus_rows
from bonus_outros_valores import attach_outros_valores_to_bonus_rows, bonus_row_options


MAX_UPLOAD_BYTES = 40 * 1024 * 1024
VALID_EXTENSIONS = {".xlsx", ".xlsm"}
SUPPORTED_FUNCTIONS = {"VENDEDOR", "GERENTE", "MECANICO"}

# Mapeamento por NOME do cabeçalho da aba BONUS FINAL.
#
# A posição física das colunas pode mudar sem quebrar a importação. O sistema
# normaliza acentos, espaços e pontuação e procura cada métrica pelos aliases
# abaixo. ``valor_parcial`` permanece opcional para compatibilidade com bases
# antigas; quando a coluna não existir, o valor é gravado como zero.
COLUMN_LAYOUT: dict[str, dict[str, Any]] = {
    "funcao": {
        "headers": {"FUNCAO", "CARGO", "PERFIL"},
        "label": "FUNÇÃO",
        "required": True,
    },
    "emp": {
        "headers": {"EMP", "EMPRESA", "LOJA"},
        "label": "EMP",
        "required": True,
    },
    "usuario_nome": {
        "headers": {"FUNCIONARIO", "USUARIO", "VENDEDOR", "COLABORADOR"},
        "label": "FUNCIONARIO",
        "required": True,
    },
    "produto_vendedor": {
        "headers": {"PRODUTOVENDEDOR", "PRODUTOSVENDEDOR", "BONUSPRODUTOVENDEDOR"},
        "label": "PRODUTO VENDEDOR",
        "required": True,
    },
    "produto_gerente": {
        "headers": {"PRODUTOGERENTE", "PRODUTOSGERENTE", "BONUSPRODUTOGERENTE"},
        "label": "PRODUTO GERENTE",
        "required": True,
    },
    "mecanico_faturado": {
        "headers": {"MECANICO", "SERVICO", "FATURAMENTOMECANICO", "VALORFATURADOMECANICO"},
        "label": "SERVIÇO / MECÂNICO",
        "required": True,
    },
    "venda_anterior": {
        "headers": {"VENDAANTERIOR", "VALORINDIVIDUALANTERIOR", "FATURAMENTOANTERIOR"},
        "label": "VENDA ANTERIOR",
        "required": True,
    },
    "venda_atual": {
        "headers": {"VENDAATUAL", "VALORINDIVIDUALATUAL", "FATURAMENTOATUAL"},
        "label": "VENDA ATUAL",
        "required": True,
    },
    "crescimento": {
        "headers": {"CRESCIMENTO", "CRESCIMENTOVENDEDOR", "PERCENTUALCRESCIMENTO"},
        "label": "CRESCIMENTO",
        "required": True,
    },
    "loja_anterior": {
        "headers": {"LOJAANTERIOR", "VENDAANTERIORLOJA", "FATURAMENTOLOJAANTERIOR"},
        "label": "LOJA ANTERIOR",
        "required": True,
    },
    "loja_atual": {
        "headers": {"LOJAATUAL", "VENDAATUALLOJA", "FATURAMENTOLOJAATUAL"},
        "label": "LOJA ATUAL",
        "required": True,
    },
    "importado_vendedor": {
        "headers": {"IMPORTADOVENDEDOR", "IMPORTADOINDIVIDUAL"},
        "label": "IMPORTADO VENDEDOR",
        "required": True,
    },
    "importado_loja": {
        "headers": {"IMPORTADOLOJA", "IMPORTADOEMPRESA"},
        "label": "IMPORTADO LOJA",
        "required": True,
    },
    "bonus_importado": {
        "headers": {"BONUSIMPORTADO", "PREMIOIMPORTADO"},
        "label": "BONUS IMPORTADO",
        "required": True,
    },
    "valor_meta": {
        "headers": {"VALORMETA", "PREMIOMETA"},
        "label": "VALOR META",
        "required": True,
    },
    "valor_parcial": {
        "headers": {"VALORPARCIAL", "BONUSPARCIAL", "PREMIOPROVISORIO", "VALORPROVISORIO"},
        "label": "VALOR PARCIAL",
        "required": False,
    },
    "bonus_final": {
        "headers": {"VALORFINAL", "BONUSFINAL", "PREMIOFINAL"},
        "label": "VALOR FINAL",
        "required": True,
    },
}


ROLE_FIELDS = {
    "VENDEDOR": (
        "produto_vendedor",
        "venda_anterior",
        "venda_atual",
        "crescimento",
        "importado_vendedor",
        "bonus_importado",
        "valor_meta",
        "valor_parcial",
        "bonus_final",
    ),
    "GERENTE": (
        # Alguns gerentes também realizam vendas próprias. A coluna E deve ser
        # preservada junto aos valores agregados da loja.
        "produto_vendedor",
        "loja_anterior",
        "loja_atual",
        "produto_gerente",
        "importado_loja",
        "bonus_importado",
        "valor_meta",
        "valor_parcial",
        "bonus_final",
    ),
    "MECANICO": (
        "mecanico_faturado",
        "valor_parcial",
        "bonus_final",
    ),
}

MONEY_FIELDS = {
    "produto_vendedor",
    "produto_gerente",
    "mecanico_faturado",
    "venda_anterior",
    "venda_atual",
    "loja_anterior",
    "loja_atual",
    "importado_vendedor",
    "importado_loja",
    "bonus_importado",
    "valor_meta",
    "valor_parcial",
    "bonus_final",
}
PERCENT_FIELDS = {"crescimento"}
# Campos que devem ser importados para toda linha válida, independentemente
# da função. Loja Anterior/Loja Atual são o retrato da EMP e Valor Final é o
# valor liberado pronto vindo da planilha. A função do usuário só controla as
# demais colunas auxiliares e a visualização; nunca altera o Valor Final.
GLOBAL_FIELDS = ("loja_anterior", "loja_atual", "bonus_final")
ALL_IMPORTED_FIELDS = MONEY_FIELDS | PERCENT_FIELDS
REQUIRED_CALCULATED_FIELDS = {"bonus_final"}
ROLE_SORT_ORDER = {"GERENTE": 0, "VENDEDOR": 1, "MECANICO": 2}


def register_bonus_importados_routes(app) -> None:
    # URL canônica do módulo de varejo. O endpoint é preservado para não quebrar
    # templates e redirecionamentos existentes.
    app.add_url_rule(
        "/bonus-varejo",
        endpoint="bonus_importados",
        view_func=bonus_importados,
        methods=["GET"],
    )

    # Compatibilidade com favoritos e links antigos.
    def _bonus_legacy_redirect():
        return redirect(url_for("bonus_importados", **request.args.to_dict(flat=True)))

    app.add_url_rule(
        "/bonus",
        endpoint="bonus_importados_legacy",
        view_func=_bonus_legacy_redirect,
        methods=["GET"],
    )
    app.add_url_rule(
        "/admin/bonus/importar",
        endpoint="admin_bonus_importar",
        view_func=admin_bonus_importar,
        methods=["POST"],
    )


def _strip_accents(value: object) -> str:
    raw = str(value or "")
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch)
    )


def _norm_header(value: object) -> str:
    raw = _strip_accents(value).strip().upper().replace("%", " PERCENTUAL ")
    return re.sub(r"[^A-Z0-9]", "", raw)


def _norm_username(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def _norm_function(value: object) -> str:
    raw = re.sub(r"\s+", " ", _strip_accents(value).strip()).upper()
    aliases = {
        "MEC": "MECANICO",
        "OFICINA": "MECANICO",
        "VENDEDORA": "VENDEDOR",
        "GERENCIA": "GERENTE",
        "MANAGER": "GERENTE",
    }
    return aliases.get(raw, raw)


def _norm_emp(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).upper()
    if isinstance(value, (int, float, Decimal)):
        try:
            number = Decimal(str(value))
            if number == number.to_integral_value():
                return str(int(number))
        except Exception:
            pass
    raw = str(value).strip()
    if re.fullmatch(r"[+-]?\d+[.,]0+", raw):
        return raw.split(".", 1)[0].split(",", 1)[0]
    return raw.upper()


def _decimal_from_cell(
    cell,
    *,
    field: str,
    is_percent: bool,
) -> Decimal | None:
    value = cell.value
    if getattr(cell, "data_type", None) == "e" or (
        isinstance(value, str) and value.strip().startswith("#")
    ):
        raise ValueError(f"erro do Excel ({value})")

    if value is None or (isinstance(value, str) and not value.strip()):
        if field in REQUIRED_CALCULATED_FIELDS:
            raise ValueError(
                "sem valor calculado; abra a planilha no Excel, recalcule e salve antes de importar"
            )
        return None if is_percent else Decimal("0")

    had_percent_symbol = False
    if isinstance(value, str):
        raw = value.strip()
        had_percent_symbol = "%" in raw
        raw = raw.replace("R$", "").replace("%", "").replace(" ", "")
        if not raw:
            return None if is_percent else Decimal("0")
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        elif raw.count(".") > 1:
            raw = raw.replace(".", "")
        value = raw

    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"valor numérico inválido ({value})") from exc
    if not result.is_finite():
        raise ValueError(f"valor numérico inválido ({value})")

    number_format = str(getattr(cell, "number_format", "") or "")
    if is_percent and not had_percent_symbol and "%" in number_format:
        # O Excel armazena 10% como 0,10. No banco persistimos 10,000000.
        result *= Decimal("100")

    quantum = Decimal("0.000001") if is_percent else Decimal("0.0001")
    return result.quantize(quantum, rounding=ROUND_HALF_UP)


FINAL_BONUS_SHEET_ALIASES = {"BONUSFINAL", "FINALBONUS"}


def _find_final_bonus_sheet(workbook):
    """Localiza a aba de origem sem correr o risco de ler uma cópia antiga.

    A planilha atual usa ``Final Bonus``. Bases antigas usavam ``BONUS FINAL``.
    Quando os dois nomes existem no mesmo arquivo, ``Final Bonus`` visível tem
    prioridade. Se houver mais de uma aba visível com o mesmo nome normalizado,
    a importação é interrompida em vez de escolher uma aba arbitrariamente.
    """

    matches = []
    for position, sheet in enumerate(workbook.worksheets):
        normalized = _norm_header(sheet.title)
        if normalized not in FINAL_BONUS_SHEET_ALIASES:
            continue
        visible = str(getattr(sheet, "sheet_state", "visible") or "visible").lower() == "visible"
        # A planilha atual (FINALBONUS) tem prioridade sobre o alias legado.
        name_rank = 0 if normalized == "FINALBONUS" else 1
        visibility_rank = 0 if visible else 1
        matches.append((name_rank, visibility_rank, position, sheet))

    if not matches:
        available = ", ".join(sheet.title for sheet in workbook.worksheets)
        raise ValueError(
            "A aba 'Final Bonus' (ou 'BONUS FINAL') não foi encontrada. "
            "Abas disponíveis: " + available
        )

    # Evita ler silenciosamente uma cópia antiga quando existem duas abas
    # visíveis chamadas, por exemplo, 'Final Bonus' e 'Final  Bonus'.
    visible_current = [
        item for item in matches if item[0] == 0 and item[1] == 0
    ]
    if len(visible_current) > 1:
        names = ", ".join(item[3].title for item in visible_current)
        raise ValueError(
            "Existem várias abas visíveis equivalentes a 'Final Bonus': " + names
        )

    matches.sort(key=lambda item: (item[0], item[1], item[2]))
    return matches[0][3]


def _excel_column_letter(zero_based_index: int | None) -> str:
    """Converte índice zero-based em letra de coluna para diagnóstico do admin."""
    if zero_based_index is None or zero_based_index < 0:
        return "?"
    number = int(zero_based_index) + 1
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _map_header_cells(cells: list[Any]) -> dict[str, int]:
    """Resolve campos pelo texto do cabeçalho, sem depender da letra da coluna."""

    positions: dict[str, list[int]] = defaultdict(list)
    for index, cell in enumerate(cells):
        normalized = _norm_header(getattr(cell, "value", None))
        if normalized:
            positions[normalized].append(index)

    mapping: dict[str, int] = {}
    for field, config in COLUMN_LAYOUT.items():
        matches: set[int] = set()
        for alias in config["headers"]:
            matches.update(positions.get(alias, []))
        if len(matches) > 1:
            raise ValueError(
                f"Cabeçalho ambíguo para '{config['label']}': existem colunas duplicadas/compatíveis na mesma linha."
            )
        if matches:
            mapping[field] = next(iter(matches))
    return mapping


def _find_header(sheet) -> tuple[int, dict[str, int]]:
    required_fields = {
        field for field, config in COLUMN_LAYOUT.items() if bool(config.get("required", True))
    }
    best_row = 1
    best_mapping: dict[str, int] = {}
    best_score = -1
    max_col = max(int(sheet.max_column or 1), 1)

    for row_number, cells_tuple in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 30), max_col=max_col),
        start=1,
    ):
        cells = list(cells_tuple)
        mapping = _map_header_cells(cells)
        score = len(required_fields.intersection(mapping))
        if score > best_score:
            best_row, best_mapping, best_score = row_number, mapping, score
        if required_fields.issubset(mapping):
            return row_number, mapping

    missing = [
        COLUMN_LAYOUT[field]["label"]
        for field in COLUMN_LAYOUT
        if field in required_fields and field not in best_mapping
    ]
    found = [COLUMN_LAYOUT[field]["label"] for field in best_mapping]
    detail = ", ".join(missing) if missing else "cabeçalho não identificado"
    found_detail = ", ".join(found) if found else "nenhum cabeçalho reconhecido"
    raise ValueError(
        "Não foi possível identificar a estrutura da aba BONUS FINAL pelos nomes das colunas. "
        f"Campos obrigatórios ausentes: {detail}. Reconhecidos: {found_detail}."
    )


def _read_bonus_workbook(content: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - depende do ambiente do deploy
        raise RuntimeError(
            "A biblioteca openpyxl não está disponível no servidor. Inclua openpyxl nas dependências."
        ) from exc

    try:
        workbook = load_workbook(
            filename=BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ValueError("Não foi possível abrir o arquivo Excel enviado.") from exc

    try:
        sheet = _find_final_bonus_sheet(workbook)
        header_row, mapping = _find_header(sheet)

        records: list[dict[str, Any]] = []
        warnings: list[str] = []
        duplicate_keys: dict[tuple[str, str], int] = {}
        rows_read = 0
        rows_skipped = 0

        for source_row, cells_tuple in enumerate(
            sheet.iter_rows(min_row=header_row + 1, max_col=max(mapping.values()) + 1), start=header_row + 1
        ):
            cells = list(cells_tuple)
            identity_values = [
                cells[mapping[field]].value for field in ("funcao", "emp", "usuario_nome")
            ]
            if not any(value not in (None, "") for value in identity_values):
                continue

            rows_read += 1
            username = _norm_username(cells[mapping["usuario_nome"]].value)
            function = _norm_function(cells[mapping["funcao"]].value)
            emp = _norm_emp(cells[mapping["emp"]].value)

            identity_missing = []
            if not username:
                identity_missing.append("usuário")
            if not function:
                identity_missing.append("função")
            if not emp:
                identity_missing.append("EMP")
            if identity_missing:
                rows_skipped += 1
                warnings.append(
                    f"Linha {source_row} ignorada: sem {', '.join(identity_missing)}."
                )
                continue

            if function not in SUPPORTED_FUNCTIONS:
                rows_skipped += 1
                warnings.append(
                    f"Linha {source_row} ignorada: função '{function}' não é VENDEDOR, GERENTE ou MECANICO."
                )
                continue

            record: dict[str, Any] = {
                "linha_origem": source_row,
                "usuario_nome": username,
                "funcao": function,
                "emp": emp,
            }
            for field in MONEY_FIELDS:
                record[field] = Decimal("0")
            for field in PERCENT_FIELDS:
                record[field] = None

            numeric_error = None

            # Loja Anterior, Loja Atual e principalmente Valor Final são lidos
            # diretamente pelos nomes dos cabeçalhos para TODAS as funções.
            # O Valor Final é o valor liberado pronto da planilha e nunca deve
            # depender de VENDEDO/GERENTE/MECANICO dentro do SistemaVendas.
            for field in GLOBAL_FIELDS:
                if field not in mapping:
                    numeric_error = (
                        f"campo obrigatório '{COLUMN_LAYOUT[field]['label']}' não encontrado"
                    )
                    break
                try:
                    record[field] = _decimal_from_cell(
                        cells[mapping[field]],
                        field=field,
                        is_percent=False,
                    )
                except ValueError as exc:
                    numeric_error = f"campo '{COLUMN_LAYOUT[field]['label']}': {exc}"
                    break

            if numeric_error is None:
                for field in ROLE_FIELDS[function]:
                    # Os campos globais já foram lidos acima pelo cabeçalho.
                    if field in GLOBAL_FIELDS:
                        continue
                    # Alguns layouts mais novos deixaram de trazer a provisão/valor
                    # parcial. Campos explicitamente opcionais são zerados sem
                    # impedir a importação da competência.
                    if field not in mapping:
                        if not COLUMN_LAYOUT[field].get("required", True):
                            record[field] = None if field in PERCENT_FIELDS else Decimal("0")
                            continue
                        numeric_error = f"campo obrigatório '{COLUMN_LAYOUT[field]['label']}' não encontrado"
                        break
                    try:
                        record[field] = _decimal_from_cell(
                            cells[mapping[field]],
                            field=field,
                            is_percent=(field in PERCENT_FIELDS),
                        )
                    except ValueError as exc:
                        if field == "crescimento":
                            # Crescimento pode ficar indisponível quando a venda anterior é zero.
                            # Mantemos o usuário e exibimos "não disponível" em vez de inventar 0%.
                            record[field] = None
                            warnings.append(
                                f"Linha {source_row}: Crescimento não disponível para {username} ({exc})."
                            )
                            continue
                        numeric_error = f"campo '{COLUMN_LAYOUT[field]['label']}': {exc}"
                        break

            if numeric_error:
                rows_skipped += 1
                warnings.append(f"Linha {source_row} ignorada: {numeric_error}.")
                continue

            key = (emp, username)
            if key in duplicate_keys:
                first_row = duplicate_keys[key]
                raise ValueError(
                    f"Duplicidade na aba BONUS FINAL: usuário {username}, EMP {emp}, "
                    f"linhas {first_row} e {source_row}."
                )
            duplicate_keys[key] = source_row
            records.append(record)

        if not records:
            raise ValueError("Nenhuma linha válida foi encontrada na aba BONUS FINAL.")

        # Loja Anterior/Atual são valores agregados da EMP e normalmente se
        # repetem em todas as linhas da mesma loja. Divergências dentro da mesma
        # EMP são sinal de planilha parcialmente atualizada e precisam aparecer
        # para o administrador em vez de passarem despercebidas.
        loja_values: dict[str, dict[str, set[Decimal]]] = defaultdict(
            lambda: {"loja_anterior": set(), "loja_atual": set()}
        )
        for record in records:
            for field in ("loja_anterior", "loja_atual"):
                value = record.get(field)
                if value is not None:
                    loja_values[record["emp"]][field].add(Decimal(str(value)))
        for emp, values in loja_values.items():
            for field in ("loja_anterior", "loja_atual"):
                if len(values[field]) > 1:
                    formatted = ", ".join(str(value) for value in sorted(values[field]))
                    warnings.append(
                        f"EMP {emp}: '{COLUMN_LAYOUT[field]['label']}' possui mais de um valor na aba "
                        f"{sheet.title}: {formatted}."
                    )

        manager_records = [record for record in records if record["funcao"] == "GERENTE"]
        manager_final_nonzero = sum(
            1 for record in manager_records if Decimal(str(record.get("bonus_final") or 0)) != 0
        )

        return {
            "records": records,
            "warnings": warnings,
            "rows_read": rows_read,
            "rows_skipped": rows_skipped,
            "sheet_name": sheet.title,
            "column_mapping": dict(mapping),
            "manager_count": len(manager_records),
            "manager_final_nonzero": manager_final_nonzero,
        }
    finally:
        workbook.close()


def _role_from_function(function: str) -> str:
    return {
        "VENDEDOR": "vendedor",
        "GERENTE": "gerente",
        "MECANICO": "mecanico",
    }.get(function, function.lower())


def _bind_users_and_validate(db, records: list[dict[str, Any]], warnings: list[str]) -> None:
    users = db.query(Usuario).all()
    users_by_name = {_norm_username(user.username): user for user in users}
    links = db.query(UsuarioEmp).filter(UsuarioEmp.ativo.is_(True)).all()
    emps_by_user: dict[int, set[str]] = defaultdict(set)
    for link in links:
        emps_by_user[int(link.usuario_id)].add(_norm_emp(link.emp))

    for record in records:
        user = users_by_name.get(record["usuario_nome"])
        if user is None:
            record["usuario_id"] = None
            warnings.append(
                f"Linha {record['linha_origem']}: usuário {record['usuario_nome']} ainda não está cadastrado no sistema."
            )
            continue

        record["usuario_id"] = user.id
        current_role = normalize_role(getattr(user, "role", None))
        sheet_role = _role_from_function(record["funcao"])
        if current_role != sheet_role:
            warnings.append(
                f"Linha {record['linha_origem']}: função da planilha ({record['funcao']}) difere do cadastro ({current_role.upper()})."
            )

        allowed = set(emps_by_user.get(int(user.id), set()))
        legacy_emp = _norm_emp(getattr(user, "emp", None))
        if legacy_emp:
            allowed.add(legacy_emp)
        if allowed and record["emp"] not in allowed:
            warnings.append(
                f"Linha {record['linha_origem']}: EMP {record['emp']} não está vinculada ao usuário {record['usuario_nome']}."
            )


def _safe_period(value: object, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(str(value).strip())
    except Exception:
        return default
    return max(minimum, min(maximum, parsed))


def admin_bonus_importar():
    red = _login_required()
    if red:
        return red
    if _role() != "admin":
        flash("Acesso restrito ao administrador.", "warning")
        return redirect(url_for("bonus_importados"))

    today = business_today()
    ano = _safe_period(request.form.get("ano"), today.year, 2000, 2100)
    mes = _safe_period(request.form.get("mes"), today.month, 1, 12)
    uploaded = request.files.get("arquivo")

    if uploaded is None or not uploaded.filename:
        flash("Selecione a planilha que contém a aba BONUS FINAL.", "warning")
        return redirect(url_for("bonus_importados", ano=ano, mes=mes))

    filename = secure_filename(uploaded.filename) or "bonus_final.xlsx"
    extension = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in VALID_EXTENSIONS:
        flash("Envie um arquivo .xlsx ou .xlsm.", "warning")
        return redirect(url_for("bonus_importados", ano=ano, mes=mes))

    content = uploaded.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        flash("A planilha excede o limite de 40 MB.", "danger")
        return redirect(url_for("bonus_importados", ano=ano, mes=mes))
    if not content:
        flash("O arquivo enviado está vazio.", "warning")
        return redirect(url_for("bonus_importados", ano=ano, mes=mes))

    try:
        parsed = _read_bonus_workbook(content)
        ensure_bonus_importados_schema()

        username = _norm_username(_usuario_logado())
        user_id = session.get("user_id")
        imported_at = datetime.utcnow()
        warnings = list(parsed["warnings"])

        with SessionLocal() as db:
            _bind_users_and_validate(db, parsed["records"], warnings)

            # Substituição atômica: uma falha preserva a competência anterior.
            db.query(BonusUsuarioImportado).filter(
                BonusUsuarioImportado.ano == ano,
                BonusUsuarioImportado.mes == mes,
            ).delete(synchronize_session=False)

            batch = BonusImportacaoLote(
                ano=ano,
                mes=mes,
                arquivo_origem=filename,
                importado_por_user_id=int(user_id) if user_id else None,
                importado_por=username,
                importado_em=imported_at,
                linhas_lidas=int(parsed["rows_read"]),
                linhas_importadas=len(parsed["records"]),
                linhas_ignoradas=int(parsed["rows_skipped"]),
                avisos_json=json.dumps(warnings, ensure_ascii=False),
            )
            db.add(batch)
            db.flush()

            rows = []
            expected_by_key: dict[tuple[str, str], dict[str, Decimal]] = {}
            for record in parsed["records"]:
                payload = {
                    **record,
                    "lote_id": batch.id,
                    "ano": ano,
                    "mes": mes,
                    "importado_em": imported_at,
                }
                rows.append(BonusUsuarioImportado(**payload))
                expected_by_key[(record["emp"], record["usuario_nome"])] = {
                    "loja_anterior": Decimal(str(record.get("loja_anterior") or 0)),
                    "loja_atual": Decimal(str(record.get("loja_atual") or 0)),
                    "bonus_final": Decimal(str(record.get("bonus_final") or 0)),
                }
            db.add_all(rows)

            # Força INSERTs dentro da mesma transação e lê de volta os três
            # campos mais críticos. Se banco/ORM/trigger alterar qualquer valor,
            # abortamos tudo e a competência anterior permanece intacta.
            db.flush()
            persisted_rows = (
                db.query(BonusUsuarioImportado)
                .filter(
                    BonusUsuarioImportado.ano == ano,
                    BonusUsuarioImportado.mes == mes,
                    BonusUsuarioImportado.lote_id == batch.id,
                )
                .all()
            )
            persisted_by_key = {
                (row.emp, row.usuario_nome): row for row in persisted_rows
            }
            if len(persisted_by_key) != len(expected_by_key):
                raise ValueError(
                    "Validação pós-gravação falhou: quantidade de registros no banco difere da planilha."
                )
            for key, expected in expected_by_key.items():
                persisted = persisted_by_key.get(key)
                if persisted is None:
                    raise ValueError(
                        f"Validação pós-gravação falhou: registro {key[1]} / EMP {key[0]} não foi encontrado."
                    )
                for field in ("loja_anterior", "loja_atual", "bonus_final"):
                    actual = Decimal(str(getattr(persisted, field, 0) or 0)).quantize(Decimal("0.0001"))
                    wanted = Decimal(str(expected[field] or 0)).quantize(Decimal("0.0001"))
                    if actual != wanted:
                        raise ValueError(
                            f"Validação pós-gravação falhou em {key[1]} / EMP {key[0]}: "
                            f"{COLUMN_LAYOUT[field]['label']} esperado {wanted}, gravado {actual}."
                        )

            db.commit()

        audit(
            "bonus_final_snapshot_imported",
            ano=ano,
            mes=mes,
            arquivo=filename,
            linhas=len(parsed["records"]),
            ignoradas=parsed["rows_skipped"],
            avisos=len(warnings),
        )
        mapping = parsed.get("column_mapping") or {}
        loja_atual_col = _excel_column_letter(mapping.get("loja_atual"))
        valor_final_col = _excel_column_letter(mapping.get("bonus_final"))
        flash(
            f"{parsed.get('sheet_name', 'Final Bonus')} de {mes:02d}/{ano} importado: "
            f"{len(parsed['records'])} usuários. Linhas ignoradas: {parsed['rows_skipped']}. "
            f"Loja Atual={loja_atual_col}; Valor Final={valor_final_col}. "
            f"Gerentes com Valor Final diferente de zero: "
            f"{parsed.get('manager_final_nonzero', 0)}/{parsed.get('manager_count', 0)}.",
            "success",
        )
        if warnings:
            flash(
                f"A importação terminou com {len(warnings)} aviso(s). Consulte a validação da última importação.",
                "warning",
            )
    except Exception as exc:
        current_app.logger.exception("Falha ao importar aba BONUS FINAL")
        audit(
            "bonus_final_snapshot_import_failed",
            ano=ano,
            mes=mes,
            arquivo=filename,
            erro=str(exc),
        )
        flash(f"Não foi possível importar a planilha: {exc}", "danger")

    return redirect(url_for("bonus_importados", ano=ano, mes=mes))


def _sum_field(rows: list[BonusUsuarioImportado], field: str) -> Decimal:
    total = Decimal("0")
    for row in rows:
        value = getattr(row, field, None)
        if value is not None:
            total += Decimal(str(value))
    return total


def _period_options(db) -> list[tuple[int, int]]:
    rows = (
        db.query(BonusUsuarioImportado.ano, BonusUsuarioImportado.mes)
        .distinct()
        .order_by(BonusUsuarioImportado.ano.desc(), BonusUsuarioImportado.mes.desc())
        .all()
    )
    return [(int(row[0]), int(row[1])) for row in rows]


def _load_warnings(batch: BonusImportacaoLote | None) -> list[str]:
    if not batch or not batch.avisos_json:
        return []
    try:
        value = json.loads(batch.avisos_json)
        return [str(item) for item in value] if isinstance(value, list) else []
    except Exception:
        return []


def bonus_importados():
    red = _login_required()
    if red:
        return red

    try:
        ensure_bonus_importados_schema()
    except Exception:
        current_app.logger.exception("Falha ao garantir schema de bônus")
        flash(
            "As tabelas/colunas do módulo de bônus ainda não estão disponíveis. Execute o SQL de implantação no Supabase.",
            "danger",
        )
        return render_template(
            "bonus_importados.html",
            role=_role(),
            ano=business_today().year,
            mes=business_today().month,
            rows=[],
            own_rows=[],
            team_rows=[],
            seller_rows=[],
            manager_rows=[],
            mechanic_rows=[],
            selected=None,
            periods=[],
            emp_options=[],
            user_options=[],
            summary={},
            batch=None,
            batch_warnings=[],
            team_by_emp=[],
            admin_user_options=[],
            db_unavailable=True,
        )

    role = _role()
    current_username = _norm_username(_usuario_logado())
    today = business_today()

    with SessionLocal() as db:
        periods = _period_options(db)
        default_year, default_month = today.year, today.month
        ano = _safe_period(request.args.get("ano"), default_year, 2000, 2100)
        mes = _safe_period(request.args.get("mes"), default_month, 1, 12)

        period_query = db.query(BonusUsuarioImportado).filter(
            BonusUsuarioImportado.ano == ano,
            BonusUsuarioImportado.mes == mes,
        )

        all_period_rows = period_query.all() if role == "admin" else []
        emp_filter = _norm_emp(request.args.get("emp")) if role == "admin" else ""
        user_filter = _norm_username(request.args.get("usuario")) if role == "admin" else ""

        if role == "admin":
            query = period_query
            if emp_filter:
                query = query.filter(BonusUsuarioImportado.emp == emp_filter)
            if user_filter:
                query = query.filter(BonusUsuarioImportado.usuario_nome == user_filter)
            rows = query.all()
            emp_options = sorted(
                {row.emp for row in all_period_rows if row.emp}, key=emp_sort_key
            )
            user_options = sorted(
                {row.usuario_nome for row in all_period_rows if row.usuario_nome}
            )
        elif role == "gerente":
            allowed_emps = sorted(
                {_norm_emp(emp) for emp in _allowed_emps() if _norm_emp(emp)},
                key=emp_sort_key,
            )
            if allowed_emps:
                rows = (
                    period_query.filter(BonusUsuarioImportado.emp.in_(allowed_emps))
                    .filter(
                        or_(
                            BonusUsuarioImportado.usuario_nome == current_username,
                            BonusUsuarioImportado.funcao.in_(["VENDEDOR", "MECANICO"]),
                        )
                    )
                    .all()
                )
            else:
                rows = period_query.filter(
                    BonusUsuarioImportado.usuario_nome == current_username
                ).all()
            emp_options = []
            user_options = []
        else:
            # Vendedor e mecânico jamais recebem registros de outro usuário.
            rows = period_query.filter(
                BonusUsuarioImportado.usuario_nome == current_username
            ).all()
            emp_options = []
            user_options = []

        rows.sort(
            key=lambda row: (
                emp_sort_key(row.emp),
                ROLE_SORT_ORDER.get(row.funcao, 99),
                row.usuario_nome,
            )
        )
        try:
            ensure_itens_parados_snapshot_schema()
            itens_parados_summary = attach_saldos_to_bonus_rows(
                db, rows, ano=ano, mes=mes
            )
        except Exception:
            current_app.logger.exception(
                "Falha ao carregar saldo de itens parados no Bônus Varejo"
            )
            itens_parados_summary = {
                "total_visivel": Decimal("0"),
                "total_lojas": Decimal("0"),
            }
            for row in rows:
                row.saldo_itens_parados = Decimal("0")
                row.saldo_itens_parados_usuario = Decimal("0")
                row.saldo_itens_parados_loja = Decimal("0")

        try:
            ensure_bonus_outros_valores_schema()
            outros_valores_summary = attach_outros_valores_to_bonus_rows(
                db,
                rows,
                ano=ano,
                mes=mes,
                origem="VAREJO",
                bonus_field="bonus_final",
            )
        except Exception:
            current_app.logger.exception(
                "Falha ao carregar Outros Valores no Bônus Varejo"
            )
            outros_valores_summary = {
                "total_visivel": Decimal("0"),
                "quantidade": 0,
            }
            for row in rows:
                row.outros_valores_total = Decimal("0")
                row.outros_valores_detalhes = []
                row.valor_bonus_base = Decimal(str(row.bonus_final or 0))
                row.total_geral = row.valor_bonus_base + Decimal(
                    str(getattr(row, "saldo_itens_parados", 0) or 0)
                )

        seller_rows = [row for row in rows if row.funcao == "VENDEDOR"]
        manager_rows = [row for row in rows if row.funcao == "GERENTE"]
        mechanic_rows = [row for row in rows if row.funcao == "MECANICO"]
        own_rows = [row for row in rows if row.usuario_nome == current_username]
        team_rows = []
        if role == "gerente":
            team_rows = [
                row
                for row in rows
                if row.usuario_nome != current_username
                and row.funcao in {"VENDEDOR", "MECANICO"}
            ]

        selected_id = request.args.get("registro", type=int)
        detail_candidates = rows if role in {"admin", "gerente"} else own_rows
        selected = next(
            (row for row in detail_candidates if selected_id and row.id == selected_id),
            None,
        )
        if selected is None and detail_candidates:
            selected = next(
                (row for row in detail_candidates if row.usuario_nome == current_username),
                detail_candidates[0],
            )

        batch = (
            db.query(BonusImportacaoLote)
            .filter(BonusImportacaoLote.ano == ano, BonusImportacaoLote.mes == mes)
            .order_by(BonusImportacaoLote.importado_em.desc(), BonusImportacaoLote.id.desc())
            .first()
        )
        batch_warnings = _load_warnings(batch)

        team_by_emp_map: dict[str, dict[str, Any]] = {}
        for row in team_rows:
            item = team_by_emp_map.setdefault(
                row.emp,
                {
                    "emp": row.emp,
                    "quantidade": 0,
                    "valor_parcial": Decimal("0"),
                    "bonus_final": Decimal("0"),
                    "itens_parados": Decimal("0"),
                    "outros_valores": Decimal("0"),
                    "total_geral": Decimal("0"),
                },
            )
            item["quantidade"] += 1
            item["valor_parcial"] += Decimal(str(row.valor_parcial or 0))
            item["bonus_final"] += Decimal(str(row.bonus_final or 0))
            item["itens_parados"] += Decimal(
                str(getattr(row, "saldo_itens_parados_usuario", 0) or 0)
            )
            item["outros_valores"] += Decimal(
                str(getattr(row, "outros_valores_total", 0) or 0)
            )
            item["total_geral"] += (
                Decimal(str(row.bonus_final or 0))
                + Decimal(str(getattr(row, "saldo_itens_parados_usuario", 0) or 0))
                + Decimal(str(getattr(row, "outros_valores_total", 0) or 0))
            )
        team_by_emp = sorted(
            team_by_emp_map.values(), key=lambda item: emp_sort_key(item["emp"])
        )

        bonus_visivel = _sum_field(rows, "bonus_final")
        bonus_proprio = _sum_field(own_rows, "bonus_final")
        bonus_equipe = _sum_field(team_rows, "bonus_final")
        outros_proprio = sum(
            (Decimal(str(getattr(row, "outros_valores_total", 0) or 0)) for row in own_rows),
            Decimal("0"),
        )
        outros_equipe = sum(
            (Decimal(str(getattr(row, "outros_valores_total", 0) or 0)) for row in team_rows),
            Decimal("0"),
        )
        itens_proprio = sum(
            (Decimal(str(getattr(row, "saldo_itens_parados", 0) or 0)) for row in own_rows),
            Decimal("0"),
        )
        itens_equipe = sum(
            (Decimal(str(getattr(row, "saldo_itens_parados_usuario", 0) or 0)) for row in team_rows),
            Decimal("0"),
        )
        summary = {
            "registros": len(rows),
            "valor_parcial_visivel": _sum_field(rows, "valor_parcial"),
            "bonus_visivel": bonus_visivel,
            "bonus_vendedores": _sum_field(seller_rows, "bonus_final"),
            "bonus_gerentes": _sum_field(manager_rows, "bonus_final"),
            "bonus_mecanicos": _sum_field(mechanic_rows, "bonus_final"),
            "valor_parcial_proprio": _sum_field(own_rows, "valor_parcial"),
            "bonus_proprio": bonus_proprio,
            "valor_parcial_equipe": _sum_field(team_rows, "valor_parcial"),
            "bonus_equipe": bonus_equipe,
            "quantidade_equipe": len(team_rows),
            "itens_parados_visivel": itens_parados_summary["total_visivel"],
            "itens_parados_lojas": itens_parados_summary["total_lojas"],
            "outros_valores_visivel": outros_valores_summary["total_visivel"],
            "outros_valores_proprio": outros_proprio,
            "outros_valores_equipe": outros_equipe,
            "total_geral_visivel": (
                bonus_visivel
                + itens_parados_summary["total_visivel"]
                + outros_valores_summary["total_visivel"]
            ),
            "total_geral_proprio": bonus_proprio + itens_proprio + outros_proprio,
            "total_geral_equipe": bonus_equipe + itens_equipe + outros_equipe,
            "nao_vinculados": sum(1 for row in rows if row.usuario_id is None),
        }
        admin_user_options = bonus_row_options(all_period_rows) if role == "admin" else []

        return render_template(
            "bonus_importados.html",
            role=role,
            ano=ano,
            mes=mes,
            rows=rows,
            own_rows=own_rows,
            team_rows=team_rows,
            seller_rows=seller_rows,
            manager_rows=manager_rows,
            mechanic_rows=mechanic_rows,
            selected=selected,
            periods=periods,
            emp_options=emp_options,
            user_options=user_options,
            emp_filter=emp_filter,
            user_filter=user_filter,
            summary=summary,
            batch=batch,
            batch_warnings=batch_warnings,
            team_by_emp=team_by_emp,
            admin_user_options=admin_user_options,
            db_unavailable=False,
        )
