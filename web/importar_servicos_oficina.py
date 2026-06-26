from __future__ import annotations

"""Importação simples de mão de obra/oficina.

Formato recomendado:
  USUARIO | EMP | VALOR_SERVICO

A competência pode vir no formulário (mês de referência) ou na planilha por:
  ANO + MES, COMPETENCIA (YYYY-MM ou MM/YYYY) ou DATA.
"""

import datetime as dt
import os
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, Iterable

import pandas as pd
from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert as pg_insert

from db import OficinaServico, SessionLocal, engine

USUARIO_ALIASES = ["USUARIO", "MECANICO", "MECÂNICO", "VENDEDOR", "FUNCIONARIO", "FUNCIONÁRIO"]
VALOR_ALIASES = ["VALOR_SERVICO", "VALOR_SERVIÇO", "SERVICO", "SERVIÇO", "VALOR", "TOTAL"]
EMP_ALIASES = ["EMP", "EMPRESA", "LOJA"]
ANO_ALIASES = ["ANO"]
MES_ALIASES = ["MES", "MÊS"]
COMPETENCIA_ALIASES = ["COMPETENCIA", "COMPETÊNCIA", "REFERENCIA", "REFERÊNCIA"]
DATA_ALIASES = ["DATA", "MOVIMENTO"]
OBS_ALIASES = ["OBS", "OBSERVACAO", "OBSERVAÇÃO"]


def ensure_oficina_servicos_schema() -> None:
    """Garante a tabela antes de importar, mesmo se AUTO_MIGRATE estiver desligado."""
    with engine.connect() as conn:
        conn = conn.execution_options(isolation_level="AUTOCOMMIT")
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS oficina_servicos (
                id SERIAL PRIMARY KEY,
                ano INTEGER NOT NULL,
                mes INTEGER NOT NULL,
                emp VARCHAR(30) NOT NULL,
                usuario VARCHAR(80) NOT NULL,
                valor_servico DOUBLE PRECISION NOT NULL DEFAULT 0,
                observacao VARCHAR(240),
                arquivo_origem VARCHAR(255),
                importado_por VARCHAR(80),
                importado_em TIMESTAMP NOT NULL DEFAULT NOW(),
                ativo BOOLEAN NOT NULL DEFAULT TRUE,
                CONSTRAINT uq_oficina_servico_periodo_usuario UNIQUE (ano, mes, emp, usuario)
            );
        """))
        conn.execute(text("ALTER TABLE oficina_servicos ADD COLUMN IF NOT EXISTS observacao varchar(240);"))
        conn.execute(text("ALTER TABLE oficina_servicos ADD COLUMN IF NOT EXISTS arquivo_origem varchar(255);"))
        conn.execute(text("ALTER TABLE oficina_servicos ADD COLUMN IF NOT EXISTS importado_por varchar(80);"))
        conn.execute(text("ALTER TABLE oficina_servicos ADD COLUMN IF NOT EXISTS importado_em timestamp NOT NULL DEFAULT now();"))
        conn.execute(text("ALTER TABLE oficina_servicos ADD COLUMN IF NOT EXISTS ativo boolean NOT NULL DEFAULT true;"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ux_oficina_servicos_periodo_usuario ON oficina_servicos (ano, mes, emp, usuario);"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_oficina_servicos_emp_periodo ON oficina_servicos (emp, ano, mes);"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_oficina_servicos_usuario_periodo ON oficina_servicos (usuario, ano, mes);"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_oficina_servicos_periodo ON oficina_servicos (ano, mes);"))


def _norm_col(value: Any) -> str:
    import unicodedata
    s = str(value or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return s


def _norm_cols(cols: Iterable[Any]) -> list[str]:
    return [_norm_col(c) for c in cols]


def _pick(row: Any, aliases: list[str]) -> Any:
    norm_aliases = {_norm_col(a) for a in aliases}
    for a in norm_aliases:
        try:
            if a in row:
                return row.get(a)
        except Exception:
            pass
    return None


def _to_float(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value)
        s = str(value).strip()
        if not s:
            return None
        # Aceita R$ 1.234,56, 1234,56 e 1234.56
        s = s.replace("R$", "").replace(" ", "")
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return None


def _money(value: Any) -> float:
    try:
        return float(Decimal(str(value if value is not None else 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except Exception:
        return 0.0


def _norm_str(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _norm_usuario(value: Any) -> str:
    return _norm_str(value).upper()


def _parse_periodo(row: Any, ano_default: int | None, mes_default: int | None) -> tuple[int, int] | None:
    ano = _to_float(_pick(row, ANO_ALIASES))
    mes = _to_float(_pick(row, MES_ALIASES))
    if ano and mes:
        return int(ano), int(mes)

    comp = _norm_str(_pick(row, COMPETENCIA_ALIASES))
    if comp:
        s = comp.replace("/", "-").strip()
        parts = [p for p in s.split("-") if p]
        try:
            if len(parts) >= 2:
                if len(parts[0]) == 4:
                    return int(parts[0]), int(parts[1])
                return int(parts[1]), int(parts[0])
        except Exception:
            pass

    data = _pick(row, DATA_ALIASES)
    if data is not None and not (isinstance(data, float) and pd.isna(data)):
        try:
            d = pd.to_datetime(data).date()
            return int(d.year), int(d.month)
        except Exception:
            pass

    if ano_default and mes_default:
        return int(ano_default), int(mes_default)
    return None


def _find_missing(cols: list[str], ano_default: int | None, mes_default: int | None) -> list[str]:
    colset = set(cols)
    missing = []
    if not any(_norm_col(a) in colset for a in USUARIO_ALIASES):
        missing.append("USUARIO")
    if not any(_norm_col(a) in colset for a in EMP_ALIASES):
        missing.append("EMP")
    if not any(_norm_col(a) in colset for a in VALOR_ALIASES):
        missing.append("VALOR_SERVICO")
    has_periodo_cols = (
        any(_norm_col(a) in colset for a in COMPETENCIA_ALIASES)
        or any(_norm_col(a) in colset for a in DATA_ALIASES)
        or (any(_norm_col(a) in colset for a in ANO_ALIASES) and any(_norm_col(a) in colset for a in MES_ALIASES))
    )
    if not has_periodo_cols and not (ano_default and mes_default):
        missing.append("COMPETENCIA ou ANO+MES ou DATA")
    return missing


def _rows_from_xlsx(filepath: str):
    from openpyxl import load_workbook
    wb = None
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            yield None, None
            return
        cols = _norm_cols(header)
        for values in rows:
            yield cols, {cols[i]: values[i] if i < len(values) else None for i in range(len(cols))}
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def _build_stmt(records: list[dict], modo: str):
    stmt = pg_insert(OficinaServico.__table__).values(records)
    conflict_cols = ["ano", "mes", "emp", "usuario"]
    if modo == "somar":
        return stmt.on_conflict_do_update(
            index_elements=conflict_cols,
            set_={
                "valor_servico": OficinaServico.__table__.c.valor_servico + stmt.excluded.valor_servico,
                "observacao": stmt.excluded.observacao,
                "arquivo_origem": stmt.excluded.arquivo_origem,
                "importado_por": stmt.excluded.importado_por,
                "importado_em": stmt.excluded.importado_em,
                "ativo": True,
            },
        )
    return stmt.on_conflict_do_update(
        index_elements=conflict_cols,
        set_={
            "valor_servico": stmt.excluded.valor_servico,
            "observacao": stmt.excluded.observacao,
            "arquivo_origem": stmt.excluded.arquivo_origem,
            "importado_por": stmt.excluded.importado_por,
            "importado_em": stmt.excluded.importado_em,
            "ativo": True,
        },
    )


def importar_servicos_oficina(
    filepath: str,
    *,
    ano: int | None = None,
    mes: int | None = None,
    modo: str = "substituir",
    arquivo_origem: str | None = None,
    importado_por: str | None = None,
    batch_size: int = 500,
) -> Dict[str, Any]:
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in {".xlsx", ".csv"}:
        return {"ok": False, "msg": "Formato não suportado. Use .xlsx ou .csv."}

    modo = "somar" if str(modo or "").strip().lower() == "somar" else "substituir"
    ano_default = int(ano) if ano else None
    mes_default = int(mes) if mes else None

    total_linhas = validas = erros_linha = 0
    total_valor = 0.0
    periodos: set[tuple[str, int, int]] = set()
    aggregate: dict[tuple[int, int, str, str], dict[str, Any]] = {}
    missing_checked = False
    seen_cols: list[str] = []

    def consume_row(cols: list[str], row: Any):
        nonlocal total_linhas, validas, erros_linha, total_valor, missing_checked, seen_cols
        if not missing_checked:
            seen_cols = list(cols or [])
            missing = _find_missing(seen_cols, ano_default, mes_default)
            if missing:
                raise ValueError("Colunas faltando: " + ", ".join(missing))
            missing_checked = True

        total_linhas += 1
        usuario = _norm_usuario(_pick(row, USUARIO_ALIASES))
        emp = _norm_str(_pick(row, EMP_ALIASES))
        valor = _to_float(_pick(row, VALOR_ALIASES))
        periodo = _parse_periodo(row, ano_default, mes_default)
        if not usuario or not emp or valor is None or not periodo:
            erros_linha += 1
            return
        ano_r, mes_r = periodo
        if mes_r < 1 or mes_r > 12:
            erros_linha += 1
            return
        valor_m = _money(valor)
        key = (int(ano_r), int(mes_r), emp, usuario)
        if key not in aggregate:
            aggregate[key] = {
                "ano": int(ano_r),
                "mes": int(mes_r),
                "emp": emp,
                "usuario": usuario,
                "valor_servico": 0.0,
                "observacao": _norm_str(_pick(row, OBS_ALIASES))[:240] or None,
                "arquivo_origem": (arquivo_origem or "")[:255] or None,
                "importado_por": (importado_por or "")[:80] or None,
                "importado_em": dt.datetime.utcnow(),
                "ativo": True,
            }
        aggregate[key]["valor_servico"] = _money(float(aggregate[key]["valor_servico"] or 0.0) + valor_m)
        validas += 1
        total_valor = _money(total_valor + valor_m)
        periodos.add((emp, int(ano_r), int(mes_r)))

    try:
        if ext == ".xlsx":
            any_row = False
            for cols, row in _rows_from_xlsx(filepath):
                if cols is None:
                    return {"ok": False, "msg": "Planilha vazia."}
                any_row = True
                consume_row(cols, row)
            if not any_row:
                return {"ok": False, "msg": "Planilha vazia."}
        else:
            first = True
            for chunk in pd.read_csv(filepath, chunksize=2000, dtype=str, encoding_errors="ignore"):
                chunk.columns = _norm_cols(list(chunk.columns))
                if first and chunk.empty:
                    return {"ok": False, "msg": "Arquivo CSV vazio."}
                first = False
                cols = list(chunk.columns)
                for _, row in chunk.iterrows():
                    consume_row(cols, row)
    except ValueError as exc:
        return {"ok": False, "msg": str(exc), "faltando": str(exc).replace("Colunas faltando: ", "").split(", "), "lidas": seen_cols}
    except Exception as exc:
        return {"ok": False, "msg": f"Falha ao ler arquivo: {exc}"}

    records = list(aggregate.values())
    inseridas_ou_atualizadas = 0
    if records:
        ensure_oficina_servicos_schema()
        db = SessionLocal()
        try:
            for i in range(0, len(records), int(batch_size or 500)):
                batch = records[i:i + int(batch_size or 500)]
                stmt = _build_stmt(batch, modo)
                res = db.execute(stmt)
                db.commit()
                inseridas_ou_atualizadas += int(res.rowcount or 0)
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

    return {
        "ok": True,
        "msg": "Importação de mão de obra finalizada.",
        "total_linhas": int(total_linhas),
        "validas": int(validas),
        "erros_linha": int(erros_linha),
        "registros_competencia": int(len(records)),
        "inseridas_ou_atualizadas": int(inseridas_ou_atualizadas),
        "modo": modo,
        "total_servico": float(total_valor),
        "periodos": sorted(list(periodos)),
        "affected_periods": sorted(list(periodos)),
    }
